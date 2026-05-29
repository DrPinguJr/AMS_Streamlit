from __future__ import annotations

import base64
import os
import platform
import re
import subprocess
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


try:
    st.set_page_config(layout="wide", page_title="Recruitment Tracker")
except st.errors.StreamlitAPIException:
    pass


BASE_DIR = Path("Streamlit Version")
WORKBOOK_NAME = "Hiring Folder.xlsx"
WORKBOOK_PATH = BASE_DIR / WORKBOOK_NAME
RESUMES_DIR = BASE_DIR / "resumes"
JD_DIR = BASE_DIR / "jd"
LOGS_DIR = BASE_DIR / "logs"
EXPORTS_DIR = BASE_DIR / "exports"

ROLES_SHEET = "Roles"
CANDIDATES_SHEET = "Candidates"
ACTIVITY_LOG_SHEET = "ActivityLog"

ROLE_COLUMNS = [
    "RoleID",
    "RoleName",
    "JDFileName",
    "JDPath",
    "DateCreated",
    "LastUpdated",
    "Active",
]

CANDIDATE_COLUMNS = [
    "CandidateID",
    "Platform",
    "Name",
    "Contact",
    "RoleID",
    "Role",
    "Remarks",
    "Result",
    "ResumeFileName",
    "ResumePath",
    "DateAdded",
    "LastUpdated",
]

ACTIVITY_COLUMNS = [
    "Timestamp",
    "Action",
    "Role",
    "CandidateID",
    "Name",
    "Contact",
    "Details",
]

RESULT_OPTIONS = [
    "Pending",
    "Contacted",
    "Called",
    "Pass",
    "Fail",
    "KIV",
    "No Response",
    "Rejected",
    "Hired",
]

RESULT_COLOURS = {
    "Pass": "C6EFCE",
    "Hired": "70AD47",
    "Fail": "FFC7CE",
    "Rejected": "F4B183",
    "KIV": "FFF2CC",
    "No Response": "D9EAD3",
    "Contacted": "DDEBF7",
    "Called": "E4DFEC",
}

COMMON_RESUME_WORDS = {
    "resume",
    "resumes",
    "cv",
    "curriculum",
    "vitae",
    "profile",
    "candidate",
}


def today_str() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def timestamp_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def ensure_folder_structure() -> None:
    for folder in (BASE_DIR, RESUMES_DIR, JD_DIR, LOGS_DIR, EXPORTS_DIR):
        folder.mkdir(parents=True, exist_ok=True)


def ensure_workbook() -> None:
    ensure_folder_structure()
    if not WORKBOOK_PATH.exists():
        wb = Workbook()
        default = wb.active
        default.title = ROLES_SHEET
        for sheet_name in (CANDIDATES_SHEET, ACTIVITY_LOG_SHEET):
            wb.create_sheet(sheet_name)
        _write_headers(wb[ROLES_SHEET], ROLE_COLUMNS)
        _write_headers(wb[CANDIDATES_SHEET], CANDIDATE_COLUMNS)
        _write_headers(wb[ACTIVITY_LOG_SHEET], ACTIVITY_COLUMNS)
        wb.save(WORKBOOK_PATH)
        return

    wb = load_workbook(WORKBOOK_PATH)
    required = {
        ROLES_SHEET: ROLE_COLUMNS,
        CANDIDATES_SHEET: CANDIDATE_COLUMNS,
        ACTIVITY_LOG_SHEET: ACTIVITY_COLUMNS,
    }
    changed = False
    for sheet_name, columns in required.items():
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            _write_headers(ws, columns)
            changed = True
        else:
            changed = _ensure_headers(wb[sheet_name], columns) or changed
    if changed:
        wb.save(WORKBOOK_PATH)


def _write_headers(ws: Any, columns: list[str]) -> None:
    for col_idx, column in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=column)


def _ensure_headers(ws: Any, columns: list[str]) -> bool:
    existing = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    changed = False
    if not any(existing):
        _write_headers(ws, columns)
        return True
    for column in columns:
        if column not in existing:
            ws.cell(row=1, column=len(existing) + 1, value=column)
            existing.append(column)
            changed = True
    return changed


def _load_sheet(sheet_name: str, columns: list[str]) -> pd.DataFrame:
    ensure_workbook()
    try:
        df = pd.read_excel(WORKBOOK_PATH, sheet_name=sheet_name, dtype=str).fillna("")
    except ValueError:
        df = pd.DataFrame(columns=columns)
    for column in columns:
        if column not in df.columns:
            df[column] = ""
    return df[columns].astype(str)


def load_roles() -> pd.DataFrame:
    return _load_sheet(ROLES_SHEET, ROLE_COLUMNS)


def load_candidates() -> pd.DataFrame:
    df = _load_sheet(CANDIDATES_SHEET, CANDIDATE_COLUMNS)
    df["Result"] = df["Result"].replace("", "Pending")
    return df


def _write_df_to_sheet(sheet_name: str, df: pd.DataFrame, columns: list[str]) -> None:
    ensure_workbook()
    wb = load_workbook(WORKBOOK_PATH)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    ws.delete_rows(1, ws.max_row)
    _write_headers(ws, columns)
    safe_df = df.copy()
    for column in columns:
        if column not in safe_df.columns:
            safe_df[column] = ""
    for row in safe_df[columns].fillna("").astype(str).itertuples(index=False, name=None):
        ws.append(list(row))
    wb.save(WORKBOOK_PATH)


def save_roles(df: pd.DataFrame) -> None:
    _write_df_to_sheet(ROLES_SHEET, df, ROLE_COLUMNS)


def save_candidates(df: pd.DataFrame) -> None:
    _write_df_to_sheet(CANDIDATES_SHEET, df, CANDIDATE_COLUMNS)
    apply_result_colours()


def append_activity_log(
    action: str,
    role: str = "",
    candidate_id: str = "",
    name: str = "",
    contact: str = "",
    details: str = "",
) -> None:
    ensure_workbook()
    wb = load_workbook(WORKBOOK_PATH)
    if ACTIVITY_LOG_SHEET not in wb.sheetnames:
        wb.create_sheet(ACTIVITY_LOG_SHEET)
        _write_headers(wb[ACTIVITY_LOG_SHEET], ACTIVITY_COLUMNS)
    ws = wb[ACTIVITY_LOG_SHEET]
    _ensure_headers(ws, ACTIVITY_COLUMNS)
    ws.append([timestamp_str(), action, role, candidate_id, name, contact, details])
    wb.save(WORKBOOK_PATH)


def apply_result_colours() -> None:
    ensure_workbook()
    wb = load_workbook(WORKBOOK_PATH)
    if CANDIDATES_SHEET not in wb.sheetnames:
        wb.save(WORKBOOK_PATH)
        return

    ws = wb[CANDIDATES_SHEET]
    headers = [cell.value for cell in ws[1]]
    if "Result" not in headers:
        wb.save(WORKBOOK_PATH)
        return
    result_col = headers.index("Result") + 1

    for row in range(2, ws.max_row + 1):
        result = str(ws.cell(row=row, column=result_col).value or "")
        fill = PatternFill(fill_type=None)
        if result in RESULT_COLOURS:
            fill = PatternFill("solid", fgColor=RESULT_COLOURS[result])
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill
    wb.save(WORKBOOK_PATH)


def clean_contact(contact: Any) -> str:
    return re.sub(r"\D", "", str(contact or ""))


def safe_filename(text: Any) -> str:
    value = str(text or "").encode("ascii", "ignore").decode("ascii")
    value = re.sub(r"[<>:\"/\\|?*\x00-\x1f]", " ", value)
    value = re.sub(r"[^A-Za-z0-9._ -]", " ", value)
    value = re.sub(r"[\s_]+", "_", value).strip("._- ")
    return (value[:90].strip("._- ") or "File")


def extract_original_keyword(filename: str) -> str:
    stem = Path(filename or "").stem
    cleaned = safe_filename(stem)
    words = [
        word
        for word in re.split(r"[_\s.-]+", cleaned)
        if word and word.lower() not in COMMON_RESUME_WORDS
    ]
    keyword = "_".join(words)
    return safe_filename(keyword) if keyword else "Resume"


def generate_role_id(role_name: str) -> str:
    return safe_filename(role_name).lower()


def generate_candidate_id(role_id: str, contact: str) -> str:
    return f"{role_id}_{clean_contact(contact)}"


def _unique_path(folder: Path, filename: str) -> Path:
    path = folder / filename
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    counter = 2
    while True:
        candidate = folder / f"{stem}_{counter}{suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def _save_uploaded_file(uploaded_file: Any, folder: Path, filename: str) -> tuple[str, str]:
    ensure_folder_structure()
    final_path = _unique_path(folder, filename)
    final_path.write_bytes(uploaded_file.getbuffer())
    return final_path.name, final_path.relative_to(BASE_DIR).as_posix()


def save_uploaded_resume(uploaded_file: Any, name: str, contact: str, role: str) -> tuple[str, str]:
    ext = Path(uploaded_file.name).suffix.lower()
    keyword = extract_original_keyword(uploaded_file.name)
    filename = f"{safe_filename(name)}_{clean_contact(contact)}_{safe_filename(role)}_{keyword}_{today_str()}{ext}"
    return _save_uploaded_file(uploaded_file, RESUMES_DIR, filename)


def save_uploaded_jd(uploaded_file: Any, role: str) -> tuple[str, str]:
    ext = Path(uploaded_file.name).suffix.lower()
    filename = f"{safe_filename(role)}_JD_{today_str()}{ext}"
    return _save_uploaded_file(uploaded_file, JD_DIR, filename)


def resolve_storage_path(relative_or_absolute: str) -> Path:
    path = Path(str(relative_or_absolute or ""))
    if path.is_absolute():
        return path
    return BASE_DIR / path


def open_local_file(path: Path) -> tuple[bool, str]:
    try:
        if not path.exists():
            return False, f"File not found: {path}"
        system = platform.system()
        if system == "Windows":
            os.startfile(path)  # type: ignore[attr-defined]
        elif system == "Darwin":
            subprocess.run(["open", str(path)], check=False)
        else:
            subprocess.run(["xdg-open", str(path)], check=False)
        return True, f"Opened: {path}"
    except Exception as exc:
        return False, f"Could not open file: {exc}"


def preview_pdf(path: Path, height: int = 560) -> None:
    try:
        encoded = base64.b64encode(path.read_bytes()).decode("utf-8")
        st.markdown(
            f"""
            <iframe
                src="data:application/pdf;base64,{encoded}"
                width="100%"
                height="{height}"
                type="application/pdf">
            </iframe>
            """,
            unsafe_allow_html=True,
        )
    except Exception as exc:
        st.warning(f"PDF preview failed: {exc}")


def preview_docx_text(path: Path) -> None:
    try:
        import docx  # type: ignore
    except Exception:
        st.info("DOCX text preview is unavailable because python-docx is not installed.")
        return

    try:
        document = docx.Document(path)
        text = "\n".join(p.text for p in document.paragraphs if p.text.strip())
        if text:
            st.text_area("Text preview", text[:8000], height=360, disabled=True)
        else:
            st.info("No readable text found in this DOCX file.")
    except Exception as exc:
        st.warning(f"DOCX preview failed: {exc}")


def preview_image(path: Path) -> None:
    try:
        st.image(str(path), use_container_width=True)
    except Exception as exc:
        st.warning(f"Image preview failed: {exc}")


def render_file_viewer(
    label: str,
    relative_path: str,
    filename: str = "",
    key_prefix: str = "",
) -> None:
    st.subheader(label)
    if not relative_path:
        st.info(f"No {label.lower()} attached.")
        return

    path = resolve_storage_path(relative_path)
    st.caption(filename or path.name)
    st.code(str(path.resolve()))
    if not path.exists():
        st.warning("File is missing from the local folder.")
        return

    open_key = f"open_{key_prefix}_{label}_{relative_path}".replace("/", "_").replace("\\", "_")
    if st.button(f"Open {label}", key=open_key):
        ok, message = open_local_file(path)
        (st.success if ok else st.error)(message)

    suffix = path.suffix.lower()
    if suffix == ".pdf":
        preview_pdf(path)
    elif suffix == ".docx":
        preview_docx_text(path)
    elif suffix == ".doc":
        st.info("DOC preview is not available in Streamlit. Use the open button.")
    elif suffix in {".png", ".jpg", ".jpeg"}:
        preview_image(path)
    else:
        st.info("Preview is not available for this file type.")


def get_active_roles(roles_df: pd.DataFrame) -> pd.DataFrame:
    if roles_df.empty:
        return roles_df
    return roles_df[roles_df["Active"].str.lower().eq("yes")].copy()


def role_label(row: pd.Series) -> str:
    status = "Active" if str(row.get("Active", "")).lower() == "yes" else "Inactive"
    return f"{row.get('RoleName', '')} ({status})"


def find_role_by_id(roles_df: pd.DataFrame, role_id: str) -> pd.Series | None:
    matches = roles_df[roles_df["RoleID"] == role_id]
    if matches.empty:
        return None
    return matches.iloc[0]


def render_metrics(candidates_df: pd.DataFrame) -> None:
    cols = st.columns(6)
    metrics = [
        ("Total candidates", len(candidates_df)),
        ("Pending", (candidates_df["Result"] == "Pending").sum() if not candidates_df.empty else 0),
        ("Pass", (candidates_df["Result"] == "Pass").sum() if not candidates_df.empty else 0),
        ("Fail", (candidates_df["Result"] == "Fail").sum() if not candidates_df.empty else 0),
        ("KIV", (candidates_df["Result"] == "KIV").sum() if not candidates_df.empty else 0),
        ("Hired", (candidates_df["Result"] == "Hired").sum() if not candidates_df.empty else 0),
    ]
    for col, (label, value) in zip(cols, metrics):
        col.metric(label, int(value))


def filter_candidates(
    candidates_df: pd.DataFrame,
    role_filter: str,
    result_filter: str,
    platform_filter: str,
    search: str,
) -> pd.DataFrame:
    filtered = candidates_df.copy()
    if role_filter != "All":
        filtered = filtered[filtered["Role"] == role_filter]
    if result_filter != "All":
        filtered = filtered[filtered["Result"] == result_filter]
    if platform_filter != "All":
        filtered = filtered[filtered["Platform"] == platform_filter]
    search_value = search.strip().lower()
    if search_value:
        haystack = (
            filtered["Name"].str.lower()
            + " "
            + filtered["Contact"].str.lower()
            + " "
            + filtered["Remarks"].str.lower()
        )
        filtered = filtered[haystack.str.contains(re.escape(search_value), na=False)]
    return filtered.reset_index(drop=True)


def export_dataframe(df: pd.DataFrame, prefix: str, file_type: str) -> Path:
    ensure_folder_structure()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if file_type == "csv":
        path = EXPORTS_DIR / f"{prefix}_{stamp}.csv"
        df.to_csv(path, index=False, encoding="utf-8-sig")
    else:
        path = EXPORTS_DIR / f"{prefix}_{stamp}.xlsx"
        df.to_excel(path, index=False)
    return path


def format_candidate_for_email(candidate: pd.Series) -> str:
    remarks = str(candidate.get("Remarks", "")).strip() or "-"
    return "\n".join(
        [
            f"Platform: {candidate.get('Platform', '') or '-'}",
            f"Role: {candidate.get('Role', '') or '-'}",
            f"Name: {candidate.get('Name', '') or '-'}",
            f"Contact: {candidate.get('Contact', '') or '-'}",
            f"Remarks: {remarks}",
        ]
    )


def format_candidates_tab_copy(candidates: pd.DataFrame) -> str:
    if candidates.empty:
        return ""
    lines = []
    for _, candidate in candidates.iterrows():
        name = str(candidate.get("Name", "")).strip()
        contact = str(candidate.get("Contact", "")).strip()
        remarks = str(candidate.get("Remarks", "")).strip()
        lines.append(f"{name}\t{contact}\t{remarks}")
    return "\n".join(lines)


def render_role_management() -> None:
    roles_df = load_roles()
    st.header("Role Management")

    with st.expander("Create or update role", expanded=True):
        with st.form("role_form", clear_on_submit=False):
            role_name = st.text_input("Role Name")
            jd_file = st.file_uploader(
                "JD upload file",
                type=["pdf", "docx", "doc", "png"],
                accept_multiple_files=False,
            )
            update_existing = st.checkbox("Update existing role")
            submitted = st.form_submit_button("Save Role")

        if submitted:
            clean_role = role_name.strip()
            if not clean_role:
                st.error("Role Name is required.")
            else:
                role_id = generate_role_id(clean_role)
                existing_mask = roles_df["RoleID"] == role_id
                exists = bool(existing_mask.any())
                if exists and not update_existing:
                    st.warning("This role already exists. Tick Update existing role to update it.")
                else:
                    jd_filename = ""
                    jd_path = ""
                    if jd_file is not None:
                        jd_filename, jd_path = save_uploaded_jd(jd_file, clean_role)

                    now = today_str()
                    if exists:
                        idx = roles_df[existing_mask].index[0]
                        roles_df.loc[idx, "RoleName"] = clean_role
                        roles_df.loc[idx, "LastUpdated"] = now
                        roles_df.loc[idx, "Active"] = roles_df.loc[idx, "Active"] or "Yes"
                        if jd_file is not None:
                            roles_df.loc[idx, "JDFileName"] = jd_filename
                            roles_df.loc[idx, "JDPath"] = jd_path
                            append_activity_log("JD uploaded", clean_role, details=jd_path)
                        append_activity_log("role updated", clean_role, details=f"RoleID={role_id}")
                    else:
                        new_row = {
                            "RoleID": role_id,
                            "RoleName": clean_role,
                            "JDFileName": jd_filename,
                            "JDPath": jd_path,
                            "DateCreated": now,
                            "LastUpdated": now,
                            "Active": "Yes",
                        }
                        roles_df = pd.concat([roles_df, pd.DataFrame([new_row])], ignore_index=True)
                        append_activity_log("role created", clean_role, details=f"RoleID={role_id}")
                        if jd_file is not None:
                            append_activity_log("JD uploaded", clean_role, details=jd_path)
                    save_roles(roles_df)
                    st.success(f"Saved role: {clean_role}")
                    st.rerun()

    st.subheader("Existing roles")
    if roles_df.empty:
        st.info("No roles yet. Create your first role above.")
        return

    display_roles = roles_df.copy()
    edited_roles = st.data_editor(
        display_roles,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Active": st.column_config.SelectboxColumn("Active", options=["Yes", "No"]),
        },
        disabled=["RoleID", "JDFileName", "JDPath", "DateCreated", "LastUpdated"],
        key="roles_editor",
    )
    if st.button("Save Role Status Changes"):
        edited_roles["LastUpdated"] = edited_roles.apply(
            lambda row: today_str()
            if row["Active"] != roles_df.loc[roles_df["RoleID"] == row["RoleID"], "Active"].iloc[0]
            else row["LastUpdated"],
            axis=1,
        )
        save_roles(edited_roles)
        append_activity_log("role status updated", details="Role active/inactive changes saved")
        st.success("Role status changes saved.")
        st.rerun()

    role_options = {role_label(row): row["RoleID"] for _, row in roles_df.iterrows()}
    selected_label = st.selectbox("Select role to preview JD", [""] + list(role_options.keys()))
    if selected_label:
        selected_role = find_role_by_id(roles_df, role_options[selected_label])
        if selected_role is not None:
            render_file_viewer(
                "JD",
                selected_role.get("JDPath", ""),
                selected_role.get("JDFileName", ""),
                key_prefix="role_management",
            )


def render_add_candidate() -> None:
    st.header("Add Candidate")
    roles_df = load_roles()
    active_roles = get_active_roles(roles_df)
    candidates_df = load_candidates()

    if active_roles.empty:
        st.info("Create an active role before adding candidates.")
        return

    role_options = {row["RoleName"]: row["RoleID"] for _, row in active_roles.iterrows()}
    with st.form("add_candidate_form", clear_on_submit=False):
        role_name = st.selectbox("Role", list(role_options.keys()))
        platform_name = st.text_input("Platform")
        name = st.text_input("Name")
        contact = st.text_input("Contact")
        remarks = st.text_area("Remarks")
        resume_file = st.file_uploader("Resume uploader", type=["pdf", "docx", "doc", "png"])
        update_existing = st.checkbox("Update existing candidate instead")
        submitted = st.form_submit_button("Add Candidate")

    if not submitted:
        return

    cleaned_contact = clean_contact(contact)
    if not role_name or not name.strip() or not cleaned_contact:
        st.error("Role, Name, and Contact are required. Contact must contain at least one digit.")
        return

    role_id = role_options[role_name]
    candidate_id = generate_candidate_id(role_id, cleaned_contact)
    existing_mask = candidates_df["CandidateID"] == candidate_id
    exists = bool(existing_mask.any())
    if exists and not update_existing:
        st.warning("This candidate already exists. Tick Update existing candidate instead to update it.")
        return

    resume_filename = ""
    resume_path = ""
    if resume_file is not None:
        resume_filename, resume_path = save_uploaded_resume(resume_file, name, cleaned_contact, role_name)

    now = today_str()
    if exists:
        idx = candidates_df[existing_mask].index[0]
        candidates_df.loc[idx, ["Platform", "Name", "Contact", "RoleID", "Role", "Remarks", "LastUpdated"]] = [
            platform_name.strip(),
            name.strip(),
            cleaned_contact,
            role_id,
            role_name,
            remarks.strip(),
            now,
        ]
        if resume_file is not None:
            candidates_df.loc[idx, "ResumeFileName"] = resume_filename
            candidates_df.loc[idx, "ResumePath"] = resume_path
            append_activity_log("resume uploaded", role_name, candidate_id, name, cleaned_contact, resume_path)
        append_activity_log("candidate updated", role_name, candidate_id, name, cleaned_contact, "Updated from add form")
    else:
        new_row = {
            "CandidateID": candidate_id,
            "Platform": platform_name.strip(),
            "Name": name.strip(),
            "Contact": cleaned_contact,
            "RoleID": role_id,
            "Role": role_name,
            "Remarks": remarks.strip(),
            "Result": "Pending",
            "ResumeFileName": resume_filename,
            "ResumePath": resume_path,
            "DateAdded": now,
            "LastUpdated": now,
        }
        candidates_df = pd.concat([candidates_df, pd.DataFrame([new_row])], ignore_index=True)
        append_activity_log("candidate added", role_name, candidate_id, name, cleaned_contact, "New candidate")
        if resume_file is not None:
            append_activity_log("resume uploaded", role_name, candidate_id, name, cleaned_contact, resume_path)

    save_candidates(candidates_df)
    st.success(f"Saved candidate: {name.strip()} ({cleaned_contact})")
    st.rerun()


def render_candidate_tracker() -> None:
    roles_df = load_roles()
    candidates_df = load_candidates()
    st.header("Candidate Tracker")
    render_metrics(candidates_df)

    if st.button("Refresh data"):
        st.rerun()

    if candidates_df.empty:
        st.info("No candidates yet. Add a candidate from the Add Candidate tab.")
        return

    role_values = sorted([x for x in candidates_df["Role"].unique().tolist() if x])
    result_values = [result for result in RESULT_OPTIONS if result in candidates_df["Result"].unique().tolist()]
    platform_values = sorted([x for x in candidates_df["Platform"].unique().tolist() if x])

    st.subheader("Filters")
    filter_cols = st.columns(4)
    role_filter = filter_cols[0].selectbox("Role filter", ["All"] + role_values)
    result_filter = filter_cols[1].selectbox("Result filter", ["All"] + result_values)
    platform_filter = filter_cols[2].selectbox("Platform filter", ["All"] + platform_values)
    search = filter_cols[3].text_input("Search name/contact/remarks")

    filtered_df = filter_candidates(candidates_df, role_filter, result_filter, platform_filter, search)

    selected_id = st.session_state.get("selected_candidate_id", "")
    visible_ids = filtered_df["CandidateID"].tolist()
    if visible_ids and selected_id not in visible_ids:
        selected_id = visible_ids[0]
        st.session_state["selected_candidate_id"] = selected_id

    left, right = st.columns([1.05, 1.25], gap="large")
    with left:
        st.subheader("Candidates")
        st.caption(f"Showing {len(filtered_df)} of {len(candidates_df)} candidates")

        if filtered_df.empty:
            st.info("No candidates match the current filters.")
        else:
            select_actions = st.columns([1, 1, 2])
            if select_actions[0].button("Select visible", key="select_visible_candidates"):
                for candidate_id in filtered_df["CandidateID"].tolist():
                    st.session_state[f"select_candidate_{candidate_id}"] = True
                st.rerun()
            if select_actions[1].button("Clear selected", key="clear_selected_candidates"):
                for candidate_id in candidates_df["CandidateID"].tolist():
                    st.session_state[f"select_candidate_{candidate_id}"] = False
                st.rerun()

            header = st.columns([0.55, 1.2, 1, 2.2, 0.65])
            header[0].markdown("**Select**")
            header[1].markdown("**Name**")
            header[2].markdown("**Contact**")
            header[3].markdown("**Remarks**")
            header[4].markdown("**Open**")
            st.divider()

            for _, row in filtered_df.iterrows():
                is_selected = row["CandidateID"] == st.session_state.get("selected_candidate_id", "")
                marker = "**" if is_selected else ""
                remarks = row["Remarks"].strip() or "-"
                if len(remarks) > 95:
                    remarks = f"{remarks[:92]}..."

                row_cols = st.columns([0.55, 1.2, 1, 2.2, 0.65], vertical_alignment="center")
                row_cols[0].checkbox(
                    "Select candidate",
                    key=f"select_candidate_{row['CandidateID']}",
                    label_visibility="collapsed",
                )
                row_cols[1].markdown(f"{marker}{row['Name'] or '-'}{marker}")
                row_cols[2].markdown(f"{marker}{row['Contact'] or '-'}{marker}")
                row_cols[3].caption(remarks)
                if row_cols[4].button("Open", key=f"open_candidate_{row['CandidateID']}"):
                    st.session_state["selected_candidate_id"] = row["CandidateID"]
                    st.rerun()

        selected_export_ids = [
            candidate_id
            for candidate_id in candidates_df["CandidateID"].tolist()
            if st.session_state.get(f"select_candidate_{candidate_id}", False)
        ]
        selected_export_df = candidates_df[candidates_df["CandidateID"].isin(selected_export_ids)].copy()
        st.caption(f"{len(selected_export_df)} candidate(s) selected for export")

        with st.expander("Copy selected candidates", expanded=not selected_export_df.empty):
            if selected_export_df.empty:
                st.info("Tick one or more candidates to build a copy-ready list.")
            else:
                st.caption("Tab-separated format: Name, Contact, Remarks. Click inside, press Ctrl+A, then Ctrl+C.")
                st.text_area(
                    "Selected candidates copy text",
                    value=format_candidates_tab_copy(selected_export_df),
                    height=max(90, min(260, 46 * len(selected_export_df))),
                    key="selected_candidates_tab_copy",
                    label_visibility="collapsed",
                )

        export_col1, export_col2, export_col3 = st.columns(3)
        with export_col1:
            if st.button("Export selected CSV", disabled=selected_export_df.empty):
                path = export_dataframe(selected_export_df, "selected_candidates", "csv")
                st.success(f"Exported to {path}")
        with export_col2:
            if st.button("Export selected Excel", disabled=selected_export_df.empty):
                path = export_dataframe(selected_export_df, "selected_candidates", "xlsx")
                st.success(f"Exported to {path}")
        with export_col3:
            if st.button("Export filtered CSV", disabled=filtered_df.empty):
                path = export_dataframe(filtered_df, "filtered_candidates", "csv")
                st.success(f"Exported to {path}")

        with st.expander("Export by selected role/JD", expanded=False):
            role_export_options = {
                f"{row['RoleName']} | JD: {row['JDFileName'] or 'No JD attached'}": row["RoleID"]
                for _, row in roles_df.iterrows()
            }
            selected_role_labels = st.multiselect(
                "Select roles/JDs to export candidates for",
                list(role_export_options.keys()),
            )
            selected_role_ids = [role_export_options[label] for label in selected_role_labels]
            role_export_df = candidates_df[candidates_df["RoleID"].isin(selected_role_ids)].copy()
            st.caption(f"{len(role_export_df)} candidate(s) match selected role/JD choice")
            role_export_cols = st.columns(2)
            if role_export_cols[0].button("Export role/JD CSV", disabled=role_export_df.empty):
                path = export_dataframe(role_export_df, "role_jd_candidates", "csv")
                st.success(f"Exported to {path}")
            if role_export_cols[1].button("Export role/JD Excel", disabled=role_export_df.empty):
                path = export_dataframe(role_export_df, "role_jd_candidates", "xlsx")
                st.success(f"Exported to {path}")

    with right:
        st.subheader("Selected candidate")
        selected_id = st.session_state.get("selected_candidate_id", "")
        selected_rows = candidates_df[candidates_df["CandidateID"] == selected_id]
        if filtered_df.empty or selected_rows.empty:
            st.info("No candidate selected.")
            return

        selected = selected_rows.iloc[0]

        st.markdown(f"**{selected['Name']}**")
        detail_cols = st.columns(2)
        detail_cols[0].write(f"Contact: {selected['Contact']}")
        detail_cols[1].write(f"Role: {selected['Role']}")
        detail_cols[0].write(f"Platform: {selected['Platform']}")
        detail_cols[1].write(f"Candidate ID: {selected['CandidateID']}")
        detail_cols[0].write(f"Date added: {selected['DateAdded'] or '-'}")
        detail_cols[1].write(f"Last updated: {selected['LastUpdated'] or '-'}")

        st.divider()
        st.markdown("**Update candidate**")
        platform_value = st.text_input(
            "Platform",
            value=selected["Platform"],
            key=f"selected_platform_{selected_id}",
        )
        result_value = st.selectbox(
            "Result",
            RESULT_OPTIONS,
            index=RESULT_OPTIONS.index(selected["Result"]) if selected["Result"] in RESULT_OPTIONS else 0,
            key=f"selected_result_{selected_id}",
        )
        remarks_value = st.text_area(
            "Remarks",
            value=selected["Remarks"],
            height=120,
            key=f"selected_remarks_{selected_id}",
        )

        if st.button("Save Candidate Changes", type="primary", key=f"save_selected_{selected_id}"):
            updated = candidates_df.copy()
            idx = updated.index[updated["CandidateID"] == selected_id][0]
            before = updated.loc[idx].copy()
            updated.loc[idx, "Platform"] = platform_value.strip()
            updated.loc[idx, "Result"] = result_value
            updated.loc[idx, "Remarks"] = remarks_value.strip()

            changed = any(
                str(before[column]) != str(updated.loc[idx, column])
                for column in ["Platform", "Result", "Remarks"]
            )
            if changed:
                updated.loc[idx, "LastUpdated"] = today_str()
                if str(before["Result"]) != str(result_value):
                    append_activity_log(
                        "result changed",
                        selected["Role"],
                        selected_id,
                        selected["Name"],
                        selected["Contact"],
                        f"{before['Result']} -> {result_value}",
                    )
                append_activity_log(
                    "candidate updated",
                    selected["Role"],
                    selected_id,
                    selected["Name"],
                    selected["Contact"],
                    "Selected candidate panel save",
                )
                save_candidates(updated)
                st.success("Candidate changes saved.")
                st.rerun()
            else:
                st.info("No changes to save.")

        with st.expander("Stored file details", expanded=False):
            st.write(f"Resume file: `{selected['ResumeFileName'] or '-'}`")
            st.write(f"Resume path: `{selected['ResumePath'] or '-'}`")

        with st.expander("Copy for email", expanded=True):
            st.text_area(
                "Candidate information",
                value=format_candidate_for_email(selected),
                height=150,
                key=f"copy_candidate_{selected_id}",
                help="Click inside the box, press Ctrl+A, then Ctrl+C.",
            )
            st.text_area(
                "Spreadsheet paste format",
                value=format_candidates_tab_copy(pd.DataFrame([selected])),
                height=80,
                key=f"copy_candidate_tab_{selected_id}",
                help="Tab-separated: Name, Contact, Remarks. Paste into Excel or an email table.",
            )

        with st.expander("Resume", expanded=True):
            render_file_viewer(
                "Resume",
                selected.get("ResumePath", ""),
                selected.get("ResumeFileName", ""),
                key_prefix=f"candidate_resume_{selected.get('CandidateID', '')}",
            )

        role_row = find_role_by_id(roles_df, selected.get("RoleID", ""))
        with st.expander("Role JD", expanded=True):
            if role_row is None:
                st.warning("No matching role found for this candidate.")
            else:
                render_file_viewer(
                    "JD",
                    role_row.get("JDPath", ""),
                    role_row.get("JDFileName", ""),
                    key_prefix=f"candidate_jd_{selected.get('CandidateID', '')}",
                )


def render_diagnostics() -> None:
    st.header("Diagnostics")
    roles_df = load_roles()
    candidates_df = load_candidates()

    if st.button("Initialize missing structure"):
        ensure_workbook()
        st.success("Folder structure and workbook sheets are initialized.")

    folder_status = {
        "BASE_DIR": BASE_DIR,
        "resumes": RESUMES_DIR,
        "jd": JD_DIR,
        "logs": LOGS_DIR,
        "exports": EXPORTS_DIR,
    }
    missing = [name for name, path in folder_status.items() if not path.exists()]

    col1, col2, col3 = st.columns(3)
    col1.metric("Roles", len(roles_df))
    col2.metric("Active roles", len(get_active_roles(roles_df)))
    col3.metric("Candidates", len(candidates_df))

    col4, col5, col6 = st.columns(3)
    col4.metric("Resume files", len(list(RESUMES_DIR.glob("*"))) if RESUMES_DIR.exists() else 0)
    col5.metric("JD files", len(list(JD_DIR.glob("*"))) if JD_DIR.exists() else 0)
    col6.metric("Workbook exists", "Yes" if WORKBOOK_PATH.exists() else "No")

    st.subheader("Paths")
    st.write(f"BASE_DIR absolute path: `{BASE_DIR.resolve()}`")
    st.write(f"Workbook: `{WORKBOOK_PATH.resolve()}`")

    if WORKBOOK_PATH.exists():
        try:
            wb = load_workbook(WORKBOOK_PATH, read_only=True)
            st.write("Sheet names:", wb.sheetnames)
            wb.close()
        except Exception as exc:
            st.error(f"Could not read workbook sheets: {exc}")

    if missing:
        st.warning(f"Missing folders: {', '.join(missing)}")
    else:
        st.success("All required folders exist.")

    with st.expander("Recent activity", expanded=False):
        log_df = _load_sheet(ACTIVITY_LOG_SHEET, ACTIVITY_COLUMNS)
        if log_df.empty:
            st.info("No activity logged yet.")
        else:
            st.dataframe(log_df.tail(50).iloc[::-1], use_container_width=True, hide_index=True)


def main() -> None:
    ensure_workbook()
    st.title("Recruitment Tracker")
    st.caption("Local-folder based recruitment workspace for roles, JDs, resumes, and candidate tracking.")

    tabs = st.tabs(["Role Management", "Candidate Tracker", "Add Candidate", "Diagnostics"])
    with tabs[0]:
        render_role_management()
    with tabs[1]:
        render_candidate_tracker()
    with tabs[2]:
        render_add_candidate()
    with tabs[3]:
        render_diagnostics()


if __name__ == "__main__":
    main()
