from __future__ import annotations

from datetime import date, time
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from Flexar.BlueSG.operation_context import OperationContext
from Flexar.BlueSG.run_metrics import create_run_result
from Flexar.BlueSG.vehicle_route_optimizer import (
    RiderState,
    export_routes_to_excel,
    load_and_validate_jobs,
    optimisation_integrity_report,
    optimise_vehicle_routes,
)


def _workbook_bytes(job_count: int = 3, *, include_fallback: bool = False) -> bytes:
    rows = [
        {
            "Car Plate": f"CAR{index:02d}",
            "Pickup Address": "Tampines",
            "Pickup Lot": str(index),
            "Drop-off Address": "Bedok" if include_fallback and index == 0 else "Tampines",
            "Date": "17/07/2026",
        }
        for index in range(job_count)
    ]
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, index=False, sheet_name="Daily RB task")
    return output.getvalue()


def test_full_parse_optimise_validate_and_export_flow() -> None:
    source = _workbook_bytes(include_fallback=True)
    jobs, missing, _ = load_and_validate_jobs(BytesIO(source))
    assert not missing
    context = OperationContext.for_window(date(2026, 7, 17), time(23), time(3))
    riders = [RiderState("R1", "Tampines", "East", max_jobs=1)]
    route_df, summary_df, warnings = optimise_vehicle_routes(jobs, riders, use_onemap=False, operation_context=context)
    integrity = optimisation_integrity_report(route_df, jobs)
    assert integrity["is_valid"]
    assert len(route_df) == 3  # Soft Max Jobs total is below selected work.
    result = create_run_result(
        route_df=route_df,
        unassigned_df=integrity["unassigned_df"],
        riders=riders,
        context=context,
        settings={"jobs_uploaded": len(jobs), "use_onemap": False},
        input_filename="fixture.xlsx",
        input_sha256="fixture-hash",
        selected_job_date="2026-07-17",
        warnings=[{"severity": "manual_review", "message": warning} for warning in warnings],
        validation={key: value for key, value in integrity.items() if key != "unassigned_df"},
    )
    output = export_routes_to_excel(route_df, summary_df, jobs_df=jobs, lookup_warnings=warnings, run_result=result)
    workbook = load_workbook(BytesIO(output), read_only=True)
    assert {
        "How To Read This",
        "Optimised Routes",
        "Map Loader",
        "Unassigned Jobs",
        "Summary",
        "Rider Instructions",
        "Manual Review",
        "Regional Capacity",
        "Regional Assignment Audit",
        "Local Search Audit",
        "Run Metadata",
    } <= set(workbook.sheetnames)
    rider_text = "\n".join(str(cell.value or "") for row in workbook["Rider Instructions"].iter_rows() for cell in row)
    manager_review_text = "\n".join(
        str(cell.value or "") for row in workbook["Manual Review"].iter_rows() for cell in row
    )
    assert "LOW-CONFIDENCE ROUTE" not in rider_text
    assert "LOW-CONFIDENCE ROUTE" in manager_review_text
    assert route_df["Assignment Tier"].isin(["primary", "support", "exceptional"]).all()


def test_thirty_selected_jobs_remain_thirty_when_feasible() -> None:
    source = _workbook_bytes(30)
    jobs, _, _ = load_and_validate_jobs(BytesIO(source))
    context = OperationContext.for_window(date(2026, 7, 17), time(23), time(3))
    rider = RiderState("R", "Tampines", "East", max_jobs=3)
    route_df, _, _ = optimise_vehicle_routes(jobs, [rider], use_onemap=False, operation_context=context)
    assert len(route_df) == 30
    assert route_df["Uploaded Row"].is_unique


def test_route_chain_and_deterministic_tie_breaking(regression_jobs, overnight_context) -> None:
    def run():
        riders = [RiderState("B", "Tampines", "East"), RiderState("A", "Tampines", "East")]
        route_df, _, _ = optimise_vehicle_routes(
            regression_jobs.iloc[:2], riders, use_onemap=False, operation_context=overnight_context
        )
        return route_df

    first = run()
    second = run()
    pd.testing.assert_frame_equal(first.reset_index(drop=True), second.reset_index(drop=True))
    assert first["Route Validation Status"].eq("OK").all()
    for _, routes in first.sort_values(["Rider", "Sequence"]).groupby("Rider"):
        if len(routes) > 1:
            assert routes.iloc[1]["Start From"] == routes.iloc[0]["Drop-off Address"]
