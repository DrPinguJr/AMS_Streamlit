from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from Flexar.BlueSG.build_optimised_vehicle_routes import (
    RiderState,
    SUMMARY_COLUMNS,
    build_jobs_by_stable_id,
    export_routes_to_excel,
    format_summary_output,
    rebuild_outputs_from_sequences,
)
from Flexar.BlueSG.job_import_staging import (
    CANONICAL_JOB_COLUMNS,
    commit_staged_jobs,
    parse_csv_jobs,
    parse_delimited_text,
    parse_excel_jobs,
    parse_html_job_list,
    validate_staged_jobs,
)
from Flexar.BlueSG.manual_route_assignment_editing_and_recalculation import (
    UNASSIGNED_LANE,
    assignment_from_routes,
    incremental_recalculate,
)
from Flexar.BlueSG.optimiser_workflow_state import (
    begin_rider_draft,
    cancel_rider_draft,
    commit_job_draft,
    commit_optimiser_result,
    initialise_workflow_state,
    save_rider_draft,
    streamlit_key_value_table,
    validate_and_commit_job_import,
    validate_assignment_draft,
)


def _excel_bytes(rows: list[list[object]]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, index=False, header=False)
    return output.getvalue()


def _riders() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Rider Name": "A",
                "Start Location": "Tampines",
                "Start Zone": "East",
                "Max Jobs": 5,
                "Rider Load": "Medium",
            },
            {
                "Rider Name": "B",
                "Start Location": "Yishun",
                "Start Zone": "North",
                "Max Jobs": 5,
                "Rider Load": "Medium",
            },
        ]
    )


def _jobs() -> pd.DataFrame:
    return commit_staged_jobs(
        pd.DataFrame(
            [
                {
                    "Job ID": "J1",
                    "Car Plate": "SPE1001A",
                    "Pickup Address": "Tampines",
                    "Pickup Lot": "A1",
                    "Drop-off Address": "Bedok",
                    "Drop-off Lot": "B1",
                    "Deadline": "2026-07-31 16:00",
                    "_original_order": 0,
                },
                {
                    "Job ID": "J2",
                    "Car Plate": "SPE1002B",
                    "Pickup Address": "Yishun",
                    "Pickup Lot": "A2",
                    "Drop-off Address": "Woodlands",
                    "Drop-off Lot": "B2",
                    "Deadline": "2026-07-31 16:30",
                    "_original_order": 1,
                },
            ]
        )
    )


def test_old_flexar_excel_header_format_maps_driver_and_locations() -> None:
    rows = [
        [
            "Date",
            "CarPlate",
            "Fuel %",
            "Pick up location",
            "Lots number",
            "Pick Up Time",
            "Unlock Time",
            "End location",
            "Lots number",
            "Drop off Time",
            "Driver",
            "Remarks",
        ],
        [
            "31/07/2026",
            "SPE1001A",
            50,
            "Tampines",
            "A1",
            "",
            "",
            "Bedok",
            "B2",
            "",
            "Rider A",
            "Check",
        ],
    ]
    result = parse_excel_jobs(_excel_bytes(rows))
    job = result.dataframe.iloc[0]
    assert job["Car Plate"] == "SPE1001A"
    assert job["Pickup Address"] == "Tampines"
    assert job["Drop-off Address"] == "Bedok"
    assert job["Worker"] == "Rider A"


def test_excel_scans_extra_title_rows_before_header() -> None:
    result = parse_excel_jobs(
        _excel_bytes(
            [
                ["Provided by Flexar", "", "", ""],
                ["By Antares", "", "", ""],
                ["Car Plate", "Pickup", "Pickup Lot", "Destination"],
                ["SPE1001A", "Tampines", "A1", "Bedok"],
            ]
        )
    )
    assert result.report.rows_detected == 1
    assert result.dataframe.iloc[0]["Pickup Address"] == "Tampines"


def test_duplicate_lots_number_columns_resolve_by_position() -> None:
    result = parse_excel_jobs(
        _excel_bytes(
            [
                ["CarPlate", "Pick up location", "Lots number", "End location", "Lots number"],
                ["SPE1001A", "Tampines", "P-12", "Bedok", "D-34"],
            ]
        )
    )
    assert result.dataframe.iloc[0]["Pickup Lot"] == "P-12"
    assert result.dataframe.iloc[0]["Drop-off Lot"] == "D-34"


def test_csv_import_normalises_aliases() -> None:
    payload = (
        b"task id,licence plate,origin,pickup parking lot,destination,end lot\n"
        b"J1,SPE1001A,Tampines,A1,Bedok,B2\n"
    )
    result = parse_csv_jobs(payload)
    assert result.dataframe.loc[0, CANONICAL_JOB_COLUMNS[:6]].tolist() == [
        "J1",
        "SPE1001A",
        "Tampines",
        "A1",
        "Bedok",
        "B2",
    ]


def test_flexar_task_export_csv_maps_station_schema() -> None:
    payload = (
        b"Task ID,Task status,Escalation,Vehicle plate,Start station,End station,Deadline (SGT),Assigned to\n"
        b"48124,Lapsed,No,SPE9345E,Near Tengah Bus Interchange - 322 Tengah Drive S690322,"
        b"Near Clementi Mall - 441 Commonwealth Avenue West S120441,2026-07-30 23:04,\n"
    )
    result = parse_csv_jobs(payload)
    row = result.dataframe.iloc[0]
    assert row["Job ID"] == "48124"
    assert row["Status"] == "Lapsed"
    assert row["Car Plate"] == "SPE9345E"
    assert row["Pickup Address"] == "Near Tengah Bus Interchange - 322 Tengah Drive S690322"
    assert row["Drop-off Address"] == "Near Clementi Mall - 441 Commonwealth Avenue West S120441"
    assert row["Deadline"] == "2026-07-30 23:04"
    assert row["Worker"] == ""
    assert row["Escalation"] == "No"
    assert result.report.rows_accepted == 1
    assert result.report.rows_requiring_correction == 0


def test_html_list_import_keeps_list_fields_and_flags_missing_locations() -> None:
    result = parse_html_job_list(
        """
        <table>
          <tr><th>Status</th><th>ID</th><th>Job</th><th>Vehicle</th><th>Created At</th><th>Deadline</th><th>Worker</th></tr>
          <tr><td>Available</td><td>42</td><td>Move</td><td>SPE1001A</td><td>Today</td><td>16:00</td><td>-</td></tr>
        </table>
        """
    )
    row = result.dataframe.iloc[0]
    assert row["Job ID"] == "42"
    assert row["Car Plate"] == "SPE1001A"
    assert row["Pickup Address"] == row["Drop-off Address"] == ""
    assert result.report.missing_location_rows == 1
    assert not validate_staged_jobs(result.dataframe).is_valid


def test_pasted_tab_separated_rows() -> None:
    result = parse_delimited_text(
        "Job ID\tLP\tPickup Address\tDrop-off Address\n"
        "J1\tSPE1001A\tTampines\tBedok"
    )
    assert result.report.rows_accepted == 1
    assert result.dataframe.iloc[0]["Job ID"] == "J1"


def test_missing_required_locations_blocks_commit() -> None:
    jobs = pd.DataFrame([{"Car Plate": "SPE1001A", "Pickup Address": "", "Drop-off Address": "Bedok"}])
    validation = validate_staged_jobs(jobs)
    assert not validation.is_valid
    assert validation.report.missing_location_rows == 1
    with pytest.raises(ValueError, match="Pickup Address"):
        commit_staged_jobs(jobs)


def test_automatic_import_commits_valid_source_and_preserves_jobs_on_invalid_replacement() -> None:
    state: dict = {}
    initialise_workflow_state(state, _riders())

    valid_result, changed = validate_and_commit_job_import(state, _jobs())
    assert valid_result.is_valid
    assert changed
    committed_before = state["committed_jobs"].copy(deep=True)

    invalid = _jobs()
    invalid.loc[0, "Pickup Address"] = ""
    invalid_result, changed = validate_and_commit_job_import(state, invalid)

    assert not invalid_result.is_valid
    assert not changed
    pd.testing.assert_frame_equal(state["committed_jobs"], committed_before)


def test_page_uses_automatic_import_rider_dialog_and_single_column_steps() -> None:
    page = (
        Path(__file__).parents[1]
        / "pages"
        / "create_optimised_vehicle_routes_page.py"
    ).read_text(encoding="utf-8")

    assert "Load Into Staging" not in page
    assert "Validate Jobs" not in page
    assert "Commit Job List" not in page
    assert '@st.dialog("Today\'s riders", width="large")' in page
    assert '@st.dialog("OneMap token", width="small")' in page
    assert '"OneMap token"' in page
    assert 'st.subheader("1. Upload jobs")' in page
    assert 'st.subheader("2. Run optimiser")' in page
    assert 'st.subheader("3. Review results")' in page
    assert "action_col = st.container()" in page
    assert "review_col = st.container()" in page
    assert "action_col, review_col = st.columns" not in page


def test_duplicate_plates_with_different_job_ids_are_retained_and_flagged() -> None:
    jobs = pd.DataFrame(
        [
            {"Job ID": "J1", "Car Plate": "SPE1001A", "Pickup Address": "Tampines", "Drop-off Address": "Bedok"},
            {"Job ID": "J2", "Car Plate": "SPE1001A", "Pickup Address": "Yishun", "Drop-off Address": "Woodlands"},
        ]
    )
    validation = validate_staged_jobs(jobs)
    assert validation.is_valid
    assert any(issue.code == "duplicate_car_plate" for issue in validation.warnings)
    assert len(commit_staged_jobs(jobs)) == 2


def test_rider_draft_cancel_discards_changes() -> None:
    state: dict = {}
    initialise_workflow_state(state, _riders())
    draft = begin_rider_draft(state)
    draft.loc[0, "Rider Name"] = "Changed"
    state["rider_draft"] = draft
    cancel_rider_draft(state)
    assert state["rider_draft"] is None
    assert state["committed_riders"].iloc[0]["Rider Name"] == "A"


def test_rider_draft_save_commits_all_changes_and_marks_result_stale() -> None:
    state: dict = {}
    initialise_workflow_state(state, _riders())
    state["optimiser_result"] = {"route_df": pd.DataFrame()}
    draft = begin_rider_draft(state)
    draft.loc[0, "Maximum Jobs"] = 8
    assert save_rider_draft(state, draft)
    assert state["committed_riders"].loc[0, "Maximum Jobs"] == 8
    assert state["result_is_stale"]


def test_progress_detail_table_is_arrow_safe_with_bytes_and_numbers() -> None:
    import pyarrow as pa

    table = streamlit_key_value_table(
        [
            ("Address", b"Tampines"),
            ("Orders before this", 2),
            ("Duration", 5.5),
        ]
    )
    arrow_table = pa.Table.from_pandas(table)
    assert table.dtypes.astype(str).tolist() == ["string", "string"]
    assert arrow_table.column("Value").to_pylist() == ["Tampines", "2", "5.5"]


def test_assignment_draft_rejects_duplicate_job() -> None:
    draft = pd.DataFrame(
        [
            {"Job ID": "J1", "Rider": "A", "Sequence": 1},
            {"Job ID": "J1", "Rider": "B", "Sequence": 1},
            {"Job ID": "J2", "Rider": "B", "Sequence": 2},
        ]
    )
    result = validate_assignment_draft(draft, {"J1", "J2"}, {"A", "B"})
    assert not result.is_valid
    assert result.duplicate_job_ids == ("J1",)


def test_assignment_draft_rejects_disappearing_job() -> None:
    draft = pd.DataFrame([{"Job ID": "J1", "Rider": "A", "Sequence": 1}])
    result = validate_assignment_draft(draft, {"J1", "J2"}, {"A", "B"})
    assert not result.is_valid
    assert result.missing_job_ids == ("J2",)


def _summary_builder(route_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for rider, routes in route_df.groupby("Rider", sort=False):
        total = float(pd.to_numeric(routes["Total Duration Min"], errors="coerce").fillna(0).sum())
        rows.append(
            {
                "Rider": rider,
                "Total Jobs": len(routes),
                "Total Route Duration Min": total,
                "Adjusted Route Duration Min": total * 1.2,
                "Within 3 Hours": "OK",
                "Final Location": routes.iloc[-1]["Drop-off Address"],
            }
        )
    frame = pd.DataFrame(rows)
    for column in SUMMARY_COLUMNS:
        if column not in frame.columns:
            frame[column] = 0 if "Distance" in column or "Duration" in column else ""
    return format_summary_output(frame[SUMMARY_COLUMNS], route_df)


def test_recalculation_after_changing_a_rider_assignment() -> None:
    jobs = _jobs()
    riders = [
        RiderState("A", "Tampines", "East", max_jobs=5),
        RiderState("B", "Yishun", "North", max_jobs=5),
    ]
    ids = list(build_jobs_by_stable_id(jobs))
    original_sequences = {"A": [ids[0]], "B": [ids[1]]}
    routes, _, _ = rebuild_outputs_from_sequences(
        original_sequences,
        riders,
        build_jobs_by_stable_id(jobs),
        jobs_df=jobs,
        use_onemap=False,
    )
    rider_df = _riders()
    original = assignment_from_routes(routes, jobs, ["A", "B"])
    changed = {"A": [], "B": ids, UNASSIGNED_LANE: []}
    result = incremental_recalculate(
        confirmed_routes=routes,
        confirmed_assignment=original,
        draft_assignment=changed,
        rider_df=rider_df,
        jobs_df=jobs,
        settings={"use_onemap": False},
        summary_builder=_summary_builder,
    )
    assert set(result.affected_riders) == {"A", "B"}
    assert set(result.route_df["Rider"]) == {"B"}
    assert result.route_df.sort_values("Sequence")["Sequence"].tolist() == [1, 2]


def test_stale_result_detection_after_committed_job_change() -> None:
    state: dict = {}
    initialise_workflow_state(state, _riders())
    first = _jobs()
    commit_job_draft(state, first)
    commit_optimiser_result(state, {"route_df": pd.DataFrame()}, jobs=state["committed_jobs"], riders=state["committed_riders"])
    changed = first.copy()
    changed.loc[0, "Pickup Address"] = "Simei"
    commit_job_draft(state, changed)
    assert state["result_is_stale"]


def test_existing_export_compatibility_and_flexar_assignment_sheet() -> None:
    jobs = _jobs()
    riders = [RiderState("A", "Tampines", "East", max_jobs=5)]
    ids = list(build_jobs_by_stable_id(jobs))
    route_df, summary_df, _ = rebuild_outputs_from_sequences(
        {"A": ids},
        riders,
        build_jobs_by_stable_id(jobs),
        jobs_df=jobs,
        use_onemap=False,
    )
    payload = export_routes_to_excel(route_df, summary_df, jobs_df=jobs)
    workbook = load_workbook(BytesIO(payload), read_only=True)
    assert {
        "How To Read This",
        "Optimised Routes",
        "Map Loader",
        "Unassigned Jobs",
        "Summary",
        "Rider Instructions",
        "Flexar Assignment List",
    } <= set(workbook.sheetnames)
    sheet = workbook["Flexar Assignment List"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=3, max_row=3))]
    assert headers == [
        "Sequence",
        "Rider",
        "Car Plate",
        "Job ID",
        "Pickup Address",
        "Pickup Lot",
        "Drop-off Address",
        "Drop-off Lot",
        "Deadline",
        "Status",
        "Operator Confirmation",
        "Remarks",
    ]
    rows = list(sheet.iter_rows(min_row=4, values_only=True))
    assert [row[2] for row in rows] == ["SPE1001A", "SPE1002B"]
    assert all(row[10] is None and row[11] is None for row in rows)
