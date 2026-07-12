from __future__ import annotations

import json
import hashlib
import itertools
import math
import os
import re
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Callable
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

import pandas as pd
try:
    import streamlit as st
except ImportError:  # pragma: no cover - Streamlit is present in the app, but keep helpers importable.
    st = None
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REQUIRED_JOB_HEADERS = ["Car Plate", "Pickup Address", "Pickup Lot", "Drop-off Address"]
OPTIONAL_JOB_HEADERS = ["Date", "Fuel %", "Pickup Time", "Notes"]
RIDER_COLUMNS = ["Rider Name", "Start Location", "Start Zone", "Max Jobs", "Rider Load"]
RIDER_LOAD_LEVELS = ["Low", "Medium", "High", "Very High"]
RIDER_LOAD_ALIASES = {
    "Normal": "Medium",
}
RIDER_LOAD_POLICIES = {
    "Low": {
        "job_score_adjustment": 45.0,
        "job_escalation": 18.0,
        "empty_duration_soft_limit": 15.0,
        "empty_duration_penalty_per_min": 6.0,
        "different_pickup_zone_penalty": 110.0,
        "cross_zone_route_penalty": 55.0,
        "same_area_bonus": -30.0,
        "cluster_pressure_multiplier": 0.25,
        "cluster_jump_penalty": 110.0,
        "minimum_target_multiplier": 0.0,
    },
    "Medium": {
        "job_score_adjustment": 0.0,
        "job_escalation": 0.0,
        "empty_duration_soft_limit": 999.0,
        "empty_duration_penalty_per_min": 0.0,
        "different_pickup_zone_penalty": 0.0,
        "cross_zone_route_penalty": 0.0,
        "same_area_bonus": 0.0,
        "cluster_pressure_multiplier": 1.0,
        "cluster_jump_penalty": 25.0,
        "minimum_target_multiplier": 1.0,
    },
    "High": {
        "job_score_adjustment": -25.0,
        "job_escalation": -4.0,
        "empty_duration_soft_limit": 999.0,
        "empty_duration_penalty_per_min": 0.0,
        "different_pickup_zone_penalty": 0.0,
        "cross_zone_route_penalty": 0.0,
        "same_area_bonus": -8.0,
        "cluster_pressure_multiplier": 1.3,
        "cluster_jump_penalty": 15.0,
        "minimum_target_multiplier": 1.0,
    },
    "Very High": {
        "job_score_adjustment": -42.0,
        "job_escalation": -8.0,
        "empty_duration_soft_limit": 999.0,
        "empty_duration_penalty_per_min": 0.0,
        "different_pickup_zone_penalty": 0.0,
        "cross_zone_route_penalty": 0.0,
        "same_area_bonus": -18.0,
        "cluster_pressure_multiplier": 1.8,
        "cluster_jump_penalty": 5.0,
        "minimum_target_multiplier": 1.2,
    },
}
WEEKDAY_SHEETS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
BASE_DIR = Path(__file__).resolve().parent
ROSTER_FILE = BASE_DIR / "rider_roster.xlsx"

ROUTE_COLUMNS = [
    "Rider",
    "Sequence",
    "Uploaded Row",
    "Start From",
    "Empty Travel To Pickup",
    "Empty PT Instructions",
    "Empty Route Path",
    "Car Plate",
    "Pickup Address",
    "Pickup Lot",
    "Drop-off Address",
    "Loaded Travel / Car Movement",
    "Loaded Drive Instructions",
    "Loaded Route Path",
    "Empty Distance KM",
    "Empty Duration Min",
    "Loaded Distance KM",
    "Loaded Duration Min",
    "Total Distance KM",
    "Total Duration Min",
    "Assignment Score",
    "Zone Adjustment",
    "Same Zone Pickup",
    "Same Zone Route",
    "Route Zone Priority",
    "Empty Weight",
    "Loaded Weight",
    "Workload Penalty",
    "Duration Penalty",
    "Max Jobs Penalty",
    "Projected Rider Duration Min",
    "Projected Adjusted Duration Min",
    "First Positioning PT Duration Min",
    "First Pickup ETA",
    "Latest Departure Time",
    "In-Window Route Duration Min",
    "Final Completion ETA",
    "Cluster Name / Zone",
    "Cluster Job Count",
    "Feasibility Status",
    "Reason if Unassigned",
    "Cost Source",
    "Route Validation Status",
]
SUMMARY_COLUMNS = [
    "Rider",
    "Total Jobs",
    "Total Empty Distance KM",
    "Total Empty Duration Min",
    "Total Loaded Distance KM",
    "Total Loaded Duration Min",
    "Total Route Distance KM",
    "Total Route Duration Min",
    "Adjusted Route Duration Min",
    "Total Route Duration Hours",
    "Within 3 Hours",
    "Final Location",
    "Empty Travel %",
    "Loaded Travel %",
    "Workload Comment",
]

ZONE_KEYWORDS = {
    "North": ["Woodlands", "Admiralty", "Yishun", "Sembawang", "Canberra"],
    "North-East": ["Sengkang", "Punggol", "Hougang", "Serangoon"],
    "East": ["Pasir Ris", "Tampines", "Simei", "Bedok", "Paya Lebar", "PLQ"],
    "Central": ["Toa Payoh", "Bishan", "Ang Mo Kio", "Novena", "Kallang"],
    "West": ["Jurong", "Bukit Batok", "Choa Chu Kang", "Clementi", "Bukit Panjang"],
    "South/CBD": ["Orchard", "Marina", "Raffles", "Tanjong Pagar", "HarbourFront"],
}

FALLBACK_ZONE_MINUTES = {
    "North": {"North": 10, "North-East": 20, "East": 35, "Central": 25, "West": 32, "South/CBD": 38},
    "North-East": {"North": 20, "North-East": 10, "East": 22, "Central": 22, "West": 38, "South/CBD": 34},
    "East": {"North": 35, "North-East": 22, "East": 10, "Central": 25, "West": 45, "South/CBD": 32},
    "Central": {"North": 25, "North-East": 22, "East": 25, "Central": 10, "West": 25, "South/CBD": 16},
    "West": {"North": 32, "North-East": 38, "East": 45, "Central": 25, "West": 10, "South/CBD": 33},
    "South/CBD": {"North": 38, "North-East": 34, "East": 32, "Central": 16, "West": 33, "South/CBD": 10},
}

FALLBACK_ZONE_KM = {
    "North": {"North": 4, "North-East": 10, "East": 18, "Central": 13, "West": 16, "South/CBD": 20},
    "North-East": {"North": 10, "North-East": 4, "East": 11, "Central": 11, "West": 20, "South/CBD": 17},
    "East": {"North": 18, "North-East": 11, "East": 4, "Central": 13, "West": 25, "South/CBD": 17},
    "Central": {"North": 13, "North-East": 11, "East": 13, "Central": 4, "West": 13, "South/CBD": 7},
    "West": {"North": 16, "North-East": 20, "East": 25, "Central": 13, "West": 4, "South/CBD": 17},
    "South/CBD": {"North": 20, "North-East": 17, "East": 17, "Central": 7, "West": 17, "South/CBD": 4},
}

UNKNOWN_DISTANCE_KM = 18.0
UNKNOWN_DURATION_MIN = 40.0
SAME_UNKNOWN_DISTANCE_KM = 6.0
SAME_UNKNOWN_DURATION_MIN = 18.0
DEFAULT_EMPTY_WEIGHT = 4.5
DEFAULT_LOADED_WEIGHT = 0.8
DEFAULT_SOFT_WORKLOAD_MIN = 115.0
DEFAULT_WORKLOAD_PENALTY_PER_MIN = 2.0
DEFAULT_SOFT_ADJUSTED_DURATION_MIN = 165.0
DEFAULT_DURATION_PENALTY_PER_MIN = 2.0
DEFAULT_MAX_JOB_OVERAGE_PENALTY = 60.0
DEFAULT_DURATION_BUFFER_MULTIPLIER = 1.2
DEFAULT_MAX_ADJUSTED_DURATION_MIN = 180.0
DEFAULT_EMPTY_TRAVEL_DURATION_MULTIPLIER = 1.5
DEFAULT_EMPTY_TRAVEL_WAIT_BUFFER_MIN = 6.0
DEFAULT_CLUSTER_PRESSURE_BONUS_PER_JOB = 30.0
DEFAULT_MIN_JOBS_PER_RIDER = 2
DEFAULT_SELECTIVE_CHANGED_RIDER_PENALTY = 20.0
DEFAULT_SELECTIVE_MOVED_JOB_PENALTY = 10.0
DEFAULT_SELECTIVE_SEQUENCE_CHANGE_PENALTY = 5.0
MAX_SELECTIVE_RESHUFFLE_CANDIDATES = 50_000
DEFAULT_SELECTIVE_BEAM_WIDTH = 100
EXCEL_CELL_MAX_CHARS = 32767
JOB_WINDOW_START_MIN = 14 * 60
JOB_WINDOW_END_MIN = 17 * 60
FIRST_POSITIONING_WEIGHT = 0.2
SHORT_WALK_DISTANCE_KM = 1.0
SHORT_WALK_DURATION_MIN = 15.0
ONEMAP_SEARCH_URL = "https://www.onemap.gov.sg/api/common/elastic/search"
ONEMAP_ROUTE_URL = "https://www.onemap.gov.sg/api/public/routingsvc/route"
ONEMAP_AUTH_URL = "https://www.onemap.gov.sg/api/auth/post/getToken"
CACHE_DIR = BASE_DIR / "cache"
PROJECT_ROOT = Path(__file__).resolve().parents[2]
ENV_FILE = PROJECT_ROOT / ".env"
RUNTIME_CACHE_DIR = Path(
    os.getenv(
        "BLUESG_RUNTIME_CACHE_DIR",
        str(CACHE_DIR / "runtime"),
    )
)
GEOCODE_SEED_CACHE_FILE = CACHE_DIR / "onemap_geocode_cache.csv"
ROUTE_SEED_CACHE_FILE = CACHE_DIR / "onemap_route_cache.csv"
GEOCODE_CACHE_FILE = RUNTIME_CACHE_DIR / "onemap_geocode_cache.csv"
ROUTE_CACHE_FILE = RUNTIME_CACHE_DIR / "onemap_route_cache.csv"
ProgressCallback = Callable[[dict[str, Any]], None]
GEOCODE_MEMORY_CACHE: dict[str, GeocodeResult] = {}
ROUTE_MEMORY_CACHE: dict[str, TravelCost] = {}
GEOCODE_DISK_CACHE_LOADED = False
ROUTE_DISK_CACHE_LOADED = False
SINGAPORE_TZ = timezone(timedelta(hours=8))
ONEMAP_MEMORY_TOKEN: str = ""
ONEMAP_MEMORY_TOKEN_EXPIRY: datetime | None = None


@dataclass
class RiderState:
    name: str
    start_location: str
    start_zone: str | None = None
    max_jobs: int | None = None
    load_level: str = "Medium"
    current_location: str = ""
    current_zone: str | None = None
    assigned_count: int = 0
    empty_distance_km: float = 0
    empty_duration_min: float = 0
    loaded_distance_km: float = 0
    loaded_duration_min: float = 0

    @classmethod
    def from_row(cls, row: pd.Series) -> "RiderState":
        start_location = clean_text(row.get("Start Location"))
        start_zone = clean_text(row.get("Start Zone")) or infer_zone(start_location)
        return cls(
            name=clean_text(row.get("Rider Name")),
            start_location=start_location,
            start_zone=start_zone,
            max_jobs=parse_optional_int(row.get("Max Jobs")),
            load_level=normalise_rider_load_level(row.get("Rider Load")),
            current_location=start_location,
            current_zone=start_zone,
        )

    @property
    def has_capacity(self) -> bool:
        return self.max_jobs is None or self.assigned_count < self.max_jobs


@dataclass(frozen=True)
class GeocodeResult:
    address: str
    latitude: float | None
    longitude: float | None
    source: str
    error: str = ""

    @property
    def is_available(self) -> bool:
        return self.latitude is not None and self.longitude is not None


@dataclass(frozen=True)
class TravelCost:
    distance_km: float | None
    duration_min: float | None
    source: str
    error: str = ""
    route_text: str = ""
    route_path: list[list[float]] | None = None

    def optimisation_value(self, optimise_by: str) -> float:
        if optimise_by == "distance":
            return self.distance_km if self.distance_km is not None else self.duration_min or math.inf
        return self.duration_min if self.duration_min is not None else self.distance_km or math.inf


def clean_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def _split_lookup_warning(warning: str) -> tuple[str, str]:
    if ": " not in warning:
        return warning, ""
    lookup_text, error = warning.rsplit(": ", 1)
    return lookup_text.strip(), error.strip()


def _dedupe_lookup_warnings(warnings: list[str]) -> list[str]:
    route_warning_parts: set[tuple[str, str]] = set()
    for warning in warnings:
        lookup_text, error = _split_lookup_warning(warning)
        if not error or " -> " not in lookup_text:
            continue
        for address in lookup_text.split(" -> ", 1):
            route_warning_parts.add((address.strip(), error))

    deduped = []
    seen = set()
    for warning in warnings:
        lookup_text, error = _split_lookup_warning(warning)
        is_standalone_duplicate = error and " -> " not in lookup_text and (lookup_text, error) in route_warning_parts
        if is_standalone_duplicate or warning in seen:
            continue
        seen.add(warning)
        deduped.append(warning)
    return sorted(deduped)


def _minimum_job_target(total_jobs: int, rider_count: int) -> int:
    if rider_count <= 0:
        return 0
    return min(DEFAULT_MIN_JOBS_PER_RIDER, total_jobs // rider_count)


def _rider_minimum_job_target(rider: "RiderState", shared_target: int) -> int:
    load_level = normalise_rider_load_level(rider.load_level)
    policy = RIDER_LOAD_POLICIES.get(load_level, RIDER_LOAD_POLICIES["Medium"])
    shared_target = int(math.floor(shared_target * float(policy["minimum_target_multiplier"])))
    if rider.max_jobs is None:
        return shared_target
    return min(shared_target, rider.max_jobs)


def _minimum_job_priority(assigned_jobs: int, target_jobs: int) -> int:
    return 0 if assigned_jobs < target_jobs else 1


def parse_optional_int(value: Any) -> int | None:
    if pd.isna(value) or value == "":
        return None
    try:
        parsed = int(float(value))
    except (TypeError, ValueError):
        return None
    return parsed if parsed > 0 else None


def normalise_rider_load_level(value: Any) -> str:
    text = clean_text(value)
    text = RIDER_LOAD_ALIASES.get(text, text)
    for level in RIDER_LOAD_LEVELS:
        if text.casefold() == level.casefold():
            return level
    return "Medium"


def _read_csv_cache(path: Path, columns: list[str]) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(columns=columns)
    try:
        return pd.read_csv(path)
    except Exception:
        return pd.DataFrame(columns=columns)


def _write_csv_cache(path: Path, df: pd.DataFrame) -> None:
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        df.to_csv(path, index=False)
    except OSError:
        return


def _append_csv_cache_row(path: Path, row: dict[str, Any], columns: list[str]) -> None:
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        pd.DataFrame([row], columns=columns).to_csv(path, mode="a", index=False, header=not path.exists())
    except OSError:
        return


def _load_geocode_disk_cache_once() -> None:
    global GEOCODE_DISK_CACHE_LOADED
    if GEOCODE_DISK_CACHE_LOADED:
        return
    for cache_file in [GEOCODE_SEED_CACHE_FILE, GEOCODE_CACHE_FILE]:
        cache = _read_csv_cache(cache_file, ["address", "latitude", "longitude"])
        for row in cache.itertuples(index=False):
            address = clean_text(getattr(row, "address", ""))
            if not address:
                continue
            try:
                latitude = float(getattr(row, "latitude"))
                longitude = float(getattr(row, "longitude"))
            except (TypeError, ValueError):
                continue
            GEOCODE_MEMORY_CACHE[address.casefold()] = GeocodeResult(address, latitude, longitude, "OneMap cache")
    GEOCODE_DISK_CACHE_LOADED = True


def _load_route_disk_cache_once() -> None:
    global ROUTE_DISK_CACHE_LOADED
    if ROUTE_DISK_CACHE_LOADED:
        return
    for cache_file in [ROUTE_SEED_CACHE_FILE, ROUTE_CACHE_FILE]:
        cache = _read_csv_cache(cache_file, ["route_key", "distance_km", "duration_min", "route_text", "route_path"])
        for row in cache.itertuples(index=False):
            key = clean_text(getattr(row, "route_key", ""))
            if not key:
                continue
            route_path = None
            try:
                route_path = json.loads(getattr(row, "route_path", "") or "null")
            except (TypeError, json.JSONDecodeError):
                route_path = None
            try:
                distance_km = float(getattr(row, "distance_km"))
                duration_min = float(getattr(row, "duration_min"))
            except (TypeError, ValueError):
                continue
            ROUTE_MEMORY_CACHE[key] = TravelCost(
                distance_km,
                duration_min,
                "OneMap cache",
                route_text=clean_text(getattr(row, "route_text", "")),
                route_path=route_path,
            )
    ROUTE_DISK_CACHE_LOADED = True


def load_env_value(name: str, path: Path = ENV_FILE) -> str:
    value = os.getenv(name, "")
    if value:
        return value
    if not path.exists():
        return ""

    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except OSError:
        return ""

    for line in lines:
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, raw_value = stripped.split("=", 1)
        if key.strip() == name:
            return raw_value.strip().strip('"').strip("'")
    return ""


def get_onemap_token() -> str:
    return load_env_value("ONEMAP_TOKEN")


def _get_streamlit_session_value(key: str, default: Any = None) -> Any:
    if not _streamlit_context_available():
        return default
    try:
        return st.session_state.get(key, default)
    except Exception:
        return default


def _set_streamlit_session_value(key: str, value: Any) -> None:
    if not _streamlit_context_available():
        return
    try:
        st.session_state[key] = value
    except Exception:
        return


def _streamlit_context_available() -> bool:
    if st is None:
        return False
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx

        return get_script_run_ctx() is not None
    except Exception:
        return False


def _get_onemap_secret(name: str) -> str:
    if _streamlit_context_available():
        try:
            value = st.secrets.get(name, "")
        except Exception:
            value = ""
        if value:
            return clean_text(value)
    return load_env_value(name)


def _parse_onemap_token_expiry(payload: dict[str, Any]) -> datetime | None:
    expires_in = payload.get("expires_in") or payload.get("expiresIn")
    if expires_in is not None:
        try:
            return datetime.now(timezone.utc) + timedelta(seconds=float(expires_in))
        except (TypeError, ValueError):
            pass

    expiry_value = (
        payload.get("expiry_timestamp")
        or payload.get("expiryTimestamp")
        or payload.get("expires_at")
        or payload.get("expiresAt")
        or payload.get("expiry")
    )
    if expiry_value is None:
        return None
    try:
        if isinstance(expiry_value, (int, float)):
            timestamp = float(expiry_value)
            if timestamp > 10_000_000_000:
                timestamp /= 1000
            return datetime.fromtimestamp(timestamp, tz=timezone.utc)
        expiry_text = clean_text(expiry_value)
        if expiry_text.isdigit():
            timestamp = float(expiry_text)
            if timestamp > 10_000_000_000:
                timestamp /= 1000
            return datetime.fromtimestamp(timestamp, tz=timezone.utc)
        parsed = pd.to_datetime(expiry_text, errors="coerce", utc=True)
        if pd.isna(parsed):
            return None
        return parsed.to_pydatetime()
    except (OSError, OverflowError, TypeError, ValueError):
        return None


def _cached_onemap_token_is_valid(expiry: datetime | None) -> bool:
    if expiry is None:
        return True
    if expiry.tzinfo is None:
        expiry = expiry.replace(tzinfo=timezone.utc)
    return expiry > datetime.now(timezone.utc) + timedelta(hours=1)


def _store_active_onemap_token(token: str, expiry: datetime | None = None) -> None:
    global ONEMAP_MEMORY_TOKEN, ONEMAP_MEMORY_TOKEN_EXPIRY
    ONEMAP_MEMORY_TOKEN = token
    ONEMAP_MEMORY_TOKEN_EXPIRY = expiry
    _set_streamlit_session_value("onemap_token", token)
    _set_streamlit_session_value("onemap_token_expiry", expiry)


def request_new_onemap_token() -> tuple[str, datetime | None]:
    email = _get_onemap_secret("ONEMAP_EMAIL")
    password = _get_onemap_secret("ONEMAP_PASSWORD")
    if not email or not password:
        raise RuntimeError("OneMap token refresh failed: ONEMAP_EMAIL and ONEMAP_PASSWORD are not configured.")

    payload = json.dumps({"email": email, "password": password}).encode("utf-8")
    request = Request(
        ONEMAP_AUTH_URL,
        data=payload,
        headers={
            "User-Agent": "Lance-BlueSG-Route-Optimiser/1.0",
            "Content-Type": "application/json",
            "Accept": "application/json",
        },
        method="POST",
    )
    with urlopen(request, timeout=15) as response:
        response_payload = json.loads(response.read().decode("utf-8"))

    access_token = clean_text(
        response_payload.get("access_token")
        or response_payload.get("accessToken")
        or response_payload.get("token")
    )
    if not access_token:
        raise RuntimeError("OneMap token refresh failed: authentication response did not include an access token.")

    return access_token, _parse_onemap_token_expiry(response_payload)


def get_active_onemap_token(force_refresh: bool = False, manual_token: str | None = None) -> str | None:
    session_token = clean_text(_get_streamlit_session_value("onemap_token", ""))
    session_expiry = _get_streamlit_session_value("onemap_token_expiry")
    if isinstance(session_expiry, str):
        parsed = pd.to_datetime(session_expiry, errors="coerce", utc=True)
        session_expiry = None if pd.isna(parsed) else parsed.to_pydatetime()

    if not force_refresh and session_token and _cached_onemap_token_is_valid(session_expiry):
        return session_token

    global ONEMAP_MEMORY_TOKEN, ONEMAP_MEMORY_TOKEN_EXPIRY
    if not force_refresh and ONEMAP_MEMORY_TOKEN and _cached_onemap_token_is_valid(ONEMAP_MEMORY_TOKEN_EXPIRY):
        return ONEMAP_MEMORY_TOKEN

    if not force_refresh:
        initial_token = clean_text(manual_token) or get_onemap_token()
        if initial_token:
            return initial_token

    new_token, expiry = request_new_onemap_token()
    _store_active_onemap_token(new_token, expiry)
    return new_token


def onemap_credentials_configured() -> bool:
    return bool(_get_onemap_secret("ONEMAP_EMAIL") and _get_onemap_secret("ONEMAP_PASSWORD"))


def infer_zone(address: Any) -> str | None:
    text = clean_text(address).lower()
    if not text:
        return None

    for zone, keywords in ZONE_KEYWORDS.items():
        if any(keyword.lower() in text for keyword in keywords):
            return zone
    return None


def clean_address_for_geocoding(address: Any) -> str:
    text = clean_text(address)
    if not text:
        return ""

    lowered = text.lower()
    named_location_aliases = {
        "lyf bugis hotel": "lyf Bugis Singapore",
        "lyf bugis": "lyf Bugis Singapore",
        "suntec carpark a": "Suntec City",
        "suntec car park a": "Suntec City",
        "suntec carpark": "Suntec City",
        "suntec city carpark": "Suntec City",
        "plq entrance park place residences": "Park Place Residences PLQ",
        "park place residences": "Park Place Residences PLQ",
        "paya lebar quarter": "Paya Lebar Quarter",
        "plq mall": "PLQ Mall",
    }
    for keyword, alias in named_location_aliases.items():
        if keyword in lowered:
            return alias

    text = re.sub(r"\[[^\]]*\]", " ", text)
    text = re.sub(r",?\s+\bL\d+[A-Z]?\b.*$", " ", text, flags=re.IGNORECASE)
    text = re.sub(r",?\s+\bLevel\s+\d+[A-Z]?\b.*$", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\b(MSCP|Surface|Basement)\b.*$", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\bDeck\b.*$", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\bLot\b.*$", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\bNear\b.*$", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" ,")


def _normalise_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean_text(value).lower())


JOB_HEADER_ALIASES = {
    "Date": {"date"},
    "Car Plate": {"carplate", "carplates", "carno", "carnumber", "vehicleplate", "vehicle"},
    "Fuel %": {"fuel", "fuelpercent", "fuelpercentage"},
    "Pickup Address": {"pickupaddress", "pickuplocation", "pickuplocationaddress", "startlocation"},
    "Pickup Lot": {"pickuplot", "lot", "lots", "lotnumber", "lotsnumber", "pickuplotsnumber"},
    "Drop-off Address": {"dropoffaddress", "dropofflocation", "endlocation", "destination"},
    "Pickup Time": {"pickuptime", "starttime"},
    "Notes": {"notes", "note", "remarks", "remark"},
}


def _canonical_job_header(value: Any) -> str | None:
    normalised = _normalise_header(value)
    for canonical, aliases in JOB_HEADER_ALIASES.items():
        if normalised in aliases:
            return canonical
    return None


def _find_job_header_row(raw: pd.DataFrame) -> int:
    best_row = 0
    best_score = -1
    search_rows = min(20, len(raw))
    for row_index in range(search_rows):
        found = {
            canonical
            for value in raw.iloc[row_index].tolist()
            if (canonical := _canonical_job_header(value)) is not None
        }
        score = sum(header in found for header in REQUIRED_JOB_HEADERS)
        if score > best_score:
            best_row = row_index
            best_score = score
        if score == len(REQUIRED_JOB_HEADERS):
            return row_index
    return best_row


def _coerce_job_date(value: Any, swap_month_day: bool = False) -> Any:
    if pd.isna(value):
        return value

    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return value

    timestamp = pd.Timestamp(parsed)
    # Some supplier sheets are typed as m/d but used operationally as d/m.
    # Example: displayed/input "04/07" should mean 4 July, not 7 April.
    if swap_month_day and timestamp.day <= 12:
        timestamp = pd.Timestamp(year=timestamp.year, month=timestamp.day, day=timestamp.month)
    return timestamp.normalize()


def load_jobs_from_excel(uploaded_file: Any) -> pd.DataFrame:
    try:
        if hasattr(uploaded_file, "seek"):
            uploaded_file.seek(0)
        raw = pd.read_excel(uploaded_file, header=None)
    except Exception as exc:
        raise ValueError(f"Unable to read the Excel file: {exc}") from exc

    if raw.empty:
        return pd.DataFrame()

    header_row = _find_job_header_row(raw)
    headers = raw.iloc[header_row].tolist()
    data = raw.iloc[header_row + 1 :].copy()
    uses_supplier_day_month_dates = any(
        _normalise_header(header) in {"carplate", "pickuplocation", "endlocation"}
        for header in headers
    )

    output = pd.DataFrame(index=data.index)
    seen_lots_number = 0
    for column_index, header in enumerate(headers):
        canonical = _canonical_job_header(header)
        normalised = _normalise_header(header)
        if normalised in {"lotnumber", "lotsnumber"}:
            seen_lots_number += 1
            canonical = "Pickup Lot" if seen_lots_number == 1 else None
        if canonical is None or canonical in output.columns:
            continue
        output[canonical] = data.iloc[:, column_index]

    extra_notes = []
    for column_name in ["Driver", "Contact"]:
        matching_indexes = [
            index for index, header in enumerate(headers) if _normalise_header(header) == _normalise_header(column_name)
        ]
        if matching_indexes:
            extra_notes.append(data.iloc[:, matching_indexes[0]].apply(clean_text))
    if extra_notes:
        combined = extra_notes[0]
        for note_series in extra_notes[1:]:
            combined = combined.str.cat(note_series, sep=" ", na_rep="").str.strip()
        if "Notes" in output.columns:
            output["Notes"] = output["Notes"].apply(clean_text).str.cat(combined, sep=" ", na_rep="").str.strip()
        else:
            output["Notes"] = combined

    if "Date" in output.columns:
        output["Date"] = output["Date"].apply(
            lambda value: _coerce_job_date(value, swap_month_day=uses_supplier_day_month_dates)
        )

    output["_uploaded_row"] = data.index.astype(int) + 1
    jobs = output.dropna(how="all").reset_index(drop=True)
    jobs.columns = [clean_text(column) for column in jobs.columns]
    return jobs


def validate_jobs(jobs: pd.DataFrame) -> tuple[pd.DataFrame, list[str], list[str]]:
    warnings: list[str] = []
    uploaded_count = len(jobs)
    missing_headers = [header for header in REQUIRED_JOB_HEADERS if header not in jobs.columns]
    if missing_headers:
        return pd.DataFrame(), missing_headers, warnings

    keep_columns = REQUIRED_JOB_HEADERS + [header for header in OPTIONAL_JOB_HEADERS if header in jobs.columns]
    if "_uploaded_row" in jobs.columns:
        keep_columns.append("_uploaded_row")
    jobs = jobs.loc[:, keep_columns].copy()
    if "_uploaded_row" in jobs.columns:
        jobs["_original_order"] = pd.to_numeric(jobs["_uploaded_row"], errors="coerce").fillna(2).astype(int) - 2
    else:
        jobs["_original_order"] = range(len(jobs))

    before_drop = len(jobs)
    jobs = jobs[
        jobs["Pickup Address"].apply(clean_text).ne("")
        & jobs["Drop-off Address"].apply(clean_text).ne("")
    ].copy()
    dropped = before_drop - len(jobs)
    if dropped:
        warnings.append(f"Dropped {dropped} row(s) with blank pickup or drop-off address.")

    duplicate_plates = jobs["Car Plate"].apply(clean_text).loc[lambda series: series.ne("")].duplicated(keep=False)
    duplicate_plate_values = sorted(
        jobs["Car Plate"]
        .apply(clean_text)
        .loc[lambda series: series.ne("") & series.duplicated(keep=False)]
        .unique()
    )
    if duplicate_plates.any():
        warnings.append("Duplicate car plate values found. They are allowed, but please verify them.")

    jobs["Pickup Zone"] = jobs["Pickup Address"].apply(infer_zone)
    jobs["Drop-off Zone"] = jobs["Drop-off Address"].apply(infer_zone)
    jobs = jobs.reset_index(drop=True)
    jobs.attrs["uploaded_count"] = uploaded_count
    jobs.attrs["blank_address_rows_dropped"] = dropped
    jobs.attrs["duplicate_plate_values"] = duplicate_plate_values
    return jobs, [], warnings


def load_and_validate_jobs(uploaded_file: Any) -> tuple[pd.DataFrame, list[str], list[str]]:
    return validate_jobs(load_jobs_from_excel(uploaded_file))


def _fetch_json(url: str, params: dict[str, Any], token: str | None = None, timeout: int = 15) -> dict[str, Any]:
    def build_request(active_token: str | None) -> Request:
        headers = {"User-Agent": "Lance-BlueSG-Route-Optimiser/1.0"}
        if active_token:
            headers["Authorization"] = f"Bearer {active_token}"
        return Request(f"{url}?{urlencode(params)}", headers=headers)

    active_token = get_active_onemap_token(manual_token=token)
    try:
        with urlopen(build_request(active_token), timeout=timeout) as response:
            return json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        if exc.code != 401 or not active_token:
            raise

        print("OneMap token rejected. Requesting a fresh token.")
        try:
            refreshed_token = get_active_onemap_token(force_refresh=True)
        except Exception as refresh_exc:
            print(f"OneMap token refresh failed: {refresh_exc}")
            raise exc
        if not refreshed_token:
            print("OneMap token refresh failed: no token returned.")
            raise exc

        try:
            with urlopen(build_request(refreshed_token), timeout=timeout) as response:
                payload = json.loads(response.read().decode("utf-8"))
            print("OneMap request succeeded after token refresh.")
            return payload
        except HTTPError as retry_exc:
            print(f"OneMap request failed after token refresh: HTTP {retry_exc.code}.")
            raise retry_exc
        except Exception as retry_exc:
            print(f"OneMap request failed after token refresh: {retry_exc}")
            raise


def geocode_address_onemap(address: str, token: str | None = None) -> GeocodeResult:
    address = clean_text(address)
    if not address:
        return GeocodeResult(address, None, None, "fallback estimate", "Blank address")

    search_values = [address]
    cleaned_address = clean_address_for_geocoding(address)
    if cleaned_address and cleaned_address.casefold() != address.casefold():
        search_values.append(cleaned_address)

    last_error = "No OneMap geocoding result"
    for search_value in search_values:
        try:
            payload = _fetch_json(
                ONEMAP_SEARCH_URL,
                {"searchVal": search_value, "returnGeom": "Y", "getAddrDetails": "Y", "pageNum": 1},
                token=token,
            )
        except (HTTPError, URLError, TimeoutError, json.JSONDecodeError, OSError) as exc:
            last_error = f"OneMap geocoding failed: {exc}"
            continue

        results = payload.get("results") or []
        if not results:
            last_error = "No OneMap geocoding result"
            continue

        best = results[0]
        try:
            latitude = float(best["LATITUDE"])
            longitude = float(best["LONGITUDE"])
        except (KeyError, TypeError, ValueError) as exc:
            last_error = f"Invalid OneMap geocode result: {exc}"
            continue

        source = "OneMap" if search_value == address else "OneMap cleaned address"
        return GeocodeResult(address, latitude, longitude, source)

    return GeocodeResult(address, None, None, "fallback estimate", last_error)


def get_cached_geocode(address: str, token: str | None = None, use_onemap: bool = True) -> GeocodeResult:
    address = clean_text(address)
    if not use_onemap:
        return GeocodeResult(address, None, None, "fallback estimate", "OneMap disabled")

    _load_geocode_disk_cache_once()
    memory_key = address.casefold()
    if memory_key in GEOCODE_MEMORY_CACHE:
        cached = GEOCODE_MEMORY_CACHE[memory_key]
        if cached.is_available:
            return GeocodeResult(address, cached.latitude, cached.longitude, "OneMap cache", cached.error)

    result = geocode_address_onemap(address, token=token)
    GEOCODE_MEMORY_CACHE[memory_key] = result
    if result.is_available:
        _append_csv_cache_row(
            GEOCODE_CACHE_FILE,
            {"address": address, "latitude": result.latitude, "longitude": result.longitude},
            ["address", "latitude", "longitude"],
        )
    return result


def _parse_onemap_route(payload: dict[str, Any]) -> tuple[float | None, float | None]:
    summary = payload.get("route_summary") or payload.get("routeSummary") or {}
    distance_m = (
        summary.get("total_distance")
        or summary.get("totalDistance")
        or payload.get("total_distance")
        or payload.get("distance")
    )
    duration_s = (
        summary.get("total_time")
        or summary.get("totalTime")
        or payload.get("total_time")
        or payload.get("duration")
    )

    if distance_m is None or duration_s is None:
        itineraries = (payload.get("plan") or {}).get("itineraries") or payload.get("itineraries") or []
        if itineraries:
            itinerary = itineraries[0]
            legs = itinerary.get("legs") or []
            if duration_s is None:
                duration_s = itinerary.get("duration") or sum(float(leg.get("duration") or 0) for leg in legs)
            if distance_m is None:
                distance_m = itinerary.get("walkDistance") or sum(float(leg.get("distance") or 0) for leg in legs)

    distance_km = None
    duration_min = None
    try:
        if distance_m is not None:
            distance_km = round(float(distance_m) / 1000, 2)
    except (TypeError, ValueError):
        distance_km = None
    try:
        if duration_s is not None:
            duration_min = round(float(duration_s) / 60, 1)
    except (TypeError, ValueError):
        duration_min = None
    return distance_km, duration_min


def _decode_polyline(polyline: str) -> list[list[float]]:
    coordinates = []
    index = 0
    lat = 0
    lng = 0

    while index < len(polyline):
        shift = 0
        result = 0
        while index < len(polyline):
            byte = ord(polyline[index]) - 63
            index += 1
            result |= (byte & 0x1F) << shift
            shift += 5
            if byte < 0x20:
                break
        lat += ~(result >> 1) if result & 1 else result >> 1

        shift = 0
        result = 0
        while index < len(polyline):
            byte = ord(polyline[index]) - 63
            index += 1
            result |= (byte & 0x1F) << shift
            shift += 5
            if byte < 0x20:
                break
        lng += ~(result >> 1) if result & 1 else result >> 1
        coordinates.append([lng / 1e5, lat / 1e5])

    return coordinates


def _extract_points(value: Any) -> list[list[float]]:
    if not value:
        return []
    if isinstance(value, str):
        try:
            return _decode_polyline(value)
        except Exception:
            return []
    if isinstance(value, dict):
        return _extract_points(value.get("points") or value.get("geometry") or value.get("coordinates"))
    if isinstance(value, list):
        points = []
        for item in value:
            if isinstance(item, dict):
                lat = item.get("lat") or item.get("latitude") or item.get("LATITUDE")
                lon = item.get("lon") or item.get("lng") or item.get("longitude") or item.get("LONGITUDE")
                if lat is not None and lon is not None:
                    points.append([float(lon), float(lat)])
            elif isinstance(item, (list, tuple)) and len(item) >= 2:
                first = float(item[0])
                second = float(item[1])
                points.append([second, first] if abs(first) <= 2 and abs(second) > 90 else [first, second])
        return points
    return []


def _format_leg_instruction(leg: dict[str, Any]) -> str:
    mode = clean_text(leg.get("mode") or leg.get("transportMode") or leg.get("type")).upper()
    route = clean_text(
        leg.get("route")
        or leg.get("routeShortName")
        or leg.get("routeLongName")
        or leg.get("headsign")
        or leg.get("name")
    )
    from_name = clean_text((leg.get("from") or {}).get("name") if isinstance(leg.get("from"), dict) else leg.get("from"))
    to_name = clean_text((leg.get("to") or {}).get("name") if isinstance(leg.get("to"), dict) else leg.get("to"))
    duration = leg.get("duration")
    duration_text = ""
    try:
        duration_text = f" ({round(float(duration) / 60, 1)} min)" if duration is not None else ""
    except (TypeError, ValueError):
        duration_text = ""

    if mode in {"WALK", "WALKING"}:
        return f"Walk from {from_name or 'start'} to {to_name or 'next stop'}{duration_text}"
    if route:
        return f"Take {mode or 'PT'} {route} from {from_name or 'start'} to {to_name or 'destination'}{duration_text}"
    return f"Take {mode or 'PT'} from {from_name or 'start'} to {to_name or 'destination'}{duration_text}"


def _parse_route_details(payload: dict[str, Any], route_type: str) -> tuple[str, list[list[float]]]:
    if route_type == "drive":
        instructions = []
        for instruction in payload.get("route_instructions") or payload.get("routeInstructions") or []:
            if isinstance(instruction, list) and instruction:
                instructions.append(clean_text(instruction[-1] or instruction[0]))
            elif isinstance(instruction, dict):
                instructions.append(clean_text(instruction.get("instruction") or instruction.get("text")))
        route_text = " -> ".join([instruction for instruction in instructions if instruction])
        route_path = _extract_points(payload.get("route_geometry") or payload.get("routeGeometry"))
        return route_text, route_path

    itineraries = (payload.get("plan") or {}).get("itineraries") or payload.get("itineraries") or []
    if not itineraries:
        return "", []

    legs = itineraries[0].get("legs") or []
    route_text = " -> ".join([_format_leg_instruction(leg) for leg in legs if isinstance(leg, dict)])
    route_path: list[list[float]] = []
    for leg in legs:
        if not isinstance(leg, dict):
            continue
        route_path.extend(
            _extract_points(
                leg.get("legGeometry")
                or leg.get("geometry")
                or leg.get("points")
                or leg.get("routeGeometry")
            )
        )
    return route_text, route_path


def _route_source_label(route_type: str, mode: str | None = None, cached: bool = False) -> str:
    if route_type == "drive":
        return "OneMap cache" if cached else "OneMap"
    if route_type == "pt":
        suffix = " cache" if cached else ""
        return f"OneMap public transport {mode or 'TRANSIT'}{suffix}"
    suffix = " cache" if cached else ""
    return f"OneMap {route_type}{suffix}"


def _public_transport_route_time() -> tuple[str, str]:
    now = datetime.now(SINGAPORE_TZ)
    route_time = now.replace(hour=14, minute=0, second=0, microsecond=0)
    return route_time.strftime("%m-%d-%Y"), route_time.strftime("%H:%M:%S")


def _straight_line_distance_km(
    from_geocode: GeocodeResult,
    to_geocode: GeocodeResult,
) -> float | None:
    if not from_geocode.is_available or not to_geocode.is_available:
        return None
    lat1 = math.radians(float(from_geocode.latitude))
    lat2 = math.radians(float(to_geocode.latitude))
    delta_lat = lat2 - lat1
    delta_lon = math.radians(float(to_geocode.longitude) - float(from_geocode.longitude))
    a = (
        math.sin(delta_lat / 2) ** 2
        + math.cos(lat1) * math.cos(lat2) * math.sin(delta_lon / 2) ** 2
    )
    return 6371.0 * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def get_onemap_route_cost(
    from_geocode: GeocodeResult,
    to_geocode: GeocodeResult,
    token: str | None = None,
    route_type: str = "drive",
    mode: str | None = None,
    max_walk_distance: int = 1000,
) -> TravelCost:
    if not from_geocode.is_available or not to_geocode.is_available:
        missing = from_geocode.error or to_geocode.error or "Missing coordinates"
        return TravelCost(None, None, "fallback estimate", missing)

    route_type = route_type.casefold()
    mode = mode.upper() if mode else None
    route_date = ""
    route_time = ""
    if route_type == "pt":
        route_date, route_time = _public_transport_route_time()

    coordinate_key = (
        f"{from_geocode.latitude:.6f},{from_geocode.longitude:.6f}|"
        f"{to_geocode.latitude:.6f},{to_geocode.longitude:.6f}"
    )
    mode_key = mode or ("TRANSIT" if route_type == "pt" else route_type.upper())
    key = coordinate_key if route_type == "drive" else f"{route_type}|{mode_key}|{route_date}|{route_time}|{coordinate_key}"
    _load_route_disk_cache_once()
    if key in ROUTE_MEMORY_CACHE:
        cached = ROUTE_MEMORY_CACHE[key]
        return TravelCost(
            cached.distance_km,
            cached.duration_min,
            _route_source_label(route_type, mode, cached=True),
            cached.error,
            cached.route_text,
            cached.route_path,
        )

    params = {
        "start": f"{from_geocode.latitude},{from_geocode.longitude}",
        "end": f"{to_geocode.latitude},{to_geocode.longitude}",
        "routeType": route_type,
    }
    if route_type == "pt":
        params.update(
            {
                "date": route_date,
                "time": route_time,
                "mode": mode or "TRANSIT",
                "maxWalkDistance": max_walk_distance,
            }
        )

    last_error: Exception | None = None
    payload: dict[str, Any] | None = None
    for _ in range(2):
        try:
            payload = _fetch_json(ONEMAP_ROUTE_URL, params, token=token, timeout=8)
            break
        except (HTTPError, URLError, TimeoutError, json.JSONDecodeError, OSError) as exc:
            last_error = exc
    if payload is None:
        return TravelCost(None, None, "fallback estimate", f"OneMap routing failed after 2 attempts: {last_error}")

    distance_km, duration_min = _parse_onemap_route(payload)
    if distance_km is None and duration_min is None:
        return TravelCost(None, None, "fallback estimate", "OneMap route returned no distance or duration")

    route_text, route_path = _parse_route_details(payload, route_type)
    source = _route_source_label(route_type, mode)
    result = TravelCost(distance_km, duration_min, source, route_text=route_text, route_path=route_path or None)
    ROUTE_MEMORY_CACHE[key] = result
    _append_csv_cache_row(
        ROUTE_CACHE_FILE,
        {
            "route_key": key,
            "distance_km": distance_km or 0,
            "duration_min": duration_min or 0,
            "route_text": route_text,
            "route_path": json.dumps(route_path or []),
        },
        ["route_key", "distance_km", "duration_min", "route_text", "route_path"],
    )
    return result


def get_fallback_cost(
    from_address: Any,
    to_address: Any,
    from_zone: str | None = None,
    to_zone: str | None = None,
) -> TravelCost:
    from_text = clean_text(from_address)
    to_text = clean_text(to_address)
    if not from_text or not to_text:
        return TravelCost(UNKNOWN_DISTANCE_KM, UNKNOWN_DURATION_MIN, "fallback estimate", "Blank address")
    if from_text.casefold() == to_text.casefold():
        return TravelCost(0, 0, "fallback estimate")

    source_zone = from_zone or infer_zone(from_text)
    destination_zone = to_zone or infer_zone(to_text)
    if source_zone and destination_zone:
        distance = FALLBACK_ZONE_KM.get(source_zone, {}).get(destination_zone, UNKNOWN_DISTANCE_KM)
        duration = FALLBACK_ZONE_MINUTES.get(source_zone, {}).get(destination_zone, UNKNOWN_DURATION_MIN)
        return TravelCost(float(distance), float(duration), "fallback estimate")
    if source_zone == destination_zone and source_zone is not None:
        return TravelCost(SAME_UNKNOWN_DISTANCE_KM, SAME_UNKNOWN_DURATION_MIN, "fallback estimate")
    return TravelCost(UNKNOWN_DISTANCE_KM, UNKNOWN_DURATION_MIN, "fallback estimate", "Unknown zone")


def get_travel_cost(
    from_address: Any,
    to_address: Any,
    from_zone: str | None = None,
    to_zone: str | None = None,
    use_onemap: bool = True,
    token: str | None = None,
    route_type: str = "drive",
    mode: str | None = None,
) -> TravelCost:
    if clean_text(from_address).casefold() == clean_text(to_address).casefold():
        return TravelCost(0, 0, "OneMap cache" if use_onemap else "fallback estimate")

    if use_onemap:
        from_geocode = get_cached_geocode(clean_text(from_address), token=token, use_onemap=True)
        to_geocode = get_cached_geocode(clean_text(to_address), token=token, use_onemap=True)
        onemap_cost = get_onemap_route_cost(
            from_geocode,
            to_geocode,
            token=token,
            route_type=route_type,
            mode=mode,
        )
        if onemap_cost.distance_km is not None or onemap_cost.duration_min is not None:
            return onemap_cost
        fallback_cost = get_fallback_cost(from_address, to_address, from_zone=from_zone, to_zone=to_zone)
        return TravelCost(
            fallback_cost.distance_km,
            fallback_cost.duration_min,
            "fallback estimate",
            onemap_cost.error or fallback_cost.error,
        )

    return get_fallback_cost(from_address, to_address, from_zone=from_zone, to_zone=to_zone)


def get_empty_travel_cost(
    from_address: Any,
    to_address: Any,
    from_zone: str | None = None,
    to_zone: str | None = None,
    use_onemap: bool = True,
    token: str | None = None,
    allow_walk: bool = False,
) -> TravelCost:
    if not use_onemap:
        return get_travel_cost(
            from_address,
            to_address,
            from_zone=from_zone,
            to_zone=to_zone,
            use_onemap=False,
            token=token,
        )

    from_geocode = get_cached_geocode(clean_text(from_address), token=token, use_onemap=True)
    to_geocode = get_cached_geocode(clean_text(to_address), token=token, use_onemap=True)
    straight_line_km = _straight_line_distance_km(from_geocode, to_geocode)

    if allow_walk and straight_line_km is not None and straight_line_km <= SHORT_WALK_DISTANCE_KM:
        walk_cost = get_travel_cost(
            from_address,
            to_address,
            from_zone=from_zone,
            to_zone=to_zone,
            use_onemap=True,
            token=token,
            route_type="walk",
        )
        if (
            (walk_cost.distance_km is not None and walk_cost.distance_km <= SHORT_WALK_DISTANCE_KM)
            or (walk_cost.duration_min is not None and walk_cost.duration_min <= SHORT_WALK_DURATION_MIN)
        ):
            return TravelCost(walk_cost.distance_km, walk_cost.duration_min, "OneMap walk", walk_cost.error)

    pt_cost = get_onemap_route_cost(
        from_geocode,
        to_geocode,
        token=token,
        route_type="pt",
        mode="TRANSIT",
    )
    if pt_cost.duration_min is not None or pt_cost.distance_km is not None:
        return pt_cost

    geocode_error = from_geocode.error or to_geocode.error
    if geocode_error:
        fallback_cost = get_fallback_cost(from_address, to_address, from_zone=from_zone, to_zone=to_zone)
        return TravelCost(
            fallback_cost.distance_km,
            fallback_cost.duration_min,
            "fallback estimate",
            geocode_error or fallback_cost.error,
        )

    return get_travel_cost(
        from_address,
        to_address,
        from_zone=from_zone,
        to_zone=to_zone,
        use_onemap=True,
        token=token,
        route_type="drive",
    )


def get_cost_explanation() -> pd.DataFrame:
    rows = []
    for from_zone, destinations in FALLBACK_ZONE_MINUTES.items():
        for to_zone, duration in destinations.items():
            rows.append(
                {
                    "From Zone": from_zone,
                    "To Zone": to_zone,
                    "Fallback Distance KM": FALLBACK_ZONE_KM[from_zone][to_zone],
                    "Fallback Duration Min": duration,
                }
            )
    return pd.DataFrame(rows)


def calculate_route_zone_priority(
    rider_current_zone: str | None,
    pickup_zone: str | None,
    dropoff_zone: str | None,
) -> tuple[int, bool, bool]:
    same_zone_pickup = bool(rider_current_zone and pickup_zone and rider_current_zone == pickup_zone)
    same_zone_route = bool(pickup_zone and dropoff_zone and pickup_zone == dropoff_zone)
    stays_in_rider_zone = bool(
        rider_current_zone
        and pickup_zone
        and dropoff_zone
        and rider_current_zone == pickup_zone == dropoff_zone
    )
    if stays_in_rider_zone:
        return 0, same_zone_pickup, True
    if same_zone_route:
        return 1, same_zone_pickup, False
    return 2, same_zone_pickup, False


def calculate_zone_adjustment(
    rider_current_zone: str | None,
    pickup_zone: str | None,
    dropoff_zone: str | None,
) -> float:
    """
    Zone adjustment is strong but not absolute.

    Negative value = preferred. Positive value = penalty.
    Same-zone work is encouraged, but a clearly shorter/easier cross-zone
    movement can still win through the overall assignment score.
    """
    stays_in_rider_zone = bool(
        rider_current_zone
        and pickup_zone
        and dropoff_zone
        and rider_current_zone == pickup_zone == dropoff_zone
    )
    same_zone_pickup = bool(rider_current_zone and pickup_zone and rider_current_zone == pickup_zone)
    same_zone_route = bool(pickup_zone and dropoff_zone and pickup_zone == dropoff_zone)

    if stays_in_rider_zone:
        return -20.0
    if same_zone_pickup and not same_zone_route:
        return 20.0
    if not same_zone_pickup and same_zone_route:
        return 40.0
    if not same_zone_route:
        return 60.0
    return 0.0


def calculate_assignment_score(
    empty_distance_km: float | None,
    empty_duration_min: float | None,
    loaded_distance_km: float | None,
    loaded_duration_min: float | None,
    rider_current_zone: str | None,
    pickup_zone: str | None,
    dropoff_zone: str | None,
    rider_total_duration_min: float,
    rider_assigned_jobs: int,
    rider_max_jobs: int | None,
    rider_load_level: str = "Medium",
    optimise_by: str = "duration",
    empty_weight: float = DEFAULT_EMPTY_WEIGHT,
    loaded_weight: float = DEFAULT_LOADED_WEIGHT,
    soft_workload_min: float = DEFAULT_SOFT_WORKLOAD_MIN,
    workload_penalty_per_min: float = DEFAULT_WORKLOAD_PENALTY_PER_MIN,
    soft_adjusted_duration_min: float = DEFAULT_SOFT_ADJUSTED_DURATION_MIN,
    duration_buffer_multiplier: float = DEFAULT_DURATION_BUFFER_MULTIPLIER,
    duration_penalty_per_min: float = DEFAULT_DURATION_PENALTY_PER_MIN,
    max_jobs_overage_penalty: float = DEFAULT_MAX_JOB_OVERAGE_PENALTY,
    max_adjusted_duration_min: float = DEFAULT_MAX_ADJUSTED_DURATION_MIN,
) -> dict[str, float | int] | None:
    empty_distance = empty_distance_km if empty_distance_km is not None else UNKNOWN_DISTANCE_KM
    empty_duration = empty_duration_min if empty_duration_min is not None else UNKNOWN_DURATION_MIN
    loaded_distance = loaded_distance_km if loaded_distance_km is not None else UNKNOWN_DISTANCE_KM
    loaded_duration = loaded_duration_min if loaded_duration_min is not None else UNKNOWN_DURATION_MIN

    if optimise_by.casefold() == "distance":
        movement_score = (empty_distance * empty_weight) + (loaded_distance * loaded_weight)
    else:
        movement_score = (empty_duration * empty_weight) + (loaded_duration * loaded_weight)

    zone_adjustment = calculate_zone_adjustment(rider_current_zone, pickup_zone, dropoff_zone)
    candidate_total_duration = empty_duration + loaded_duration
    projected_duration = rider_total_duration_min + candidate_total_duration
    projected_adjusted_duration = projected_duration * duration_buffer_multiplier
    if projected_adjusted_duration > max_adjusted_duration_min:
        return None

    workload_penalty = (
        (projected_duration - soft_workload_min) * workload_penalty_per_min
        if projected_duration > soft_workload_min
        else 0.0
    )
    duration_penalty = (
        (projected_adjusted_duration - soft_adjusted_duration_min) * duration_penalty_per_min
        if projected_adjusted_duration > soft_adjusted_duration_min
        else 0.0
    )
    projected_jobs = rider_assigned_jobs + 1
    max_jobs_overage = max(0, projected_jobs - rider_max_jobs) if rider_max_jobs is not None else 0
    max_jobs_penalty = (max_jobs_overage**2) * max_jobs_overage_penalty
    load_level = normalise_rider_load_level(rider_load_level)
    load_policy = RIDER_LOAD_POLICIES.get(load_level, RIDER_LOAD_POLICIES["Medium"])
    same_zone_pickup = bool(rider_current_zone and pickup_zone and rider_current_zone == pickup_zone)
    stays_in_rider_zone = bool(
        rider_current_zone
        and pickup_zone
        and dropoff_zone
        and rider_current_zone == pickup_zone == dropoff_zone
    )
    same_zone_route = bool(pickup_zone and dropoff_zone and pickup_zone == dropoff_zone)
    rider_load_adjustment = (
        float(load_policy["job_score_adjustment"]) * projected_jobs
        + float(load_policy["job_escalation"]) * max(0, projected_jobs - 1) ** 2
    )
    empty_duration_limit = float(load_policy["empty_duration_soft_limit"])
    empty_duration_penalty = max(0.0, empty_duration - empty_duration_limit) * float(
        load_policy["empty_duration_penalty_per_min"]
    )
    load_zone_adjustment = 0.0
    if not same_zone_pickup:
        load_zone_adjustment += float(load_policy["different_pickup_zone_penalty"])
    if not same_zone_route:
        load_zone_adjustment += float(load_policy["cross_zone_route_penalty"])
    if stays_in_rider_zone:
        load_zone_adjustment += float(load_policy["same_area_bonus"])
    assignment_score = (
        movement_score
        + zone_adjustment
        + workload_penalty
        + duration_penalty
        + max_jobs_penalty
        + rider_load_adjustment
        + empty_duration_penalty
        + load_zone_adjustment
    )

    return {
        "assignment_score": round(float(assignment_score), 3),
        "zone_adjustment": float(zone_adjustment),
        "workload_penalty": round(float(workload_penalty), 3),
        "duration_penalty": round(float(duration_penalty), 3),
        "max_jobs_penalty": round(float(max_jobs_penalty), 3),
        "rider_load_adjustment": round(float(rider_load_adjustment), 3),
        "load_empty_duration_penalty": round(float(empty_duration_penalty), 3),
        "load_zone_adjustment": round(float(load_zone_adjustment), 3),
        "projected_duration_min": round(float(projected_duration), 3),
        "projected_adjusted_duration_min": round(float(projected_adjusted_duration), 3),
        "max_jobs_overage": int(max_jobs_overage),
    }


def default_rider_table(count: int = 4) -> pd.DataFrame:
    defaults = [
        {"Rider Name": "Rider A", "Start Location": "Sengkang", "Start Zone": "North-East", "Max Jobs": 5, "Rider Load": "Medium"},
        {"Rider Name": "Rider B", "Start Location": "Punggol", "Start Zone": "North-East", "Max Jobs": 5, "Rider Load": "Medium"},
        {"Rider Name": "Rider C", "Start Location": "Yishun", "Start Zone": "North", "Max Jobs": 5, "Rider Load": "Medium"},
        {"Rider Name": "Rider D", "Start Location": "Tampines", "Start Zone": "East", "Max Jobs": 5, "Rider Load": "Medium"},
    ]
    while len(defaults) < count:
        rider_number = len(defaults) + 1
        defaults.append(
            {
                "Rider Name": f"Rider {rider_number}",
                "Start Location": "",
                "Start Zone": "",
                "Max Jobs": None,
                "Rider Load": "Medium",
            }
        )
    return pd.DataFrame(defaults[:count])


def _normalise_rider_roster(rider_df: pd.DataFrame) -> pd.DataFrame:
    rider_df = rider_df.copy() if rider_df is not None else pd.DataFrame()
    for column in RIDER_COLUMNS:
        if column not in rider_df.columns:
            rider_df[column] = None
    rider_df = rider_df.loc[:, RIDER_COLUMNS].copy()
    rider_df["Rider Name"] = rider_df["Rider Name"].apply(clean_text)
    rider_df["Start Location"] = rider_df["Start Location"].apply(clean_text)
    rider_df["Start Zone"] = rider_df["Start Zone"].apply(clean_text)
    rider_df["Max Jobs"] = rider_df["Max Jobs"].apply(parse_optional_int)
    rider_df["Rider Load"] = rider_df["Rider Load"].apply(normalise_rider_load_level)
    return rider_df


def ensure_rider_roster_workbook(path: Path = ROSTER_FILE) -> Path:
    if path.exists():
        try:
            workbook = pd.ExcelFile(path)
            try:
                existing_sheets = workbook.sheet_names
            finally:
                workbook.close()
        except Exception:
            existing_sheets = []
    else:
        existing_sheets = []

    if path.exists() and all(day in existing_sheets for day in WEEKDAY_SHEETS):
        return path

    sheets: dict[str, pd.DataFrame] = {}
    if path.exists() and existing_sheets:
        try:
            sheets = {
                sheet_name: _normalise_rider_roster(pd.read_excel(path, sheet_name=sheet_name))
                for sheet_name in existing_sheets
            }
        except Exception:
            sheets = {}

    for day in WEEKDAY_SHEETS:
        sheets.setdefault(day, default_rider_table(4))

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for day in WEEKDAY_SHEETS:
            _normalise_rider_roster(sheets[day]).to_excel(writer, sheet_name=day, index=False)
    return path


def load_rider_roster(day: str, path: Path = ROSTER_FILE) -> pd.DataFrame:
    if day not in WEEKDAY_SHEETS:
        raise ValueError(f"Unknown roster day: {day}")
    ensure_rider_roster_workbook(path)
    try:
        return _normalise_rider_roster(pd.read_excel(path, sheet_name=day))
    except ValueError:
        save_rider_roster(day, default_rider_table(4), path)
        return default_rider_table(4)


def save_rider_roster(day: str, rider_df: pd.DataFrame, path: Path = ROSTER_FILE) -> Path:
    if day not in WEEKDAY_SHEETS:
        raise ValueError(f"Unknown roster day: {day}")
    ensure_rider_roster_workbook(path)
    workbook = pd.ExcelFile(path)
    try:
        sheet_names = workbook.sheet_names
    finally:
        workbook.close()
    sheets = {
        sheet_name: _normalise_rider_roster(pd.read_excel(path, sheet_name=sheet_name))
        for sheet_name in sheet_names
        if sheet_name in WEEKDAY_SHEETS
    }
    sheets[day] = _normalise_rider_roster(rider_df)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in WEEKDAY_SHEETS:
            sheets.get(sheet_name, default_rider_table(4)).to_excel(writer, sheet_name=sheet_name, index=False)
    return path


def read_rider_roster_file(path: Path = ROSTER_FILE) -> bytes:
    ensure_rider_roster_workbook(path)
    return path.read_bytes()


def validate_riders(rider_df: pd.DataFrame) -> tuple[list[RiderState], list[str]]:
    errors: list[str] = []
    required_columns = ["Rider Name", "Start Location", "Start Zone", "Max Jobs"]
    for column in required_columns:
        if column not in rider_df.columns:
            errors.append(f"Missing rider column: {column}")

    if errors:
        return [], errors

    riders: list[RiderState] = []
    for index, row in rider_df.iterrows():
        rider_name = clean_text(row.get("Rider Name"))
        start_location = clean_text(row.get("Start Location"))
        if not rider_name and not start_location:
            continue
        if not rider_name or not start_location:
            errors.append(f"Rider row {index + 1} needs both Rider Name and Start Location.")
            continue
        riders.append(RiderState.from_row(row))

    if not riders:
        errors.append("Add at least one rider with a name and start location.")
    return riders, errors


def _combined_source(empty_cost: TravelCost, loaded_cost: TravelCost) -> str:
    if "public transport adjusted" in empty_cost.source.lower():
        base_empty_source = empty_cost.source.replace(", public transport adjusted", "")
        base_source = _combined_source(
            TravelCost(empty_cost.distance_km, empty_cost.duration_min, base_empty_source, empty_cost.error),
            loaded_cost,
        )
        return f"{base_source}, public transport adjusted"

    sources = {empty_cost.source, loaded_cost.source}
    if sources == {"OneMap"}:
        return "OneMap"
    if sources <= {"OneMap", "OneMap cache"}:
        return "OneMap cache" if "OneMap cache" in sources else "OneMap"
    if "fallback estimate" in sources:
        return "fallback estimate"
    return ", ".join(sorted(sources))


def adjust_empty_travel_for_public_transport(
    empty_cost: TravelCost,
    duration_multiplier: float = DEFAULT_EMPTY_TRAVEL_DURATION_MULTIPLIER,
    wait_buffer_min: float = DEFAULT_EMPTY_TRAVEL_WAIT_BUFFER_MIN,
) -> TravelCost:
    if "public transport" in empty_cost.source.lower():
        return empty_cost
    if empty_cost.duration_min is None:
        return empty_cost
    if empty_cost.duration_min <= 0:
        return empty_cost

    adjusted_duration = round((empty_cost.duration_min * duration_multiplier) + wait_buffer_min, 1)
    source = empty_cost.source
    if "public transport adjusted" not in source.lower():
        source = f"{source}, public transport adjusted"
    return TravelCost(empty_cost.distance_km, adjusted_duration, source, empty_cost.error)


def _job_uploaded_row(job: dict[str, Any]) -> int:
    return int(job.get("_original_order", 0)) + 2


def _job_id(job: dict[str, Any]) -> int:
    return _job_uploaded_row(job)


def stable_job_id_from_values(
    uploaded_row: Any,
    car_plate: Any = "",
    pickup_address: Any = "",
    dropoff_address: Any = "",
) -> str:
    """Stable route-editor identifier that survives sequence and rider changes."""
    row_text = clean_text(uploaded_row)
    if row_text:
        try:
            row_text = str(int(float(row_text)))
        except (TypeError, ValueError):
            pass
    parts = [
        row_text or "row-unknown",
        clean_text(car_plate).casefold(),
        clean_text(pickup_address).casefold(),
        clean_text(dropoff_address).casefold(),
    ]
    digest = hashlib.sha1("|".join(parts).encode("utf-8")).hexdigest()[:12]
    return f"job-{row_text or digest}-{digest}"


def stable_job_id_from_job(job: dict[str, Any]) -> str:
    uploaded_row = job.get("Uploaded Row")
    if uploaded_row is None:
        uploaded_row = _job_uploaded_row(job)
    return stable_job_id_from_values(
        uploaded_row,
        job.get("Car Plate"),
        job.get("Pickup Address"),
        job.get("Drop-off Address"),
    )


def stable_job_id_from_route_row(row: pd.Series | dict[str, Any]) -> str:
    getter = row.get
    return stable_job_id_from_values(
        getter("Uploaded Row"),
        getter("Car Plate"),
        getter("Pickup Address"),
        getter("Drop-off Address"),
    )


def build_jobs_by_stable_id(jobs_df: pd.DataFrame | None) -> dict[str, dict[str, Any]]:
    if jobs_df is None or jobs_df.empty:
        return {}
    jobs = jobs_df.copy()
    if "_original_order" not in jobs.columns:
        if "Uploaded Row" in jobs.columns:
            jobs["_original_order"] = pd.to_numeric(jobs["Uploaded Row"], errors="coerce").fillna(2).astype(int) - 2
        else:
            jobs["_original_order"] = range(len(jobs))
    output: dict[str, dict[str, Any]] = {}
    for job in jobs.to_dict("records"):
        output[stable_job_id_from_job(job)] = job
    return output


def build_rider_sequences_from_route_df(route_df: pd.DataFrame) -> dict[str, list[str]]:
    if route_df is None or route_df.empty:
        return {}
    route_df = route_df.copy()
    route_df["_sequence_sort"] = pd.to_numeric(route_df.get("Sequence"), errors="coerce")
    fallback_sequence = pd.Series(range(1, len(route_df) + 1), index=route_df.index)
    route_df["_sequence_sort"] = route_df["_sequence_sort"].fillna(fallback_sequence)
    sequences: dict[str, list[str]] = {}
    for rider, rider_routes in route_df.sort_values(["Rider", "_sequence_sort"], kind="stable").groupby("Rider", sort=False):
        sequences[clean_text(rider)] = [stable_job_id_from_route_row(row) for _, row in rider_routes.iterrows()]
    return sequences


def _route_uploaded_rows(route_df: pd.DataFrame) -> pd.Series:
    if route_df is None or route_df.empty or "Uploaded Row" not in route_df.columns:
        return pd.Series(dtype="Int64")
    return pd.to_numeric(route_df["Uploaded Row"], errors="coerce").dropna().astype(int)


def dedupe_rider_roster(rider_df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    if rider_df is None or rider_df.empty:
        return rider_df, 0
    subset = [column for column in ["Rider Name", "Start Location", "Contact"] if column in rider_df.columns]
    if not subset:
        return rider_df, 0
    before = len(rider_df)
    deduped = rider_df.drop_duplicates(subset=subset, keep="first").reset_index(drop=True)
    return deduped, before - len(deduped)


def _rider_sequence_cache_key(
    rider: RiderState,
    sequence: list[dict[str, Any]],
    settings: dict[str, Any],
) -> tuple[Any, ...]:
    settings_key = tuple(
        (
            key,
            round(float(value), 4) if isinstance(value, (int, float)) and not isinstance(value, bool) else value,
        )
        for key, value in sorted(settings.items())
        if key
        in {
            "use_onemap",
            "optimise_by",
            "empty_weight",
            "loaded_weight",
            "soft_workload_min",
            "workload_penalty_per_min",
            "soft_adjusted_duration_min",
            "duration_penalty_per_min",
            "max_job_overage_penalty",
            "duration_buffer_multiplier",
            "empty_travel_duration_multiplier",
            "empty_travel_wait_buffer_min",
        }
    )
    return (
        rider.name,
        rider.start_location,
        rider.start_zone,
        rider.max_jobs,
        rider.load_level,
        tuple(stable_job_id_from_job(job) for job in sequence),
        settings_key,
    )


def evaluate_explicit_rider_sequence(
    rider: RiderState,
    sequence: list[dict[str, Any]],
    *,
    use_onemap: bool = True,
    optimise_by: str = "duration",
    token: str | None = None,
    empty_weight: float = DEFAULT_EMPTY_WEIGHT,
    loaded_weight: float = DEFAULT_LOADED_WEIGHT,
    soft_workload_min: float = DEFAULT_SOFT_WORKLOAD_MIN,
    workload_penalty_per_min: float = DEFAULT_WORKLOAD_PENALTY_PER_MIN,
    soft_adjusted_duration_min: float = DEFAULT_SOFT_ADJUSTED_DURATION_MIN,
    duration_penalty_per_min: float = DEFAULT_DURATION_PENALTY_PER_MIN,
    max_job_overage_penalty: float = DEFAULT_MAX_JOB_OVERAGE_PENALTY,
    duration_buffer_multiplier: float = DEFAULT_DURATION_BUFFER_MULTIPLIER,
    empty_travel_duration_multiplier: float = DEFAULT_EMPTY_TRAVEL_DURATION_MULTIPLIER,
    empty_travel_wait_buffer_min: float = DEFAULT_EMPTY_TRAVEL_WAIT_BUFFER_MIN,
) -> dict[str, Any]:
    current_location = rider.start_location
    current_zone = rider.start_zone
    total_empty_distance = 0.0
    total_empty_duration = 0.0
    in_window_empty_distance = 0.0
    in_window_empty_duration = 0.0
    loaded_distance = 0.0
    loaded_duration = 0.0
    in_window_duration = 0.0
    first_positioning_duration: float | None = None
    first_positioning_distance = 0.0
    first_pickup_eta = _format_minutes_as_time(JOB_WINDOW_START_MIN)
    latest_departure_time = "-"
    rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    cluster_names = [
        _route_zone_for_job(sequence_job, "Pickup Address")
        or _route_zone_for_job(sequence_job, "Drop-off Address")
        or "Unknown"
        for sequence_job in sequence
    ]
    cluster_counts = {cluster_name: cluster_names.count(cluster_name) for cluster_name in set(cluster_names)}

    for sequence_index, job in enumerate(sequence, start=1):
        pickup_address = clean_text(job["Pickup Address"])
        dropoff_address = clean_text(job["Drop-off Address"])
        pickup_zone = _route_zone_for_job(job, "Pickup Address")
        dropoff_zone = _route_zone_for_job(job, "Drop-off Address")
        cluster_name = pickup_zone or dropoff_zone or "Unknown"

        empty_cost = get_empty_travel_cost(
            current_location,
            pickup_address,
            current_zone,
            pickup_zone,
            use_onemap=use_onemap,
            token=token,
            allow_walk=sequence_index > 1,
        )
        empty_cost = adjust_empty_travel_for_public_transport(
            empty_cost,
            duration_multiplier=empty_travel_duration_multiplier,
            wait_buffer_min=empty_travel_wait_buffer_min,
        )
        loaded_cost = get_travel_cost(
            pickup_address,
            dropoff_address,
            pickup_zone,
            dropoff_zone,
            use_onemap=use_onemap,
            token=token,
        )
        route_zone_priority, same_zone_pickup, route_stays_current_zone = calculate_route_zone_priority(
            current_zone,
            pickup_zone,
            dropoff_zone,
        )
        empty_duration_for_score = (empty_cost.duration_min or 0) * (
            FIRST_POSITIONING_WEIGHT if sequence_index == 1 else 1.0
        )
        score_data = calculate_assignment_score(
            empty_distance_km=empty_cost.distance_km,
            empty_duration_min=empty_duration_for_score,
            loaded_distance_km=loaded_cost.distance_km,
            loaded_duration_min=loaded_cost.duration_min,
            rider_current_zone=current_zone,
            pickup_zone=pickup_zone,
            dropoff_zone=dropoff_zone,
            rider_total_duration_min=in_window_duration,
            rider_assigned_jobs=sequence_index - 1,
            rider_max_jobs=rider.max_jobs,
            rider_load_level=rider.load_level,
            optimise_by=optimise_by,
            empty_weight=empty_weight,
            loaded_weight=loaded_weight,
            soft_workload_min=soft_workload_min,
            workload_penalty_per_min=workload_penalty_per_min,
            soft_adjusted_duration_min=soft_adjusted_duration_min,
            duration_buffer_multiplier=duration_buffer_multiplier,
            duration_penalty_per_min=duration_penalty_per_min,
            max_jobs_overage_penalty=max_job_overage_penalty,
            max_adjusted_duration_min=math.inf,
        )
        if score_data is None:
            return {
                "valid": False,
                "rows": rows,
                "raw_duration": in_window_duration,
                "adjusted_duration": in_window_duration * duration_buffer_multiplier,
                "reason": "route exceeded configured limits",
                "warnings": warnings,
            }

        if empty_cost.error:
            warnings.append(f"{current_location} -> {pickup_address}: {empty_cost.error}")
        if loaded_cost.error:
            warnings.append(f"{pickup_address} -> {dropoff_address}: {loaded_cost.error}")

        cost_source = _combined_source(empty_cost, loaded_cost)
        if sequence_index == 1:
            first_positioning_duration = empty_cost.duration_min
            first_positioning_distance = empty_cost.distance_km or 0
            latest_departure_time = _format_minutes_as_time(
                JOB_WINDOW_START_MIN - float(first_positioning_duration or 0)
            )
            row_in_window_duration = loaded_cost.duration_min or 0
        else:
            row_in_window_duration = (empty_cost.duration_min or 0) + (loaded_cost.duration_min or 0)
            in_window_empty_distance += empty_cost.distance_km or 0
            in_window_empty_duration += empty_cost.duration_min or 0
        in_window_duration += row_in_window_duration
        is_window_valid, feasibility_status = _job_window_status(in_window_duration, first_positioning_duration)
        final_completion_eta = _format_minutes_as_time(JOB_WINDOW_START_MIN + in_window_duration)
        rows.append(
            {
                "Rider": rider.name,
                "Sequence": sequence_index,
                "Uploaded Row": _job_uploaded_row(job),
                "Start From": current_location,
                "Empty Travel To Pickup": f"{current_location} -> {pickup_address}",
                "Empty PT Instructions": empty_cost.route_text,
                "Empty Route Path": json.dumps(empty_cost.route_path or []),
                "Car Plate": clean_text(job.get("Car Plate")),
                "Pickup Address": pickup_address,
                "Pickup Lot": clean_text(job.get("Pickup Lot")),
                "Drop-off Address": dropoff_address,
                "Loaded Travel / Car Movement": f"{pickup_address} -> {dropoff_address}",
                "Loaded Drive Instructions": loaded_cost.route_text,
                "Loaded Route Path": json.dumps(loaded_cost.route_path or []),
                "Empty Distance KM": empty_cost.distance_km,
                "Empty Duration Min": empty_cost.duration_min,
                "Loaded Distance KM": loaded_cost.distance_km,
                "Loaded Duration Min": loaded_cost.duration_min,
                "Total Distance KM": round((empty_cost.distance_km or 0) + (loaded_cost.distance_km or 0), 2),
                "Total Duration Min": round(row_in_window_duration, 1),
                "Assignment Score": score_data["assignment_score"],
                "Zone Adjustment": score_data["zone_adjustment"],
                "Same Zone Pickup": "Yes" if same_zone_pickup else "No",
                "Same Zone Route": "Yes" if route_stays_current_zone else "No",
                "Route Zone Priority": route_zone_priority,
                "Empty Weight": empty_weight,
                "Loaded Weight": loaded_weight,
                "Workload Penalty": score_data["workload_penalty"],
                "Duration Penalty": score_data["duration_penalty"],
                "Max Jobs Penalty": score_data["max_jobs_penalty"],
                "Projected Rider Duration Min": round(float(in_window_duration), 1),
                "Projected Adjusted Duration Min": round(float(in_window_duration * duration_buffer_multiplier), 1),
                "First Positioning PT Duration Min": round(float(first_positioning_duration or 0), 1),
                "First Pickup ETA": first_pickup_eta,
                "Latest Departure Time": latest_departure_time,
                "In-Window Route Duration Min": round(float(in_window_duration), 1),
                "Final Completion ETA": final_completion_eta,
                "Cluster Name / Zone": cluster_name,
                "Cluster Job Count": cluster_counts.get(cluster_name, 1),
                "Feasibility Status": feasibility_status,
                "Reason if Unassigned": "",
                "Cost Source": cost_source,
                "Route Validation Status": "",
            }
        )

        total_empty_distance += empty_cost.distance_km or 0
        total_empty_duration += empty_cost.duration_min or 0
        loaded_distance += loaded_cost.distance_km or 0
        loaded_duration += loaded_cost.duration_min or 0
        current_location = dropoff_address
        current_zone = dropoff_zone

        if not is_window_valid:
            return {
                "valid": False,
                "rows": rows,
                "empty_distance": in_window_empty_distance,
                "empty_duration": in_window_empty_duration,
                "total_empty_distance": total_empty_distance,
                "total_empty_duration": total_empty_duration,
                "positioning_distance": first_positioning_distance,
                "positioning_duration": first_positioning_duration or 0,
                "loaded_distance": loaded_distance,
                "loaded_duration": loaded_duration,
                "raw_duration": in_window_duration,
                "adjusted_duration": in_window_duration * duration_buffer_multiplier,
                "final_location": current_location,
                "final_zone": current_zone,
                "reason": feasibility_status,
                "warnings": warnings,
            }

    raw_duration = in_window_duration
    valid, reason = _job_window_status(raw_duration, first_positioning_duration if sequence else 0)
    return {
        "valid": valid,
        "rows": rows,
        "empty_distance": in_window_empty_distance,
        "empty_duration": in_window_empty_duration,
        "total_empty_distance": total_empty_distance,
        "total_empty_duration": total_empty_duration,
        "positioning_distance": first_positioning_distance or 0,
        "positioning_duration": first_positioning_duration or 0,
        "loaded_distance": loaded_distance,
        "loaded_duration": loaded_duration,
        "raw_duration": raw_duration,
        "adjusted_duration": raw_duration * duration_buffer_multiplier,
        "final_location": current_location,
        "final_zone": current_zone,
        "final_completion_eta": _format_minutes_as_time(JOB_WINDOW_START_MIN + raw_duration),
        "reason": reason,
        "warnings": warnings,
    }


def rebuild_outputs_from_sequences(
    rider_sequences: dict[str, list[str]],
    riders: list[RiderState],
    jobs_by_id: dict[str, dict[str, Any]],
    jobs_df: pd.DataFrame | None = None,
    **settings: Any,
) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    route_rows: list[dict[str, Any]] = []
    lookup_warnings: list[str] = []
    rider_by_name = {rider.name: RiderState(
        name=rider.name,
        start_location=rider.start_location,
        start_zone=rider.start_zone,
        max_jobs=rider.max_jobs,
        load_level=rider.load_level,
        current_location=rider.start_location,
        current_zone=rider.start_zone,
    ) for rider in riders}

    for rider_name, rider in rider_by_name.items():
        sequence_ids = rider_sequences.get(rider_name, [])
        sequence_jobs = [jobs_by_id[job_id] for job_id in sequence_ids if job_id in jobs_by_id]
        evaluation = evaluate_explicit_rider_sequence(rider, sequence_jobs, **settings)
        lookup_warnings.extend(evaluation.get("warnings", []))
        route_rows.extend(evaluation.get("rows", []))
        rider.assigned_count = len(sequence_jobs)
        rider.empty_distance_km = float(evaluation.get("empty_distance", 0) or 0)
        rider.empty_duration_min = float(evaluation.get("empty_duration", 0) or 0)
        rider.loaded_distance_km = float(evaluation.get("loaded_distance", 0) or 0)
        rider.loaded_duration_min = float(evaluation.get("loaded_duration", 0) or 0)
        rider.current_location = clean_text(evaluation.get("final_location", rider.start_location))
        rider.current_zone = evaluation.get("final_zone") or rider.start_zone

    route_df = format_route_output(pd.DataFrame(route_rows, columns=ROUTE_COLUMNS), list(rider_by_name.values()))
    if jobs_df is not None:
        validate_optimisation_integrity(route_df, jobs_df)

    duration_buffer_multiplier = float(settings.get("duration_buffer_multiplier", DEFAULT_DURATION_BUFFER_MULTIPLIER))
    summary_df = pd.DataFrame(
        [
            {
                "Rider": rider.name,
                "Total Jobs": rider.assigned_count,
                "Total Empty Distance KM": round(rider.empty_distance_km, 2),
                "Total Empty Duration Min": round(rider.empty_duration_min, 1),
                "Total Loaded Distance KM": round(rider.loaded_distance_km, 2),
                "Total Loaded Duration Min": round(rider.loaded_duration_min, 1),
                "Total Route Distance KM": round(rider.empty_distance_km + rider.loaded_distance_km, 2),
                "Total Route Duration Min": round(rider.empty_duration_min + rider.loaded_duration_min, 1),
                "Adjusted Route Duration Min": round(
                    (rider.empty_duration_min + rider.loaded_duration_min) * duration_buffer_multiplier,
                    1,
                ),
                "Within 3 Hours": _route_status_for_adjusted_duration(
                    (rider.empty_duration_min + rider.loaded_duration_min) * duration_buffer_multiplier,
                    DEFAULT_MAX_ADJUSTED_DURATION_MIN,
                ),
                "Final Location": rider.current_location,
            }
            for rider in rider_by_name.values()
        ],
        columns=SUMMARY_COLUMNS,
    )
    summary_df = format_summary_output(summary_df, route_df)
    return route_df, summary_df, _dedupe_lookup_warnings(lookup_warnings)


def _normalise_sequence_map(sequences: dict[str, list[str]]) -> dict[str, list[str]]:
    return {clean_text(rider): [clean_text(job_id) for job_id in jobs if clean_text(job_id)] for rider, jobs in sequences.items()}


def _sequence_plan_signature(sequences: dict[str, list[str]]) -> tuple[tuple[str, tuple[str, ...]], ...]:
    return tuple((rider, tuple(jobs)) for rider, jobs in sorted(_normalise_sequence_map(sequences).items()))


def _job_positions(sequences: dict[str, list[str]]) -> dict[str, tuple[str, int]]:
    positions: dict[str, tuple[str, int]] = {}
    for rider, jobs in sequences.items():
        for index, job_id in enumerate(jobs):
            positions[job_id] = (rider, index)
    return positions


def _changed_riders(original: dict[str, list[str]], proposed: dict[str, list[str]]) -> list[str]:
    rider_names = sorted(set(original) | set(proposed))
    return [rider for rider in rider_names if original.get(rider, []) != proposed.get(rider, [])]


def _moved_jobs(original: dict[str, list[str]], proposed: dict[str, list[str]], candidate_jobs: set[str]) -> list[dict[str, Any]]:
    original_positions = _job_positions(original)
    proposed_positions = _job_positions(proposed)
    moved = []
    for job_id in sorted(candidate_jobs):
        before = original_positions.get(job_id)
        after = proposed_positions.get(job_id)
        if before != after:
            moved.append(
                {
                    "job_id": job_id,
                    "from_rider": before[0] if before else "",
                    "from_sequence": before[1] + 1 if before else "",
                    "to_rider": after[0] if after else "",
                    "to_sequence": after[1] + 1 if after else "",
                    "changed_rider": bool(before and after and before[0] != after[0]),
                }
            )
    return moved


def _route_score(summary_df: pd.DataFrame) -> float:
    if summary_df is None or summary_df.empty or "Adjusted Route Duration Min" not in summary_df.columns:
        return 0.0
    return float(summary_df["Adjusted Route Duration Min"].fillna(0).astype(float).sum())


def _latest_completion(route_df: pd.DataFrame, riders: list[str] | None = None) -> str:
    if route_df is None or route_df.empty or "Final Completion ETA" not in route_df.columns:
        return "-"
    routes = route_df
    if riders is not None and "Rider" in routes.columns:
        routes = routes[routes["Rider"].astype(str).isin(riders)]
    values = [clean_text(value) for value in routes["Final Completion ETA"].tolist() if clean_text(value) and clean_text(value) != "-"]
    return max(values) if values else "-"


def _selective_plan_result(
    original_sequences: dict[str, list[str]],
    proposed_sequences: dict[str, list[str]],
    route_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    baseline_score: float,
    movable_job_ids: set[str],
    candidate_count: int,
    *,
    changed_rider_penalty: float,
    moved_job_penalty: float,
    sequence_change_penalty: float,
) -> dict[str, Any]:
    changed = _changed_riders(original_sequences, proposed_sequences)
    moved = _moved_jobs(original_sequences, proposed_sequences, movable_job_ids)
    changed_rider_moves = sum(1 for item in moved if item["changed_rider"])
    sequence_changes = len(moved) - changed_rider_moves
    score_after = _route_score(summary_df)
    duration_delta = score_after - baseline_score
    disruption_penalty = (
        len(changed) * changed_rider_penalty
        + len(moved) * moved_job_penalty
        + sequence_changes * sequence_change_penalty
        + max(0.0, duration_delta)
    )
    plan_score = score_after + disruption_penalty
    return {
        "success": True,
        "original_sequences": {rider: list(jobs) for rider, jobs in original_sequences.items()},
        "proposed_sequences": {rider: list(jobs) for rider, jobs in proposed_sequences.items()},
        "route_df": route_df,
        "summary_df": summary_df,
        "changed_riders": changed,
        "moved_jobs": moved,
        "score_before": round(float(baseline_score), 3),
        "score_after": round(float(score_after), 3),
        "duration_delta": round(float(duration_delta), 3),
        "plan_score": round(float(plan_score), 3),
        "disruption_penalty": round(float(disruption_penalty), 3),
        "candidate_count": candidate_count,
        "latest_completion_after": _latest_completion(route_df, changed or None),
        "reason": "Beneficial reshuffle found." if duration_delta < -0.1 else "Best feasible alternative found; it does not improve adjusted duration.",
    }


def _validate_selective_inputs(
    current_rider_sequences: dict[str, list[str]],
    jobs_by_id: dict[str, dict[str, Any]],
    locked_riders: set[str],
    locked_job_ids: set[str],
    reshuffle_job_ids: set[str],
    eligible_receiver_riders: set[str],
) -> tuple[bool, str]:
    all_sequence_jobs = [job_id for jobs in current_rider_sequences.values() for job_id in jobs]
    duplicate_jobs = sorted({job_id for job_id in all_sequence_jobs if all_sequence_jobs.count(job_id) > 1})
    if duplicate_jobs:
        return False, "Current route state has duplicate jobs: " + ", ".join(duplicate_jobs[:5])
    if not reshuffle_job_ids:
        return False, "Select at least one job for the reshuffle pool."
    missing = sorted(job_id for job_id in reshuffle_job_ids if job_id not in jobs_by_id or job_id not in all_sequence_jobs)
    if missing:
        return False, "Selected job no longer exists in the current route state."
    locked_selected = sorted(reshuffle_job_ids & locked_job_ids)
    if locked_selected:
        return False, "A selected reshuffle job is locked. Unlock it before searching."
    job_positions = _job_positions(current_rider_sequences)
    locked_rider_selected = sorted(
        job_id for job_id in reshuffle_job_ids if job_positions.get(job_id, ("", 0))[0] in locked_riders
    )
    if locked_rider_selected:
        return False, "A selected reshuffle job belongs to a locked rider. Unlock that rider before searching."
    if not eligible_receiver_riders and len(reshuffle_job_ids) > 1:
        return False, "Select at least one eligible receiver, or use a single selected job that can stay where it is."
    return True, ""


def find_best_selective_reshuffle(
    current_rider_sequences: dict[str, list[str]],
    jobs_by_id: dict[str, dict[str, Any]],
    riders: list[RiderState],
    jobs_df: pd.DataFrame | None = None,
    locked_riders: set[str] | None = None,
    locked_job_ids: set[str] | None = None,
    reshuffle_job_ids: set[str] | None = None,
    eligible_receiver_riders: set[str] | None = None,
    top_n: int = 5,
    beam_width: int = DEFAULT_SELECTIVE_BEAM_WIDTH,
    max_candidates: int = MAX_SELECTIVE_RESHUFFLE_CANDIDATES,
    changed_rider_penalty: float = DEFAULT_SELECTIVE_CHANGED_RIDER_PENALTY,
    moved_job_penalty: float = DEFAULT_SELECTIVE_MOVED_JOB_PENALTY,
    sequence_change_penalty: float = DEFAULT_SELECTIVE_SEQUENCE_CHANGE_PENALTY,
    **settings: Any,
) -> dict[str, Any]:
    locked_riders = {clean_text(rider) for rider in (locked_riders or set()) if clean_text(rider)}
    locked_job_ids = {clean_text(job_id) for job_id in (locked_job_ids or set()) if clean_text(job_id)}
    reshuffle_job_ids = {clean_text(job_id) for job_id in (reshuffle_job_ids or set()) if clean_text(job_id)}
    eligible_receiver_riders = {
        clean_text(rider)
        for rider in (eligible_receiver_riders or set())
        if clean_text(rider) and clean_text(rider) not in locked_riders
    }
    original_sequences = _normalise_sequence_map(current_rider_sequences)
    ok, reason = _validate_selective_inputs(
        original_sequences,
        jobs_by_id,
        locked_riders,
        locked_job_ids,
        reshuffle_job_ids,
        eligible_receiver_riders,
    )
    if not ok:
        return {"success": False, "reason": reason, "alternatives": [], "candidate_count": 0}

    rider_names = [rider.name for rider in riders]
    for rider_name in rider_names:
        original_sequences.setdefault(rider_name, [])
    original_positions = _job_positions(original_sequences)
    original_receivers = {original_positions[job_id][0] for job_id in reshuffle_job_ids}
    receiver_names = sorted((eligible_receiver_riders | original_receivers) - locked_riders)
    if not receiver_names:
        return {"success": False, "reason": "No eligible receiver rider is available.", "alternatives": [], "candidate_count": 0}

    base_sequences = {
        rider: [job_id for job_id in jobs if job_id not in reshuffle_job_ids]
        for rider, jobs in original_sequences.items()
    }

    try:
        baseline_route_df, baseline_summary_df, baseline_warnings = rebuild_outputs_from_sequences(
            original_sequences,
            riders,
            jobs_by_id,
            jobs_df=jobs_df,
            **settings,
        )
    except Exception as exc:
        return {"success": False, "reason": f"Current route state is invalid: {exc}", "alternatives": [], "candidate_count": 0}

    baseline_score = _route_score(baseline_summary_df)
    baseline_latest = _latest_completion(baseline_route_df, sorted(original_receivers))
    sequence_eval_cache: dict[tuple[Any, ...], dict[str, Any]] = {}

    def route_is_valid(sequences: dict[str, list[str]], affected_riders: set[str]) -> bool:
        rider_by_name = {rider.name: rider for rider in riders}
        for rider_name in affected_riders:
            rider = rider_by_name.get(rider_name)
            if rider is None:
                return False
            sequence_jobs = [jobs_by_id[job_id] for job_id in sequences.get(rider_name, []) if job_id in jobs_by_id]
            cache_key = _rider_sequence_cache_key(rider, sequence_jobs, settings)
            if cache_key not in sequence_eval_cache:
                sequence_eval_cache[cache_key] = evaluate_explicit_rider_sequence(rider, sequence_jobs, **settings)
            if not sequence_eval_cache[cache_key].get("valid"):
                return False
        return True

    candidate_count = 0
    search_limited = False
    unique_signatures: set[tuple[tuple[str, tuple[str, ...]], ...]] = set()
    alternatives: list[dict[str, Any]] = []
    use_exhaustive = len(reshuffle_job_ids) <= 4

    def consider_plan(sequences: dict[str, list[str]]) -> None:
        nonlocal candidate_count, search_limited
        if candidate_count >= max_candidates:
            search_limited = True
            return
        signature = _sequence_plan_signature(sequences)
        if signature in unique_signatures:
            return
        unique_signatures.add(signature)
        candidate_count += 1

        changed = set(_changed_riders(original_sequences, sequences))
        affected = changed | original_receivers
        if not route_is_valid(sequences, affected):
            return
        try:
            route_df, summary_df, _ = rebuild_outputs_from_sequences(
                sequences,
                riders,
                jobs_by_id,
                jobs_df=jobs_df,
                **settings,
            )
        except Exception:
            return
        alternatives.append(
            _selective_plan_result(
                original_sequences,
                sequences,
                route_df,
                summary_df,
                baseline_score,
                reshuffle_job_ids,
                candidate_count,
                changed_rider_penalty=changed_rider_penalty,
                moved_job_penalty=moved_job_penalty,
                sequence_change_penalty=sequence_change_penalty,
            )
        )
        alternatives.sort(key=lambda item: (float(item["plan_score"]), float(item["score_after"])))
        del alternatives[top_n * 4 :]

    def insert_job(sequences: dict[str, list[str]], rider_name: str, job_id: str, position: int) -> dict[str, list[str]]:
        next_sequences = {rider: list(jobs) for rider, jobs in sequences.items()}
        next_sequences.setdefault(rider_name, [])
        next_sequences[rider_name] = next_sequences[rider_name][:position] + [job_id] + next_sequences[rider_name][position:]
        return next_sequences

    movable_jobs = sorted(reshuffle_job_ids, key=lambda job_id: original_positions[job_id])
    if use_exhaustive:
        for job_order in itertools.permutations(movable_jobs):
            partials = [base_sequences]
            for job_id in job_order:
                next_partials = []
                for partial in partials:
                    for rider_name in receiver_names:
                        if rider_name in locked_riders:
                            continue
                        sequence_len = len(partial.get(rider_name, []))
                        for position in range(sequence_len + 1):
                            next_partials.append(insert_job(partial, rider_name, job_id, position))
                            if len(next_partials) + candidate_count >= max_candidates:
                                search_limited = True
                                break
                        if search_limited:
                            break
                    if search_limited:
                        break
                partials = next_partials
                if search_limited:
                    break
            for plan in partials:
                consider_plan(plan)
                if search_limited:
                    break
            if search_limited:
                break
    else:
        partials = [(base_sequences, 0.0)]
        for job_id in movable_jobs:
            next_partials: list[tuple[dict[str, list[str]], float]] = []
            for partial, _ in partials:
                for rider_name in receiver_names:
                    sequence_len = len(partial.get(rider_name, []))
                    for position in range(sequence_len + 1):
                        candidate = insert_job(partial, rider_name, job_id, position)
                        moved_so_far = len(_moved_jobs(original_sequences, candidate, set(movable_jobs)))
                        changed_so_far = len(_changed_riders(original_sequences, candidate))
                        heuristic = moved_so_far * moved_job_penalty + changed_so_far * changed_rider_penalty
                        next_partials.append((candidate, heuristic))
                        if len(next_partials) >= max_candidates:
                            search_limited = True
                            break
                    if search_limited:
                        break
                if search_limited:
                    break
            next_partials.sort(key=lambda item: item[1])
            partials = next_partials[:beam_width]
            if search_limited:
                break
        for plan, _ in partials:
            consider_plan(plan)
            if search_limited:
                break

    alternatives = sorted(alternatives, key=lambda item: (float(item["plan_score"]), float(item["score_after"])))[:top_n]
    if not alternatives:
        return {
            "success": False,
            "reason": "No feasible selective reshuffle proposal was found.",
            "alternatives": [],
            "candidate_count": candidate_count,
            "search_limited": search_limited,
            "score_before": round(float(baseline_score), 3),
            "latest_completion_before": baseline_latest,
            "baseline_route_df": baseline_route_df,
            "baseline_summary_df": baseline_summary_df,
            "baseline_lookup_warnings": baseline_warnings,
        }

    for alternative in alternatives:
        alternative["candidate_count"] = candidate_count
        alternative["search_limited"] = search_limited
        alternative["latest_completion_before"] = baseline_latest
    best = alternatives[0].copy()
    best["alternatives"] = alternatives
    return best


def optimisation_integrity_report(route_df: pd.DataFrame, jobs_df: pd.DataFrame | None) -> dict[str, Any]:
    assigned_rows = _route_uploaded_rows(route_df)
    duplicate_rows = sorted(assigned_rows[assigned_rows.duplicated(keep=False)].unique().tolist())
    unassigned_df = build_unassigned_jobs_df(jobs_df, route_df)
    unassigned_rows = _route_uploaded_rows(unassigned_df)

    if jobs_df is None or jobs_df.empty:
        total_rows = pd.Series(dtype="Int64")
    elif "Uploaded Row" in jobs_df.columns:
        total_rows = pd.to_numeric(jobs_df["Uploaded Row"], errors="coerce").dropna().astype(int)
    elif "_original_order" in jobs_df.columns:
        total_rows = jobs_df["_original_order"].astype(int) + 2
    else:
        total_rows = pd.Series(range(2, len(jobs_df) + 2), dtype="int")

    overlap_rows = sorted(set(assigned_rows.tolist()) & set(unassigned_rows.tolist()))
    assigned_unique_count = int(assigned_rows.nunique())
    assigned_row_count = int(len(assigned_rows))
    unassigned_count = int(unassigned_rows.nunique())
    total_jobs = int(total_rows.nunique())
    is_valid = (
        assigned_row_count == assigned_unique_count
        and assigned_unique_count + unassigned_count == total_jobs
        and not overlap_rows
        and not duplicate_rows
    )

    duplicate_details = []
    if duplicate_rows and route_df is not None and not route_df.empty:
        for uploaded_row in duplicate_rows:
            rows = route_df[pd.to_numeric(route_df["Uploaded Row"], errors="coerce") == uploaded_row]
            first = rows.iloc[0]
            duplicate_details.append(
                {
                    "Uploaded Row": int(uploaded_row),
                    "Car Plate": clean_text(first.get("Car Plate")),
                    "Pickup Address": clean_text(first.get("Pickup Address")),
                    "Drop-off Address": clean_text(first.get("Drop-off Address")),
                    "Riders": ", ".join(sorted(rows["Rider"].astype(str).unique())),
                }
            )

    return {
        "is_valid": is_valid,
        "message": "" if is_valid else "Invalid optimisation result: duplicate job assignment detected.",
        "total_valid_jobs": total_jobs,
        "assigned_unique_jobs": assigned_unique_count,
        "assigned_route_rows": assigned_row_count,
        "unassigned_jobs": unassigned_count,
        "duplicate_uploaded_rows": duplicate_rows,
        "overlap_uploaded_rows": overlap_rows,
        "duplicate_details": duplicate_details,
        "unassigned_df": unassigned_df,
    }


def validate_optimisation_integrity(route_df: pd.DataFrame, jobs_df: pd.DataFrame | None) -> None:
    report = optimisation_integrity_report(route_df, jobs_df)
    if report["is_valid"]:
        return
    details = report.get("duplicate_details") or []
    detail_text = "; ".join(
        f"Uploaded Row {detail['Uploaded Row']} ({detail['Car Plate']}) assigned to {detail['Riders']}"
        for detail in details
    )
    raise ValueError(f"{report['message']} {detail_text}".strip())


def _route_status_for_adjusted_duration(adjusted_duration: float, base_cap: float) -> str:
    if adjusted_duration <= base_cap:
        return "OK"
    return "Fail"


def _format_minutes_as_time(minutes_after_midnight: float | None) -> str:
    if minutes_after_midnight is None or math.isinf(float(minutes_after_midnight)):
        return "-"
    minutes = int(round(float(minutes_after_midnight))) % (24 * 60)
    return f"{minutes // 60:02d}:{minutes % 60:02d}"


def _job_window_status(in_window_duration: float, first_positioning_duration: float | None = None) -> tuple[bool, str]:
    if first_positioning_duration is None or math.isinf(float(first_positioning_duration)):
        return False, "no valid OneMap route"
    if JOB_WINDOW_START_MIN - float(first_positioning_duration) < 0:
        return False, "missed first pickup target"
    if JOB_WINDOW_START_MIN + float(in_window_duration) > JOB_WINDOW_END_MIN:
        return False, "final completion after 17:00"
    return True, "OK"


def _route_zone_for_job(job: dict[str, Any], key: str) -> str | None:
    address = clean_text(job.get(key))
    zone_key = "Pickup Zone" if key == "Pickup Address" else "Drop-off Zone"
    return job.get(zone_key) or infer_zone(address)


def _route_variant_score_adjustment(
    route_variant_index: int,
    rider: RiderState,
    job: dict[str, Any],
    assignment_round: int,
) -> float:
    if route_variant_index <= 0:
        return 0.0

    key = "|".join(
        [
            str(route_variant_index),
            str(assignment_round),
            rider.name,
            rider.start_location,
            str(_job_id(job)),
            clean_text(job.get("Pickup Address")),
            clean_text(job.get("Drop-off Address")),
        ]
    )
    digest = hashlib.sha256(key.encode("utf-8")).digest()
    bucket = int.from_bytes(digest[:2], "big") / 65535
    centred = (bucket * 2) - 1
    return centred * 30.0


def optimise_vehicle_routes(
    jobs: pd.DataFrame,
    riders: list[RiderState],
    use_onemap: bool = True,
    optimise_by: str = "duration",
    token: str | None = None,
    progress_callback: ProgressCallback | None = None,
    empty_weight: float = DEFAULT_EMPTY_WEIGHT,
    loaded_weight: float = DEFAULT_LOADED_WEIGHT,
    soft_workload_min: float = DEFAULT_SOFT_WORKLOAD_MIN,
    workload_penalty_per_min: float = DEFAULT_WORKLOAD_PENALTY_PER_MIN,
    soft_adjusted_duration_min: float = DEFAULT_SOFT_ADJUSTED_DURATION_MIN,
    duration_penalty_per_min: float = DEFAULT_DURATION_PENALTY_PER_MIN,
    max_job_overage_penalty: float = DEFAULT_MAX_JOB_OVERAGE_PENALTY,
    duration_buffer_multiplier: float = DEFAULT_DURATION_BUFFER_MULTIPLIER,
    max_adjusted_duration_min: float = DEFAULT_MAX_ADJUSTED_DURATION_MIN,
    empty_travel_duration_multiplier: float = DEFAULT_EMPTY_TRAVEL_DURATION_MULTIPLIER,
    empty_travel_wait_buffer_min: float = DEFAULT_EMPTY_TRAVEL_WAIT_BUFFER_MIN,
    force_complete_assignment: bool = True,
    cluster_pressure_bonus_per_job: float = DEFAULT_CLUSTER_PRESSURE_BONUS_PER_JOB,
    route_variant_index: int = 0,
) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    token = token or get_onemap_token()
    remaining = jobs.sort_values("_original_order").to_dict("records")
    route_rows: list[dict[str, Any]] = []
    lookup_warnings: list[str] = []
    total_jobs = len(remaining)
    active_rider_count = max(1, len(riders))
    minimum_job_target = _minimum_job_target(total_jobs, active_rider_count)
    estimated_comparisons = max(1, active_rider_count * total_jobs * (total_jobs + 1) // 2)
    comparison_count = 0
    progress_base = 0.2 if use_onemap else 0
    progress_span = 0.78 if use_onemap else 0.98
    rider_keys = {id(rider): f"{index}:{rider.name}:{rider.start_location}" for index, rider in enumerate(riders)}

    def rider_key(rider: RiderState) -> str:
        return rider_keys[id(rider)]

    rider_sequences: dict[str, list[dict[str, Any]]] = {rider_key(rider): [] for rider in riders}
    base_rider_state = {
        rider_key(rider): {
            "start_location": rider.start_location,
            "start_zone": rider.start_zone,
            "max_jobs": rider.max_jobs,
            "load_level": rider.load_level,
        }
        for rider in riders
    }
    assigned_job_ids: set[int] = set()
    rejected_candidate_audit: list[dict[str, Any]] = []
    sequence_evaluation_cache: dict[tuple[str, float, tuple[int, ...]], dict[str, Any]] = {}

    def report(status: str, **extra: Any) -> None:
        if progress_callback is None:
            return
        progress = progress_base + (progress_span * comparison_count / estimated_comparisons)
        payload = {
            "status": status,
            "assigned_jobs": len(route_rows),
            "total_jobs": total_jobs,
            "remaining_jobs": len(remaining),
            "comparison_count": comparison_count,
            "estimated_comparisons": estimated_comparisons,
            "progress": min(0.98, progress),
        }
        payload.update(extra)
        progress_callback(payload)

    def evaluate_rider_sequence(
        rider: RiderState,
        sequence: list[dict[str, Any]],
        cap: float,
    ) -> dict[str, Any]:
        key = rider_key(rider)
        cache_key = (key, round(float(cap), 3) if not math.isinf(float(cap)) else math.inf, tuple(_job_id(job) for job in sequence))
        if cache_key in sequence_evaluation_cache:
            return sequence_evaluation_cache[cache_key]

        current_location = base_rider_state[key]["start_location"]
        current_zone = base_rider_state[key]["start_zone"]
        total_empty_distance = 0.0
        total_empty_duration = 0.0
        in_window_empty_distance = 0.0
        in_window_empty_duration = 0.0
        loaded_distance = 0.0
        loaded_duration = 0.0
        in_window_duration = 0.0
        first_positioning_duration: float | None = None
        first_positioning_distance = 0.0
        first_pickup_eta = _format_minutes_as_time(JOB_WINDOW_START_MIN)
        latest_departure_time = "-"
        rows: list[dict[str, Any]] = []
        warnings: list[str] = []
        cluster_names = [
            _route_zone_for_job(sequence_job, "Pickup Address")
            or _route_zone_for_job(sequence_job, "Drop-off Address")
            or "Unknown"
            for sequence_job in sequence
        ]
        cluster_counts = {cluster_name: cluster_names.count(cluster_name) for cluster_name in set(cluster_names)}

        for sequence_index, job in enumerate(sequence, start=1):
            pickup_address = clean_text(job["Pickup Address"])
            dropoff_address = clean_text(job["Drop-off Address"])
            pickup_zone = _route_zone_for_job(job, "Pickup Address")
            dropoff_zone = _route_zone_for_job(job, "Drop-off Address")
            cluster_name = pickup_zone or dropoff_zone or "Unknown"

            empty_cost = get_empty_travel_cost(
                current_location,
                pickup_address,
                current_zone,
                pickup_zone,
                use_onemap=use_onemap,
                token=token,
                allow_walk=sequence_index > 1,
            )
            empty_cost = adjust_empty_travel_for_public_transport(
                empty_cost,
                duration_multiplier=empty_travel_duration_multiplier,
                wait_buffer_min=empty_travel_wait_buffer_min,
            )
            loaded_cost = get_travel_cost(
                pickup_address,
                dropoff_address,
                pickup_zone,
                dropoff_zone,
                use_onemap=use_onemap,
                token=token,
            )
            route_zone_priority, same_zone_pickup, route_stays_current_zone = calculate_route_zone_priority(
                current_zone,
                pickup_zone,
                dropoff_zone,
            )
            empty_duration_for_score = (empty_cost.duration_min or 0) * (
                FIRST_POSITIONING_WEIGHT if sequence_index == 1 else 1.0
            )
            score_data = calculate_assignment_score(
                empty_distance_km=empty_cost.distance_km,
                empty_duration_min=empty_duration_for_score,
                loaded_distance_km=loaded_cost.distance_km,
                loaded_duration_min=loaded_cost.duration_min,
                rider_current_zone=current_zone,
                pickup_zone=pickup_zone,
                dropoff_zone=dropoff_zone,
                rider_total_duration_min=in_window_duration,
                rider_assigned_jobs=sequence_index - 1,
                rider_max_jobs=base_rider_state[key]["max_jobs"],
                rider_load_level=base_rider_state[key]["load_level"],
                optimise_by=optimise_by,
                empty_weight=empty_weight,
                loaded_weight=loaded_weight,
                soft_workload_min=soft_workload_min,
                workload_penalty_per_min=workload_penalty_per_min,
                soft_adjusted_duration_min=soft_adjusted_duration_min,
                duration_buffer_multiplier=duration_buffer_multiplier,
                duration_penalty_per_min=duration_penalty_per_min,
                max_jobs_overage_penalty=max_job_overage_penalty,
                max_adjusted_duration_min=math.inf,
            )
            if score_data is None:
                result = {
                    "valid": False,
                    "rows": rows,
                    "raw_duration": in_window_duration,
                    "adjusted_duration": in_window_duration * duration_buffer_multiplier,
                    "reason": "exceeded max jobs only if max jobs is a hard setting",
                    "warnings": warnings,
                }
                sequence_evaluation_cache[cache_key] = result
                return result

            if empty_cost.error:
                warnings.append(f"{current_location} -> {pickup_address}: {empty_cost.error}")
            if loaded_cost.error:
                warnings.append(f"{pickup_address} -> {dropoff_address}: {loaded_cost.error}")

            cost_source = _combined_source(empty_cost, loaded_cost)
            if sequence_index == 1:
                first_positioning_duration = empty_cost.duration_min
                first_positioning_distance = empty_cost.distance_km or 0
                latest_departure_time = _format_minutes_as_time(
                    JOB_WINDOW_START_MIN - float(first_positioning_duration or 0)
                )
                row_in_window_duration = loaded_cost.duration_min or 0
            else:
                row_in_window_duration = (empty_cost.duration_min or 0) + (loaded_cost.duration_min or 0)
                in_window_empty_distance += empty_cost.distance_km or 0
                in_window_empty_duration += empty_cost.duration_min or 0
            in_window_duration += row_in_window_duration
            is_window_valid, feasibility_status = _job_window_status(in_window_duration, first_positioning_duration)
            final_completion_eta = _format_minutes_as_time(JOB_WINDOW_START_MIN + in_window_duration)
            rows.append(
                {
                    "Rider": rider.name,
                    "Sequence": sequence_index,
                    "Uploaded Row": _job_uploaded_row(job),
                    "Start From": current_location,
                    "Empty Travel To Pickup": f"{current_location} -> {pickup_address}",
                    "Empty PT Instructions": empty_cost.route_text,
                    "Empty Route Path": json.dumps(empty_cost.route_path or []),
                    "Car Plate": clean_text(job.get("Car Plate")),
                    "Pickup Address": pickup_address,
                    "Pickup Lot": clean_text(job.get("Pickup Lot")),
                    "Drop-off Address": dropoff_address,
                    "Loaded Travel / Car Movement": f"{pickup_address} -> {dropoff_address}",
                    "Loaded Drive Instructions": loaded_cost.route_text,
                    "Loaded Route Path": json.dumps(loaded_cost.route_path or []),
                    "Empty Distance KM": empty_cost.distance_km,
                    "Empty Duration Min": empty_cost.duration_min,
                    "Loaded Distance KM": loaded_cost.distance_km,
                    "Loaded Duration Min": loaded_cost.duration_min,
                    "Total Distance KM": round((empty_cost.distance_km or 0) + (loaded_cost.distance_km or 0), 2),
                    "Total Duration Min": round(row_in_window_duration, 1),
                    "Assignment Score": score_data["assignment_score"],
                    "Zone Adjustment": score_data["zone_adjustment"],
                    "Same Zone Pickup": "Yes" if same_zone_pickup else "No",
                    "Same Zone Route": "Yes" if route_stays_current_zone else "No",
                    "Route Zone Priority": route_zone_priority,
                    "Empty Weight": empty_weight,
                    "Loaded Weight": loaded_weight,
                    "Workload Penalty": score_data["workload_penalty"],
                    "Duration Penalty": score_data["duration_penalty"],
                    "Max Jobs Penalty": score_data["max_jobs_penalty"],
                    "Projected Rider Duration Min": round(float(in_window_duration), 1),
                    "Projected Adjusted Duration Min": round(float(in_window_duration * duration_buffer_multiplier), 1),
                    "First Positioning PT Duration Min": round(float(first_positioning_duration or 0), 1),
                    "First Pickup ETA": first_pickup_eta,
                    "Latest Departure Time": latest_departure_time,
                    "In-Window Route Duration Min": round(float(in_window_duration), 1),
                    "Final Completion ETA": final_completion_eta,
                    "Cluster Name / Zone": cluster_name,
                    "Cluster Job Count": cluster_counts.get(cluster_name, 1),
                    "Feasibility Status": feasibility_status,
                    "Reason if Unassigned": "",
                    "Cost Source": cost_source,
                    "Route Validation Status": "",
                }
            )

            total_empty_distance += empty_cost.distance_km or 0
            total_empty_duration += empty_cost.duration_min or 0
            loaded_distance += loaded_cost.distance_km or 0
            loaded_duration += loaded_cost.duration_min or 0
            current_location = dropoff_address
            current_zone = dropoff_zone

        raw_duration = in_window_duration
        valid, reason = _job_window_status(raw_duration, first_positioning_duration if sequence else 0)
        result = {
            "valid": valid,
            "rows": rows,
            "empty_distance": in_window_empty_distance,
            "empty_duration": in_window_empty_duration,
            "total_empty_distance": total_empty_distance,
            "total_empty_duration": total_empty_duration,
            "positioning_distance": first_positioning_distance,
            "positioning_duration": first_positioning_duration or 0,
            "loaded_distance": loaded_distance,
            "loaded_duration": loaded_duration,
            "raw_duration": raw_duration,
            "adjusted_duration": raw_duration * duration_buffer_multiplier,
            "final_location": current_location,
            "final_zone": current_zone,
            "final_completion_eta": _format_minutes_as_time(JOB_WINDOW_START_MIN + raw_duration),
            "reason": reason,
            "warnings": warnings,
        }
        sequence_evaluation_cache[cache_key] = result
        return result

    if use_onemap:
        unique_addresses = sorted(
            {
                clean_text(address)
                for address in (
                    list(jobs["Pickup Address"])
                    + list(jobs["Drop-off Address"])
                    + [rider.start_location for rider in riders]
                )
                if clean_text(address)
            }
        )
        for index, address in enumerate(unique_addresses, start=1):
            report(
                f"Geocoding address {index} of {len(unique_addresses)}",
                phase="Geocoding",
                current_address=address,
                progress=min(0.2, 0.2 * index / max(1, len(unique_addresses))),
            )
            result = get_cached_geocode(address, token=token, use_onemap=True)
            if result.error:
                lookup_warnings.append(f"{address}: {result.error}")
    else:
        report("Using fallback zone estimates only", phase="Fallback")

    # Semi-optimised greedy assignment: every round compares all remaining fixed
    # pickup-to-drop-off jobs against every rider's real current location.
    assignment_round = 0
    while remaining and riders:
        assignment_round += 1
        report("Comparing rider-job combinations", phase="Comparing")
        remaining_pickup_zone_counts: dict[str, int] = {}
        for remaining_job in remaining:
            remaining_zone = _route_zone_for_job(remaining_job, "Pickup Address") or "Unknown"
            remaining_pickup_zone_counts[remaining_zone] = remaining_pickup_zone_counts.get(remaining_zone, 0) + 1

        best_choice: tuple[
            tuple[int, float, float, float, int, int],
            int,
            RiderState,
            dict[str, Any],
            dict[str, Any],
            dict[str, Any],
        ] | None = None

        for rider in riders:
            for job_index, job in enumerate(remaining):
                if _job_id(job) in assigned_job_ids:
                    continue
                comparison_count += 1
                pickup_address = clean_text(job["Pickup Address"])
                pickup_zone = job.get("Pickup Zone") or infer_zone(pickup_address)
                key = rider_key(rider)
                candidate_sequence = rider_sequences[key] + [job]
                evaluation = evaluate_rider_sequence(rider, candidate_sequence, max_adjusted_duration_min)
                if not evaluation.get("valid"):
                    continue
                inserted_row = evaluation["rows"][-1]

                rider_minimum_target = _rider_minimum_job_target(rider, minimum_job_target)
                minimum_priority = _minimum_job_priority(rider.assigned_count, rider_minimum_target)
                load_policy = RIDER_LOAD_POLICIES.get(
                    normalise_rider_load_level(rider.load_level),
                    RIDER_LOAD_POLICIES["Medium"],
                )
                cluster_pressure_bonus = (
                    -cluster_pressure_bonus_per_job
                    * float(load_policy["cluster_pressure_multiplier"])
                    * remaining_pickup_zone_counts.get(pickup_zone or "Unknown", 0)
                )
                route_variant_adjustment = _route_variant_score_adjustment(
                    route_variant_index,
                    rider,
                    job,
                    assignment_round,
                )
                greedy_assignment_score = (
                    float(inserted_row.get("Assignment Score", 0) or 0)
                    + cluster_pressure_bonus
                    + route_variant_adjustment
                )
                cluster_jump_penalty = (
                    0
                    if rider.current_zone == pickup_zone or rider.assigned_count == 0
                    else int(load_policy["cluster_jump_penalty"])
                )
                candidate_rank = (
                    minimum_priority,
                    greedy_assignment_score,
                    float(evaluation.get("raw_duration", math.inf)),
                    cluster_jump_penalty,
                    -int(inserted_row.get("Cluster Job Count", 1) or 1),
                    int(job["_original_order"]),
                )
                if best_choice is None or candidate_rank < best_choice[0]:
                    best_choice = (
                        candidate_rank,
                        job_index,
                        rider,
                        job,
                        evaluation,
                        inserted_row,
                    )

                if comparison_count == 1 or comparison_count % 10 == 0:
                    report(
                        "Comparing rider-job combinations",
                        phase="Routing",
                        current_rider=rider.name,
                        current_pickup=pickup_address,
                    )

        if best_choice is None:
            break

        (
            _,
            job_index,
            rider,
            job,
            evaluation,
            inserted_row,
        ) = best_choice
        pickup_address = clean_text(job["Pickup Address"])
        dropoff_address = clean_text(job["Drop-off Address"])
        lookup_warnings.extend(evaluation.get("warnings", []))

        report(
            f"Assigned job {len(route_rows) + 1} of {total_jobs}",
            phase="Assigning",
            current_rider=rider.name,
            current_pickup=pickup_address,
            current_dropoff=dropoff_address,
        )

        route_rows.append(inserted_row)

        assigned_job_ids.add(_job_id(job))
        key = rider_key(rider)
        rider_sequences[key].append(job)
        rider.assigned_count = len(rider_sequences[key])
        rider.empty_distance_km = float(evaluation.get("empty_distance", 0) or 0)
        rider.empty_duration_min = float(evaluation.get("empty_duration", 0) or 0)
        rider.loaded_distance_km = float(evaluation.get("loaded_distance", 0) or 0)
        rider.loaded_duration_min = float(evaluation.get("loaded_duration", 0) or 0)
        rider.current_location = clean_text(evaluation.get("final_location", dropoff_address))
        rider.current_zone = evaluation.get("final_zone") or infer_zone(dropoff_address)
        remaining.pop(job_index)

    cap_stages = [max_adjusted_duration_min] if force_complete_assignment else []
    cap_used = max_adjusted_duration_min
    if remaining:
        for cap in cap_stages:
            cap_used = cap
            made_assignment = True
            while remaining and made_assignment:
                made_assignment = False
                best_insertion: tuple[int, float, float, int, RiderState, int, int, dict[str, Any], dict[str, Any]] | None = None

                current_evaluations = {
                    rider_key(rider): evaluate_rider_sequence(rider, rider_sequences[rider_key(rider)], cap)
                    for rider in riders
                }
                for job_index, job in enumerate(remaining):
                    if _job_id(job) in assigned_job_ids:
                        continue
                    for rider in riders:
                        key = rider_key(rider)
                        current_sequence = rider_sequences[key]
                        current_duration = float(current_evaluations[key].get("raw_duration", 0) or 0)
                        for insert_at in range(len(current_sequence) + 1):
                            candidate_sequence = current_sequence[:insert_at] + [job] + current_sequence[insert_at:]
                            evaluation = evaluate_rider_sequence(rider, candidate_sequence, cap)
                            added_duration = float(evaluation.get("raw_duration", math.inf)) - current_duration
                            projected_adjusted = float(evaluation.get("adjusted_duration", math.inf))
                            if not evaluation.get("valid"):
                                continue
                            rider_minimum_target = _rider_minimum_job_target(rider, minimum_job_target)
                            minimum_priority = _minimum_job_priority(len(current_sequence), rider_minimum_target)
                            candidate_rank = (
                                minimum_priority,
                                added_duration,
                                projected_adjusted,
                                int(job.get("_original_order", job_index)),
                            )
                            if best_insertion is None or candidate_rank < best_insertion[:4]:
                                best_insertion = (
                                    minimum_priority,
                                    added_duration,
                                    projected_adjusted,
                                    int(job.get("_original_order", job_index)),
                                    rider,
                                    insert_at,
                                    job_index,
                                    job,
                                    evaluation,
                                )

                if best_insertion is not None:
                    _, _, _, _, rider, insert_at, job_index, job, _ = best_insertion
                    if _job_id(job) in assigned_job_ids:
                        remaining.pop(job_index)
                        continue
                    rider_sequences[rider_key(rider)].insert(insert_at, job)
                    assigned_job_ids.add(_job_id(job))
                    remaining.pop(job_index)
                    made_assignment = True

            if not remaining:
                break

    def rebalance_minimum_jobs() -> None:
        if minimum_job_target <= 0:
            return

        moved_job = True
        while moved_job:
            moved_job = False
            current_evaluations = {
                rider_key(rider): evaluate_rider_sequence(rider, rider_sequences[rider_key(rider)], cap_used)
                for rider in riders
            }
            underfilled_riders = [
                rider
                for rider in riders
                if len(rider_sequences[rider_key(rider)]) < _rider_minimum_job_target(rider, minimum_job_target)
            ]
            if not underfilled_riders:
                break

            best_transfer: tuple[float, float, int, RiderState, RiderState, int, int, dict[str, Any]] | None = None
            for receiver in underfilled_riders:
                receiver_key = rider_key(receiver)
                receiver_sequence = rider_sequences[receiver_key]
                receiver_current_duration = float(current_evaluations[receiver_key].get("raw_duration", 0) or 0)

                for donor in riders:
                    if donor.name == receiver.name:
                        continue

                    donor_key = rider_key(donor)
                    donor_sequence = rider_sequences[donor_key]
                    donor_target = _rider_minimum_job_target(donor, minimum_job_target)
                    if len(donor_sequence) <= donor_target:
                        continue

                    donor_current_duration = float(current_evaluations[donor_key].get("raw_duration", 0) or 0)
                    for remove_at, job in enumerate(donor_sequence):
                        donor_candidate_sequence = donor_sequence[:remove_at] + donor_sequence[remove_at + 1 :]
                        donor_evaluation = evaluate_rider_sequence(donor, donor_candidate_sequence, cap_used)
                        if not donor_evaluation.get("valid", True):
                            continue

                        donor_saved_duration = donor_current_duration - float(donor_evaluation.get("raw_duration", 0) or 0)
                        for insert_at in range(len(receiver_sequence) + 1):
                            receiver_candidate_sequence = (
                                receiver_sequence[:insert_at] + [job] + receiver_sequence[insert_at:]
                            )
                            receiver_evaluation = evaluate_rider_sequence(receiver, receiver_candidate_sequence, cap_used)
                            if not receiver_evaluation.get("valid"):
                                continue

                            receiver_added_duration = (
                                float(receiver_evaluation.get("raw_duration", math.inf)) - receiver_current_duration
                            )
                            transfer_rank = (
                                receiver_added_duration - donor_saved_duration,
                                float(receiver_evaluation.get("adjusted_duration", math.inf)),
                                int(job.get("_original_order", remove_at)),
                            )
                            if best_transfer is None or transfer_rank < best_transfer[:3]:
                                best_transfer = (
                                    *transfer_rank,
                                    donor,
                                    receiver,
                                    remove_at,
                                    insert_at,
                                    job,
                                )

            if best_transfer is None:
                break

            _, _, _, donor, receiver, remove_at, insert_at, job = best_transfer
            donor_key = rider_key(donor)
            receiver_key = rider_key(receiver)
            donor_sequence = rider_sequences[donor_key]
            receiver_sequence = rider_sequences[receiver_key]
            rider_sequences[donor_key] = donor_sequence[:remove_at] + donor_sequence[remove_at + 1 :]
            rider_sequences[receiver_key] = receiver_sequence[:insert_at] + [job] + receiver_sequence[insert_at:]
            moved_job = True

    rebalance_minimum_jobs()

    final_evaluations = {
        rider_key(rider): evaluate_rider_sequence(rider, rider_sequences[rider_key(rider)], cap_used)
        for rider in riders
    }
    route_rows = []
    lookup_warnings.extend(
        warning
        for evaluation in final_evaluations.values()
        for warning in evaluation.get("warnings", [])
    )
    for rider in riders:
        key = rider_key(rider)
        evaluation = final_evaluations[key]
        route_rows.extend(evaluation.get("rows", []))
        rider.assigned_count = len(rider_sequences[key])
        rider.empty_distance_km = float(evaluation.get("empty_distance", 0) or 0)
        rider.empty_duration_min = float(evaluation.get("empty_duration", 0) or 0)
        rider.loaded_distance_km = float(evaluation.get("loaded_distance", 0) or 0)
        rider.loaded_duration_min = float(evaluation.get("loaded_duration", 0) or 0)
        rider.current_location = clean_text(evaluation.get("final_location", base_rider_state[key]["start_location"]))
        rider.current_zone = evaluation.get("final_zone") or base_rider_state[key]["start_zone"]

    unassigned_details: list[dict[str, Any]] = []
    for job in remaining:
        candidate_rows = []
        cap_for_audit = cap_used
        for rider in riders:
            key = rider_key(rider)
            current_sequence = rider_sequences[key]
            current_evaluation = evaluate_rider_sequence(rider, current_sequence, cap_for_audit)
            current_raw_duration = float(current_evaluation.get("raw_duration", 0) or 0)
            current_location = (
                clean_text(current_sequence[-1]["Drop-off Address"])
                if current_sequence
                else base_rider_state[key]["start_location"]
            )
            for insert_at in range(len(current_sequence) + 1):
                candidate_sequence = current_sequence[:insert_at] + [job] + current_sequence[insert_at:]
                evaluation = evaluate_rider_sequence(rider, candidate_sequence, cap_for_audit)
                inserted_rows = [
                    row for row in evaluation.get("rows", [])
                    if int(row.get("Uploaded Row", -1)) == _job_uploaded_row(job)
                ]
                inserted_row = inserted_rows[0] if inserted_rows else {}
                candidate_rows.append(
                    {
                        "Car Plate": clean_text(job.get("Car Plate")),
                        "Pickup Address": clean_text(job.get("Pickup Address")),
                        "Drop-off Address": clean_text(job.get("Drop-off Address")),
                        "Candidate Rider": rider.name,
                        "Candidate Start/Current Location": current_location,
                        "Candidate Empty Duration Min": inserted_row.get("Empty Duration Min", ""),
                        "Candidate Loaded Duration Min": inserted_row.get("Loaded Duration Min", ""),
                        "Candidate Added Duration Min": round(
                            float(evaluation.get("raw_duration", 0) or 0) - current_raw_duration,
                            1,
                        ),
                        "Projected Raw Duration Min": round(float(evaluation.get("raw_duration", 0) or 0), 1),
                        "Projected Adjusted Duration Min": round(float(evaluation.get("adjusted_duration", 0) or 0), 1),
                        "Cap Used": cap_for_audit,
                        "Reason Rejected": evaluation.get("reason") or "Not selected by rescue insertion ranking",
                    }
                )

        candidate_rows = sorted(
            candidate_rows,
            key=lambda row: float(row["Projected Adjusted Duration Min"] or math.inf),
        )
        rejected_candidate_audit.extend(candidate_rows[:3])
        best = candidate_rows[0] if candidate_rows else {}
        best_adjusted = best.get("Projected Adjusted Duration Min", "")
        best_rider = best.get("Candidate Rider", "")
        best_location = best.get("Candidate Start/Current Location", "")
        unassigned_details.append(
            {
                "Uploaded Row": _job_uploaded_row(job),
                "Best Candidate Rider": best_rider,
                "Best Candidate Final Location": best_location,
                "Best Candidate Projected Raw Duration": best.get("Projected Raw Duration Min", ""),
                "Best Candidate Projected Adjusted Duration": best_adjusted,
                "Current Cap": cap_for_audit,
                "Reason": (
                    f"Could not assign. Best rejected rider was {best_rider}. "
                    f"Reason: {best.get('Reason Rejected', 'not selected')}. "
                    f"Projected in-window duration {best.get('Projected Raw Duration Min', '')} min, "
                    f"projected adjusted duration {best_adjusted} min."
                    if best
                    else "Could not assign. No rider candidate was available."
                ),
            }
        )

    route_df = format_route_output(pd.DataFrame(route_rows, columns=ROUTE_COLUMNS), riders)
    route_df.attrs["unassigned_details"] = unassigned_details
    route_df.attrs["rejected_candidate_audit"] = rejected_candidate_audit
    route_df.attrs["force_complete_cap_used"] = cap_used
    validate_optimisation_integrity(route_df, jobs)
    summary_df = pd.DataFrame(
        [
            {
                "Rider": rider.name,
                "Total Jobs": len({_job_id(job) for job in rider_sequences[rider_key(rider)]}),
                "Total Empty Distance KM": round(rider.empty_distance_km, 2),
                "Total Empty Duration Min": round(rider.empty_duration_min, 1),
                "Total Loaded Distance KM": round(rider.loaded_distance_km, 2),
                "Total Loaded Duration Min": round(rider.loaded_duration_min, 1),
                "Total Route Distance KM": round(rider.empty_distance_km + rider.loaded_distance_km, 2),
                "Total Route Duration Min": round(rider.empty_duration_min + rider.loaded_duration_min, 1),
                "Adjusted Route Duration Min": round(
                    (rider.empty_duration_min + rider.loaded_duration_min) * duration_buffer_multiplier,
                    1,
                ),
                "Within 3 Hours": _route_status_for_adjusted_duration(
                    (rider.empty_duration_min + rider.loaded_duration_min) * duration_buffer_multiplier,
                    DEFAULT_MAX_ADJUSTED_DURATION_MIN,
                ),
                "Final Location": rider.current_location,
            }
            for rider in riders
        ],
        columns=SUMMARY_COLUMNS,
    )
    summary_df = format_summary_output(summary_df, route_df)
    report("Finished route optimisation", phase="Finished", progress=1.0)
    return route_df, summary_df, _dedupe_lookup_warnings(lookup_warnings)


def validate_route_chain(route_df: pd.DataFrame, riders: list[RiderState]) -> dict[tuple[str, int], str]:
    validation: dict[tuple[str, int], str] = {}
    start_locations = {rider.name: rider.start_location for rider in riders}

    for rider_name, rider_routes in route_df.groupby("Rider", sort=False):
        expected_start = start_locations.get(rider_name, "")
        for _, row in rider_routes.sort_values("Sequence").iterrows():
            sequence = int(row["Sequence"])
            actual_start = clean_text(row["Start From"])
            if actual_start.casefold() == clean_text(expected_start).casefold():
                validation[(rider_name, sequence)] = "OK"
            else:
                validation[(rider_name, sequence)] = f"Check: expected start from {expected_start}"
            expected_start = clean_text(row["Drop-off Address"])

    return validation


def format_route_output(route_df: pd.DataFrame, riders: list[RiderState]) -> pd.DataFrame:
    if route_df.empty:
        return pd.DataFrame(columns=ROUTE_COLUMNS)
    validation = validate_route_chain(route_df, riders)
    route_df = route_df.copy()
    route_df["Route Validation Status"] = route_df.apply(
        lambda row: validation.get((row["Rider"], int(row["Sequence"])), "Check route chain"),
        axis=1,
    )
    for column in ROUTE_COLUMNS:
        if column not in route_df.columns:
            route_df[column] = ""
    return route_df.loc[:, ROUTE_COLUMNS]


def format_summary_output(summary_df: pd.DataFrame, route_df: pd.DataFrame) -> pd.DataFrame:
    if summary_df.empty:
        return pd.DataFrame(columns=SUMMARY_COLUMNS)

    summary_df = summary_df.copy()
    if not route_df.empty and "Uploaded Row" in route_df.columns and "Rider" in route_df.columns:
        unique_jobs_by_rider = (
            route_df.dropna(subset=["Uploaded Row"])
            .assign(**{"Uploaded Row": pd.to_numeric(route_df["Uploaded Row"], errors="coerce")})
            .dropna(subset=["Uploaded Row"])
            .drop_duplicates(subset=["Rider", "Uploaded Row"])
            .groupby("Rider")["Uploaded Row"]
            .count()
        )
        summary_df["Total Jobs"] = summary_df["Rider"].map(unique_jobs_by_rider).fillna(0).astype(int)
    if "Adjusted Route Duration Min" not in summary_df.columns or summary_df["Adjusted Route Duration Min"].isna().all():
        summary_df["Adjusted Route Duration Min"] = (
            summary_df["Total Route Duration Min"].fillna(0).astype(float) * DEFAULT_DURATION_BUFFER_MULTIPLIER
        ).round(1)
    summary_df["Total Route Duration Hours"] = (
        summary_df["Total Route Duration Min"].fillna(0).astype(float) / 60
    ).round(2)
    if "Within 3 Hours" not in summary_df.columns or summary_df["Within 3 Hours"].isna().all():
        summary_df["Within 3 Hours"] = summary_df["Adjusted Route Duration Min"].apply(
            lambda value: "OK" if float(value or 0) <= DEFAULT_MAX_ADJUSTED_DURATION_MIN else "Fail"
        )

    total_duration = summary_df["Total Route Duration Min"].fillna(0).astype(float)
    summary_df["Empty Travel %"] = summary_df.apply(
        lambda row: round(
            (float(row["Total Empty Duration Min"] or 0) / float(row["Total Route Duration Min"])) * 100,
            1,
        )
        if float(row["Total Route Duration Min"] or 0) > 0
        else 0,
        axis=1,
    )
    summary_df["Loaded Travel %"] = summary_df.apply(
        lambda row: round(
            (float(row["Total Loaded Duration Min"] or 0) / float(row["Total Route Duration Min"])) * 100,
            1,
        )
        if float(row["Total Route Duration Min"] or 0) > 0
        else 0,
        axis=1,
    )

    active_durations = total_duration[summary_df["Total Jobs"].fillna(0).astype(int) > 0]
    average_duration = float(active_durations.mean()) if not active_durations.empty else 0

    fallback_share_by_rider: dict[str, float] = {}
    if not route_df.empty:
        for rider, rider_routes in route_df.groupby("Rider"):
            fallback_rows = rider_routes["Cost Source"].astype(str).str.contains("fallback", case=False, na=False)
            fallback_share_by_rider[str(rider)] = float(fallback_rows.mean()) if len(fallback_rows) else 0

    comments = []
    for _, row in summary_df.iterrows():
        rider = clean_text(row["Rider"])
        duration = float(row["Total Route Duration Min"] or 0)
        adjusted_duration = float(row["Adjusted Route Duration Min"] or 0)
        jobs = int(row["Total Jobs"] or 0)
        fallback_share = fallback_share_by_rider.get(rider, 0)

        if jobs == 0:
            comment = "No jobs assigned"
        elif adjusted_duration > DEFAULT_MAX_ADJUSTED_DURATION_MIN:
            comment = "Fails 3-hour adjusted cap"
        elif fallback_share >= 0.5:
            comment = "Estimate only - verify if needed"
        elif average_duration > 0 and duration > average_duration * 1.2:
            comment = "Heavier route"
        elif average_duration > 0 and duration < average_duration * 0.8:
            comment = "Lighter route"
        else:
            comment = "Balanced"
        comments.append(comment)

    summary_df["Workload Comment"] = comments
    return summary_df.loc[:, SUMMARY_COLUMNS]


def _write_title(ws: Any, title: str, row: int = 1) -> None:
    ws.cell(row=row, column=1, value=_excel_safe_value(title))
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1F2937")


def _excel_safe_value(value: Any) -> Any:
    if isinstance(value, str) and len(value) > EXCEL_CELL_MAX_CHARS:
        suffix = "\n\n[Truncated for Excel cell limit]"
        return value[: EXCEL_CELL_MAX_CHARS - len(suffix)] + suffix
    return value


def _excel_safe_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    safe_df = df.copy()
    for column in safe_df.columns:
        safe_df[column] = safe_df[column].map(_excel_safe_value)
    return safe_df


def _write_rows(ws: Any, rows: list[list[Any]], start_row: int = 1) -> int:
    for row_offset, values in enumerate(rows):
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=start_row + row_offset, column=column, value=_excel_safe_value(value))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    return start_row + len(rows)


def _style_header_row(ws: Any, row: int, max_column: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for column in range(1, max_column + 1):
        cell = ws.cell(row=row, column=column)
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def _autosize_columns(ws: Any, max_width: int = 45) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 10
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, min(len(str(cell.value)), max_width))
        ws.column_dimensions[column_letter].width = max_length + 2


def _style_route_sheet(ws: Any, header_row: int, table_start_row: int, table_end_row: int) -> None:
    _style_header_row(ws, header_row, len(ROUTE_COLUMNS))
    ws.freeze_panes = f"A{header_row + 1}"
    address_columns = {
        "Start From",
        "Empty Travel To Pickup",
        "Pickup Address",
        "Drop-off Address",
        "Loaded Travel / Car Movement",
    }
    header_map = {ws.cell(header_row, col).value: col for col in range(1, len(ROUTE_COLUMNS) + 1)}
    previous_rider = None
    rider_fill_a = PatternFill("solid", fgColor="F8FAFC")
    rider_fill_b = PatternFill("solid", fgColor="EEF6FF")
    fallback_fill = PatternFill("solid", fgColor="FFE7A3")
    onemap_fill = PatternFill("solid", fgColor="C6EFCE")
    ok_fill = PatternFill("solid", fgColor="C6EFCE")
    warning_fill = PatternFill("solid", fgColor="F4B183")
    rider_band = 0

    for row in range(table_start_row, table_end_row + 1):
        rider = ws.cell(row, header_map["Rider"]).value
        if rider != previous_rider:
            rider_band += 1
            previous_rider = rider
        row_fill = rider_fill_a if rider_band % 2 else rider_fill_b
        for col in range(1, len(ROUTE_COLUMNS) + 1):
            cell = ws.cell(row, col)
            cell.fill = row_fill
            cell.alignment = Alignment(
                wrap_text=ws.cell(header_row, col).value in address_columns,
                vertical="top",
            )

        cost_cell = ws.cell(row, header_map["Cost Source"])
        if "fallback" in str(cost_cell.value).lower():
            cost_cell.fill = fallback_fill
        elif "onemap" in str(cost_cell.value).lower():
            cost_cell.fill = onemap_fill

        validation_cell = ws.cell(row, header_map["Route Validation Status"])
        validation_cell.fill = ok_fill if validation_cell.value == "OK" else warning_fill


def _style_summary_sheet(ws: Any, rider_summary_header_row: int) -> None:
    _style_header_row(ws, 3, 2)
    _style_header_row(ws, rider_summary_header_row, len(SUMMARY_COLUMNS))
    ws.freeze_panes = f"A{rider_summary_header_row + 1}"
    for row in range(rider_summary_header_row + 1, ws.max_row + 1):
        comment = str(ws.cell(row, SUMMARY_COLUMNS.index("Workload Comment") + 1).value or "")
        if "Heavier" in comment:
            fill = PatternFill("solid", fgColor="F4B183")
        elif "Lighter" in comment:
            fill = PatternFill("solid", fgColor="D9EAD3")
        elif "Estimate only" in comment:
            fill = PatternFill("solid", fgColor="FFE7A3")
        else:
            fill = PatternFill("solid", fgColor="E2F0D9")
        ws.cell(row, SUMMARY_COLUMNS.index("Workload Comment") + 1).fill = fill


def _overall_summary(route_df: pd.DataFrame, summary_df: pd.DataFrame, jobs_df: pd.DataFrame | None) -> list[list[Any]]:
    active_summary = summary_df[summary_df["Total Jobs"].fillna(0).astype(int) > 0]
    total_duration = float(summary_df["Total Route Duration Min"].fillna(0).sum()) if not summary_df.empty else 0
    average_duration = float(active_summary["Total Route Duration Min"].mean()) if not active_summary.empty else 0
    longest = ""
    shortest = ""
    if not active_summary.empty:
        longest_row = active_summary.loc[active_summary["Total Route Duration Min"].idxmax()]
        shortest_row = active_summary.loc[active_summary["Total Route Duration Min"].idxmin()]
        longest = f"{longest_row['Rider']} ({longest_row['Total Route Duration Min']} min)"
        shortest = f"{shortest_row['Rider']} ({shortest_row['Total Route Duration Min']} min)"

    cost_source = route_df["Cost Source"].astype(str) if not route_df.empty else pd.Series(dtype=str)
    onemap_count = int(cost_source.str.contains("OneMap", case=False, na=False).sum())
    fallback_count = int(cost_source.str.contains("fallback", case=False, na=False).sum())
    uploaded_jobs = (
        int(jobs_df.attrs.get("uploaded_count", len(jobs_df)))
        if jobs_df is not None
        else "Not provided"
    )
    fallback_warning = (
        "Warning: most route costs are fallback estimates."
        if fallback_count > onemap_count and fallback_count > 0
        else "Cost source mix looks acceptable."
    )

    return [
        ["Metric", "Value"],
        ["Total riders used", int((summary_df["Total Jobs"].fillna(0).astype(int) > 0).sum())],
        ["Total jobs uploaded", uploaded_jobs],
        ["Total jobs assigned", len(route_df)],
        ["Total estimated route duration", f"{round(total_duration, 1)} min / {round(total_duration / 60, 2)} hours"],
        ["Average duration per rider", f"{round(average_duration, 1)} min"],
        ["Longest rider route", longest],
        ["Shortest rider route", shortest],
        ["Number of OneMap route calculations", onemap_count],
        ["Number of fallback estimates used", fallback_count],
        ["Fallback warning", fallback_warning],
    ]


def _data_quality_rows(
    route_df: pd.DataFrame,
    jobs_df: pd.DataFrame | None,
    validation_warnings: list[str] | None,
    lookup_warnings: list[str] | None,
) -> list[list[Any]]:
    uploaded_jobs = len(jobs_df) if jobs_df is not None else "Not provided"
    blank_address_count = 0
    duplicate_plates = []
    if jobs_df is not None and not jobs_df.empty:
        uploaded_jobs = int(jobs_df.attrs.get("uploaded_count", len(jobs_df)))
        blank_address_count = int(jobs_df.attrs.get("blank_address_rows_dropped", 0))
        duplicate_plates = list(jobs_df.attrs.get("duplicate_plate_values", []))
        if not duplicate_plates:
            plates = jobs_df["Car Plate"].apply(clean_text)
            duplicate_plates = sorted(plates[plates.ne("") & plates.duplicated(keep=False)].unique())

    fallback_routes = route_df[route_df["Cost Source"].astype(str).str.contains("fallback", case=False, na=False)]
    validation_failures = route_df[route_df["Route Validation Status"].ne("OK")]
    geocode_failures = [
        warning
        for warning in (lookup_warnings or [])
        if "geocod" in warning.lower() or "no onemap" in warning.lower()
    ]

    action = "No major issues found. Routes can be reviewed and sent to riders."
    if not validation_failures.empty:
        action = "Do not send this route yet. Fix route chaining issue first."
    elif not fallback_routes.empty:
        action = "Some routes used fallback estimates. Review these manually if accuracy is important."

    rows = [
        ["Check", "Result"],
        ["Number of jobs uploaded", uploaded_jobs],
        ["Number of jobs assigned", len(route_df)],
        ["Number of unassigned jobs", max(0, int(uploaded_jobs) - len(route_df)) if isinstance(uploaded_jobs, int) else "Not provided"],
        ["Duplicate car plates detected", ", ".join(duplicate_plates) if duplicate_plates else "None detected"],
        ["Blank pickup/drop-off addresses detected", blank_address_count],
        ["Addresses that failed OneMap geocoding", "\n".join(geocode_failures) if geocode_failures else "None detected"],
        ["Routes that used fallback estimates", len(fallback_routes)],
        ["Route validation failures", len(validation_failures)],
        ["Validation/input warnings", "\n".join(validation_warnings or []) if validation_warnings else "None"],
        ["User Action Required", action],
    ]
    return rows


def build_unassigned_jobs_df(jobs_df: pd.DataFrame | None, route_df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Uploaded Row",
        "Car Plate",
        "Pickup Address",
        "Pickup Lot",
        "Drop-off Address",
        "Pickup Zone",
        "Drop-off Zone",
        "Best Candidate Rider",
        "Best Candidate Final Location",
        "Best Candidate Projected Raw Duration",
        "Best Candidate Projected Adjusted Duration",
        "Current Cap",
        "Reason",
    ]
    if jobs_df is None or jobs_df.empty:
        return pd.DataFrame(columns=columns)

    jobs = jobs_df.copy()
    if "_original_order" not in jobs.columns:
        jobs["_original_order"] = range(len(jobs))
    jobs["Uploaded Row"] = jobs["_original_order"].astype(int) + 2

    assigned_rows: set[int] = set()
    if route_df is not None and not route_df.empty and "Uploaded Row" in route_df.columns:
        assigned_rows = set(route_df["Uploaded Row"].dropna().astype(int).tolist())

    unassigned = jobs[~jobs["Uploaded Row"].isin(assigned_rows)].copy()
    if unassigned.empty:
        return pd.DataFrame(columns=columns)

    details = {}
    if route_df is not None:
        details = {
            int(detail.get("Uploaded Row")): detail
            for detail in route_df.attrs.get("unassigned_details", [])
            if detail.get("Uploaded Row") is not None
        }

    for detail_column in [
        "Best Candidate Rider",
        "Best Candidate Final Location",
        "Best Candidate Projected Raw Duration",
        "Best Candidate Projected Adjusted Duration",
        "Current Cap",
        "Reason",
    ]:
        unassigned[detail_column] = unassigned["Uploaded Row"].apply(
            lambda row_number: details.get(int(row_number), {}).get(detail_column, "")
        )
    unassigned["Reason"] = unassigned["Reason"].replace(
        "",
        "Could not assign under current cap. No candidate audit was available.",
    )
    for column in columns:
        if column not in unassigned.columns:
            unassigned[column] = ""
    return unassigned.loc[:, columns].reset_index(drop=True)


def _write_how_to_read_sheet(writer: pd.ExcelWriter) -> None:
    wb = writer.book
    ws = wb.create_sheet("How To Read This", 0)
    _write_title(ws, "How To Read This Export", 1)
    rows = [
        ["What this file is for", "This file assigns vehicle relocation jobs to riders and sequences them."],
        ["How each route works", "Each row has two legs: Empty Travel To Pickup and Loaded Travel / Car Movement."],
        ["Empty Travel To Pickup", "The rider travels without the car from the current location to the pickup address."],
        ["Loaded Travel / Car Movement", "The rider drives the assigned car from pickup address to drop-off address."],
        ["Important route logic", "After each drop-off, the rider's new starting point becomes the previous drop-off address. The rider does not reset to the original start location."],
        ["Example", "If Rider D drops a car at 150 Tampines Street 12, the next row for Rider D should start from 150 Tampines Street 12."],
        ["Rider", "Assigned rider or driver."],
        ["Sequence", "Order that rider should follow."],
        ["Start From", "Where the rider starts this job from."],
        ["Pickup Address / Pickup Lot", "Where the car is collected, including parking lot or deck information."],
        ["Drop-off Address", "Where the car must be moved to."],
        ["Empty / Loaded Distance and Duration", "Empty is before collecting the car. Loaded is while moving the assigned car. Total is empty plus loaded."],
        ["Cost Source", "OneMap means distance/time came from OneMap. fallback estimate means rough zone-based estimates were used because OneMap was unavailable."],
        ["Route Validation Status", "OK means sequence 1 starts from the rider start location, and later rows start from the previous drop-off address."],
        ["Fallback estimates", "Treat fallback estimates as planning estimates, not exact travel times."],
    ]
    _write_rows(ws, rows, 3)
    _style_header_row(ws, 3, 2)
    _autosize_columns(ws, 70)


def _rider_cost_accuracy_note(rider_routes: pd.DataFrame) -> str:
    if rider_routes.empty:
        return "No assigned route."
    cost_sources = rider_routes["Cost Source"].astype(str)
    fallback_count = int(cost_sources.str.contains("fallback", case=False, na=False).sum())
    if fallback_count == 0:
        return "Based on OneMap distance/time estimates."
    if fallback_count == len(rider_routes):
        return "Timing is rough estimate only."
    return "Some legs use fallback estimates. Timing should be manually reviewed."


def _route_validation_summary(rider_routes: pd.DataFrame) -> str:
    if rider_routes.empty:
        return "No route"
    return "OK" if rider_routes["Route Validation Status"].eq("OK").all() else "Check route chain"


def _cost_source_summary(rider_routes: pd.DataFrame) -> str:
    if rider_routes.empty:
        return "No route"
    cost_sources = rider_routes["Cost Source"].astype(str)
    fallback_count = int(cost_sources.str.contains("fallback", case=False, na=False).sum())
    onemap_count = int(cost_sources.str.contains("OneMap", case=False, na=False).sum())
    if fallback_count == 0:
        return "All OneMap"
    if onemap_count == 0:
        return "All fallback estimates"
    return f"Mixed: {onemap_count} OneMap, {fallback_count} fallback"


def _manager_note(rider_routes: pd.DataFrame) -> str:
    if _route_validation_summary(rider_routes) != "OK":
        return "Do not dispatch until fixed"
    if rider_routes["Cost Source"].astype(str).str.contains("fallback", case=False, na=False).any():
        return "Review timing manually"
    return "Ready to dispatch"


def _route_value(route: pd.Series, columns: list[str] | str, default: str = "-") -> str:
    if isinstance(columns, str):
        columns = [columns]
    for column in columns:
        if column in route.index:
            value = clean_text(route.get(column))
            if value:
                return value
    return default


def _sort_routes_for_export(route_df: pd.DataFrame) -> pd.DataFrame:
    if route_df.empty:
        return route_df

    sort_df = route_df.copy()
    sort_df["_export_original_order"] = range(len(sort_df))
    sort_df["_export_rider_sort"] = sort_df["Rider"].apply(clean_text) if "Rider" in sort_df.columns else "-"
    if "Sequence" in sort_df.columns:
        sort_df["_export_sequence_sort"] = pd.to_numeric(sort_df["Sequence"], errors="coerce")
    else:
        sort_df["_export_sequence_sort"] = pd.NA
    sort_df["_export_sequence_sort"] = sort_df["_export_sequence_sort"].fillna(sort_df["_export_original_order"] + 1)
    sort_df = sort_df.sort_values(
        ["_export_rider_sort", "_export_sequence_sort", "_export_original_order"],
        kind="stable",
    )
    return sort_df.drop(
        columns=["_export_original_order", "_export_rider_sort", "_export_sequence_sort"],
        errors="ignore",
    )


def _build_whatsapp_message(rider: str, rider_routes: pd.DataFrame) -> str:
    message_parts = [
        "*PLEASE REACT TO THE MESSAGE OR REPLY TO ACKNOWLEDGE YOUR JOBS.*",
        "",
        f"Hi {rider}, please complete the following vehicle movement job(s).",
        "",
        "Please take clear photos at each pickup and drop-off, including:",
        "",
        "* front, back, left, right of the car",
        "* car plate",
        "* fuel level",
        "* parking lot/location",
        "* visible vehicle condition issues",
        "",
        "After pickup photos are taken, please send and reply:",
        "\u201cJob [number] pickup photos sent. Car plate [Car Plate], Lot [Lot Number]. Please open/unlock the car.\u201d",
    ]

    for job_number, (_, route) in enumerate(rider_routes.iterrows(), start=1):
        car_plate = _route_value(route, "Car Plate")
        pickup_address = _route_value(route, ["Pickup Address", "Empty Travel To Pickup"])
        pickup_lot = _route_value(route, "Pickup Lot")
        dropoff_address = _route_value(route, "Drop-off Address")
        lot_range = _route_value(route, ["Drop-off Lot", "Lot Range", "Drop-off Lot Range"])
        zone = _route_value(route, ["Drop-off Zone", "Zone"])

        job_lines = [
            f"*Job {job_number}*",
            f"Car Plate: {car_plate}",
            f"Pickup: {pickup_address}",
            f"Pickup Lot: {pickup_lot}",
            f"Drop-off: {dropoff_address}",
        ]
        if lot_range != "-":
            job_lines.append(f"Lot Range: {lot_range}")
        if zone != "-":
            job_lines.append(f"Zone: {zone}")
        message_parts.extend(["", *job_lines])

    message_parts.extend(
        [
            "",
            "Once each job is completed, please send the drop-off photos and reply:",
            "\u201cJob [number] completed. Car plate [Car Plate]. Parked at Lot [Lot Number].\u201d",
        ]
    )
    return "\n".join(message_parts)


def _build_route_table_text(rider_routes: pd.DataFrame) -> str:
    lines = ["Step | Start From | Go To Pickup | Collect Car Plate | Pickup Lot | Drop-off At"]
    for fallback_step, (_, route) in enumerate(rider_routes.iterrows(), start=1):
        step = _route_value(route, "Sequence", str(fallback_step))
        start_from = _route_value(route, "Start From")
        pickup_address = _route_value(route, ["Pickup Address", "Empty Travel To Pickup"])
        car_plate = _route_value(route, "Car Plate")
        pickup_lot = _route_value(route, "Pickup Lot")
        dropoff_address = _route_value(route, "Drop-off Address")
        lines.append(
            f"{step} | {start_from} | {pickup_address} | {car_plate} | {pickup_lot} | {dropoff_address}"
        )
    return "\n".join(lines)


def _write_rider_instructions_sheet(writer: pd.ExcelWriter, route_df: pd.DataFrame, summary_df: pd.DataFrame) -> None:
    ws = writer.book.create_sheet("Rider Instructions")
    headers = ["Rider", "WhatsApp Message", "Route Table"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=_excel_safe_value(header))
    _style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    if route_df.empty:
        ws.cell(row=2, column=1, value=_excel_safe_value("No rider routes were assigned."))
        for col in range(1, 4):
            ws.cell(row=2, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 80
        ws.column_dimensions["C"].width = 120
        return

    sorted_routes = _sort_routes_for_export(route_df).reset_index(drop=True)
    rider_groups = (
        sorted_routes["Rider"].apply(clean_text)
        if "Rider" in sorted_routes.columns
        else pd.Series(["-"] * len(sorted_routes), index=sorted_routes.index)
    )
    row_idx = 2
    for rider, rider_routes in sorted_routes.groupby(rider_groups, sort=False):
        rider_name = clean_text(rider) or "-"
        row_values = [
            rider_name,
            _build_whatsapp_message(rider_name, rider_routes),
            _build_route_table_text(rider_routes),
        ]
        for col, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=_excel_safe_value(value))
            cell.alignment = Alignment(wrap_text=col in {2, 3}, vertical="top")
        row_idx += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 120


def _write_manager_dispatch_summary_sheet(
    writer: pd.ExcelWriter,
    route_df: pd.DataFrame,
    summary_df: pd.DataFrame,
) -> None:
    ws = writer.book.create_sheet("Manager Dispatch Summary")
    _write_title(ws, "Manager Dispatch Summary", 1)
    ws.cell(
        row=2,
        column=1,
        value=_excel_safe_value("One-row dispatch view for managers. Use Manager Notes before sending routes to riders."),
    )
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    headers = [
        "Rider",
        "Cars Assigned",
        "Start Location",
        "First Pickup",
        "Final Drop-off",
        "Total Estimated Duration Min",
        "Total Estimated Distance KM",
        "Route Validation Status",
        "Cost Source Summary",
        "Manager Notes",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=_excel_safe_value(header))
    _style_header_row(ws, 4, len(headers))
    ws.freeze_panes = "A5"

    summary_by_rider = summary_df.set_index("Rider") if not summary_df.empty else pd.DataFrame()
    row_idx = 5
    if route_df.empty:
        ws.cell(row=row_idx, column=1, value=_excel_safe_value("No rider routes were assigned."))
    else:
        for rider, rider_routes in route_df.sort_values(["Rider", "Sequence"]).groupby("Rider", sort=False):
            rider_summary = summary_by_rider.loc[rider] if rider in summary_by_rider.index else None
            first_route = rider_routes.sort_values("Sequence").iloc[0]
            final_route = rider_routes.sort_values("Sequence").iloc[-1]
            total_duration = float(rider_routes["Total Duration Min"].fillna(0).sum())
            total_distance = float(rider_routes["Total Distance KM"].fillna(0).sum())
            final_dropoff = clean_text(final_route["Drop-off Address"])
            if rider_summary is not None:
                total_duration = float(rider_summary.get("Total Route Duration Min", total_duration) or 0)
                total_distance = float(rider_summary.get("Total Route Distance KM", total_distance) or 0)
                final_dropoff = clean_text(rider_summary.get("Final Location", final_dropoff))

            values = [
                rider,
                len(rider_routes),
                clean_text(first_route["Start From"]),
                clean_text(first_route["Pickup Address"]),
                final_dropoff,
                round(total_duration, 1),
                round(total_distance, 2),
                _route_validation_summary(rider_routes),
                _cost_source_summary(rider_routes),
                _manager_note(rider_routes),
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=_excel_safe_value(value))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_idx += 1

    note_col = headers.index("Manager Notes") + 1
    for row in range(5, ws.max_row + 1):
        note = str(ws.cell(row, note_col).value or "")
        if "Ready" in note:
            fill = PatternFill("solid", fgColor="C6EFCE")
        elif "Review" in note:
            fill = PatternFill("solid", fgColor="FFE7A3")
        else:
            fill = PatternFill("solid", fgColor="F4B183")
        ws.cell(row, note_col).fill = fill
    _autosize_columns(ws, 45)


def _write_unassigned_jobs_sheet(
    writer: pd.ExcelWriter,
    jobs_df: pd.DataFrame | None,
    route_df: pd.DataFrame,
) -> None:
    ws = writer.book.create_sheet("Unassigned Jobs")
    _write_title(ws, "Unassigned Jobs", 1)
    ws.cell(
        row=2,
        column=1,
        value=_excel_safe_value("Uploaded jobs that were not assigned to any rider. Use the Uploaded Row number to find the job in the original file."),
    )
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    unassigned_df = build_unassigned_jobs_df(jobs_df, route_df)
    if unassigned_df.empty:
        ws.cell(row=4, column=1, value=_excel_safe_value("All uploaded valid jobs were assigned."))
        _autosize_columns(ws, 45)
        return

    header_row = 4
    for col, column_name in enumerate(unassigned_df.columns, start=1):
        ws.cell(row=header_row, column=col, value=_excel_safe_value(column_name))
    _style_header_row(ws, header_row, len(unassigned_df.columns))

    for row_idx, row in enumerate(unassigned_df.itertuples(index=False), start=header_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_excel_safe_value(value))
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A5"
    _autosize_columns(ws, 55)


def _write_rejected_candidate_audit_sheet(writer: pd.ExcelWriter, route_df: pd.DataFrame) -> None:
    ws = writer.book.create_sheet("Rejected Candidate Audit")
    _write_title(ws, "Rejected Candidate Audit", 1)
    ws.cell(
        row=2,
        column=1,
        value=_excel_safe_value("Top rejected rider options for jobs that remained unassigned after normal and rescue assignment passes."),
    )
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    audit_df = pd.DataFrame(route_df.attrs.get("rejected_candidate_audit", []))
    if audit_df.empty:
        ws.cell(row=4, column=1, value=_excel_safe_value("No rejected candidate audit rows for this run."))
        _autosize_columns(ws, 45)
        return

    header_row = 4
    for col, column_name in enumerate(audit_df.columns, start=1):
        ws.cell(row=header_row, column=col, value=_excel_safe_value(column_name))
    _style_header_row(ws, header_row, len(audit_df.columns))
    for row_idx, row in enumerate(audit_df.itertuples(index=False), start=header_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_excel_safe_value(value))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A5"
    _autosize_columns(ws, 55)


def export_routes_to_excel(
    route_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    jobs_df: pd.DataFrame | None = None,
    validation_warnings: list[str] | None = None,
    lookup_warnings: list[str] | None = None,
) -> bytes:
    route_df = _sort_routes_for_export(route_df).reset_index(drop=True) if not route_df.empty else route_df
    if jobs_df is not None:
        validate_optimisation_integrity(route_df, jobs_df)
    summary_df = format_summary_output(summary_df, route_df)
    export_route_df = _excel_safe_dataframe(route_df)
    export_summary_df = _excel_safe_dataframe(summary_df)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _write_how_to_read_sheet(writer)

        route_note = (
            "Follow each rider's rows in sequence order. For each row, the rider first travels from "
            "Start From to the Pickup Address without the car. After collecting the car, the rider "
            "moves it from Pickup Address to Drop-off Address. The next row starts from the previous "
            "drop-off location."
        )
        export_route_df.to_excel(writer, sheet_name="Optimised Routes", index=False, startrow=4)
        route_ws = writer.sheets["Optimised Routes"]
        _write_title(route_ws, "Optimised Routes", 1)
        route_ws.cell(row=2, column=1, value=_excel_safe_value(route_note))
        route_ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        route_ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=min(len(ROUTE_COLUMNS), 8))
        if not route_df.empty:
            _style_route_sheet(route_ws, header_row=5, table_start_row=6, table_end_row=5 + len(route_df))
        _autosize_columns(route_ws, 45)
        route_header_map = {route_ws.cell(5, col).value: col for col in range(1, route_ws.max_column + 1)}
        for hidden_column in ["Empty Route Path", "Loaded Route Path"]:
            if hidden_column in route_header_map:
                route_ws.column_dimensions[get_column_letter(route_header_map[hidden_column])].hidden = True

        _write_unassigned_jobs_sheet(writer, jobs_df, route_df)

        summary_ws = writer.book.create_sheet("Summary")
        _write_title(summary_ws, "Overall Route Summary", 1)
        overall_rows = _overall_summary(route_df, summary_df, jobs_df)
        next_row = _write_rows(summary_ws, overall_rows, 3) + 2
        _write_title(summary_ws, "Rider Workload Summary", next_row)
        rider_header_row = next_row + 2
        for col, column_name in enumerate(SUMMARY_COLUMNS, start=1):
            summary_ws.cell(row=rider_header_row, column=col, value=_excel_safe_value(column_name))
        for row_idx, row in enumerate(export_summary_df.itertuples(index=False), start=rider_header_row + 1):
            for col_idx, value in enumerate(row, start=1):
                summary_ws.cell(row=row_idx, column=col_idx, value=_excel_safe_value(value))
        _style_summary_sheet(summary_ws, rider_header_row)
        _autosize_columns(summary_ws, 45)

        _write_rider_instructions_sheet(writer, route_df, summary_df)
    return output.getvalue()


def optimise_routes(
    jobs: pd.DataFrame,
    riders: list[RiderState],
    use_onemap: bool = True,
    optimise_by: str = "duration",
) -> tuple[pd.DataFrame, pd.DataFrame]:
    route_df, summary_df, _ = optimise_vehicle_routes(
        jobs,
        riders,
        use_onemap=use_onemap,
        optimise_by=optimise_by,
        token=get_onemap_token(),
    )
    return route_df, summary_df


def build_excel_download(route_df: pd.DataFrame, summary_df: pd.DataFrame) -> bytes:
    return export_routes_to_excel(route_df, summary_df)
