from __future__ import annotations

import json
import math
import os
import re
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Callable
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REQUIRED_JOB_HEADERS = ["Car Plate", "Pickup Address", "Pickup Lot", "Drop-off Address"]
OPTIONAL_JOB_HEADERS = ["Date", "Fuel %", "Pickup Time", "Notes"]
RIDER_COLUMNS = ["Rider Name", "Start Location", "Start Zone", "Max Jobs"]
WEEKDAY_SHEETS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
ROSTER_FILE = Path(__file__).resolve().parent / "rider_roster.xlsx"

ROUTE_COLUMNS = [
    "Rider",
    "Sequence",
    "Uploaded Row",
    "Start From",
    "Empty Travel To Pickup",
    "Car Plate",
    "Pickup Address",
    "Pickup Lot",
    "Drop-off Address",
    "Loaded Travel / Car Movement",
    "Empty Distance KM",
    "Empty Duration Min",
    "Loaded Distance KM",
    "Loaded Duration Min",
    "Total Distance KM",
    "Total Duration Min",
    "Assignment Score",
    "Zone Adjustment",
    "Same Zone Pickup",
    "Same Zone Route",
    "Route Zone Priority",
    "Empty Weight",
    "Loaded Weight",
    "Workload Penalty",
    "Duration Penalty",
    "Max Jobs Penalty",
    "Projected Rider Duration Min",
    "Projected Adjusted Duration Min",
    "Cost Source",
    "Route Validation Status",
]
SUMMARY_COLUMNS = [
    "Rider",
    "Total Jobs",
    "Total Empty Distance KM",
    "Total Empty Duration Min",
    "Total Loaded Distance KM",
    "Total Loaded Duration Min",
    "Total Route Distance KM",
    "Total Route Duration Min",
    "Adjusted Route Duration Min",
    "Total Route Duration Hours",
    "Within 3 Hours",
    "Final Location",
    "Empty Travel %",
    "Loaded Travel %",
    "Workload Comment",
]

ZONE_KEYWORDS = {
    "North": ["Woodlands", "Admiralty", "Yishun", "Sembawang", "Canberra"],
    "North-East": ["Sengkang", "Punggol", "Hougang", "Serangoon"],
    "East": ["Pasir Ris", "Tampines", "Simei", "Bedok", "Paya Lebar", "PLQ"],
    "Central": ["Toa Payoh", "Bishan", "Ang Mo Kio", "Novena", "Kallang"],
    "West": ["Jurong", "Bukit Batok", "Choa Chu Kang", "Clementi", "Bukit Panjang"],
    "South/CBD": ["Orchard", "Marina", "Raffles", "Tanjong Pagar", "HarbourFront"],
}

FALLBACK_ZONE_MINUTES = {
    "North": {"North": 10, "North-East": 20, "East": 35, "Central": 25, "West": 32, "South/CBD": 38},
    "North-East": {"North": 20, "North-East": 10, "East": 22, "Central": 22, "West": 38, "South/CBD": 34},
    "East": {"North": 35, "North-East": 22, "East": 10, "Central": 25, "West": 45, "South/CBD": 32},
    "Central": {"North": 25, "North-East": 22, "East": 25, "Central": 10, "West": 25, "South/CBD": 16},
    "West": {"North": 32, "North-East": 38, "East": 45, "Central": 25, "West": 10, "South/CBD": 33},
    "South/CBD": {"North": 38, "North-East": 34, "East": 32, "Central": 16, "West": 33, "South/CBD": 10},
}

FALLBACK_ZONE_KM = {
    "North": {"North": 4, "North-East": 10, "East": 18, "Central": 13, "West": 16, "South/CBD": 20},
    "North-East": {"North": 10, "North-East": 4, "East": 11, "Central": 11, "West": 20, "South/CBD": 17},
    "East": {"North": 18, "North-East": 11, "East": 4, "Central": 13, "West": 25, "South/CBD": 17},
    "Central": {"North": 13, "North-East": 11, "East": 13, "Central": 4, "West": 13, "South/CBD": 7},
    "West": {"North": 16, "North-East": 20, "East": 25, "Central": 13, "West": 4, "South/CBD": 17},
    "South/CBD": {"North": 20, "North-East": 17, "East": 17, "Central": 7, "West": 17, "South/CBD": 4},
}

UNKNOWN_DISTANCE_KM = 18.0
UNKNOWN_DURATION_MIN = 40.0
SAME_UNKNOWN_DISTANCE_KM = 6.0
SAME_UNKNOWN_DURATION_MIN = 18.0
DEFAULT_EMPTY_WEIGHT = 4.5
DEFAULT_LOADED_WEIGHT = 0.8
DEFAULT_SOFT_WORKLOAD_MIN = 115.0
DEFAULT_WORKLOAD_PENALTY_PER_MIN = 2.0
DEFAULT_SOFT_ADJUSTED_DURATION_MIN = 165.0
DEFAULT_DURATION_PENALTY_PER_MIN = 2.0
DEFAULT_MAX_JOB_OVERAGE_PENALTY = 60.0
DEFAULT_DURATION_BUFFER_MULTIPLIER = 1.2
DEFAULT_MAX_ADJUSTED_DURATION_MIN = 180.0
DEFAULT_EMPTY_TRAVEL_DURATION_MULTIPLIER = 1.5
DEFAULT_EMPTY_TRAVEL_WAIT_BUFFER_MIN = 6.0
DEFAULT_CLUSTER_PRESSURE_BONUS_PER_JOB = 8.0
FORCE_COMPLETE_CAP_STAGES = [180.0, 195.0, 210.0]
ONEMAP_SEARCH_URL = "https://www.onemap.gov.sg/api/common/elastic/search"
ONEMAP_ROUTE_URL = "https://www.onemap.gov.sg/api/public/routingsvc/route"
CACHE_DIR = Path(__file__).resolve().parent / "cache"
PROJECT_ROOT = Path(__file__).resolve().parents[1]
ENV_FILE = PROJECT_ROOT / ".env"
GEOCODE_CACHE_FILE = CACHE_DIR / "onemap_geocode_cache.csv"
ROUTE_CACHE_FILE = CACHE_DIR / "onemap_route_cache.csv"
ProgressCallback = Callable[[dict[str, Any]], None]
GEOCODE_MEMORY_CACHE: dict[str, GeocodeResult] = {}
ROUTE_MEMORY_CACHE: dict[str, TravelCost] = {}


@dataclass
class RiderState:
    name: str
    start_location: str
    start_zone: str | None = None
    max_jobs: int | None = None
    current_location: str = ""
    current_zone: str | None = None
    assigned_count: int = 0
    empty_distance_km: float = 0
    empty_duration_min: float = 0
    loaded_distance_km: float = 0
    loaded_duration_min: float = 0

    @classmethod
    def from_row(cls, row: pd.Series) -> "RiderState":
        start_location = clean_text(row.get("Start Location"))
        start_zone = clean_text(row.get("Start Zone")) or infer_zone(start_location)
        return cls(
            name=clean_text(row.get("Rider Name")),
            start_location=start_location,
            start_zone=start_zone,
            max_jobs=parse_optional_int(row.get("Max Jobs")),
            current_location=start_location,
            current_zone=start_zone,
        )

    @property
    def has_capacity(self) -> bool:
        return self.max_jobs is None or self.assigned_count < self.max_jobs


@dataclass(frozen=True)
class GeocodeResult:
    address: str
    latitude: float | None
    longitude: float | None
    source: str
    error: str = ""

    @property
    def is_available(self) -> bool:
        return self.latitude is not None and self.longitude is not None


@dataclass(frozen=True)
class TravelCost:
    distance_km: float | None
    duration_min: float | None
    source: str
    error: str = ""

    def optimisation_value(self, optimise_by: str) -> float:
        if optimise_by == "distance":
            return self.distance_km if self.distance_km is not None else self.duration_min or math.inf
        return self.duration_min if self.duration_min is not None else self.distance_km or math.inf


def clean_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def parse_optional_int(value: Any) -> int | None:
    if pd.isna(value) or value == "":
        return None
    try:
        parsed = int(float(value))
    except (TypeError, ValueError):
        return None
    return parsed if parsed > 0 else None


def _read_csv_cache(path: Path, columns: list[str]) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(columns=columns)
    try:
        return pd.read_csv(path)
    except Exception:
        return pd.DataFrame(columns=columns)


def _write_csv_cache(path: Path, df: pd.DataFrame) -> None:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False)


def load_env_value(name: str, path: Path = ENV_FILE) -> str:
    value = os.getenv(name, "")
    if value:
        return value
    if not path.exists():
        return ""

    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except OSError:
        return ""

    for line in lines:
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, raw_value = stripped.split("=", 1)
        if key.strip() == name:
            return raw_value.strip().strip('"').strip("'")
    return ""


def get_onemap_token() -> str:
    return load_env_value("ONEMAP_TOKEN")


def infer_zone(address: Any) -> str | None:
    text = clean_text(address).lower()
    if not text:
        return None

    for zone, keywords in ZONE_KEYWORDS.items():
        if any(keyword.lower() in text for keyword in keywords):
            return zone
    return None


def clean_address_for_geocoding(address: Any) -> str:
    text = clean_text(address)
    if not text:
        return ""

    lowered = text.lower()
    named_location_aliases = {
        "plq entrance park place residences": "Park Place Residences PLQ",
        "park place residences": "Park Place Residences PLQ",
        "paya lebar quarter": "Paya Lebar Quarter",
        "plq mall": "PLQ Mall",
    }
    for keyword, alias in named_location_aliases.items():
        if keyword in lowered:
            return alias

    text = re.sub(r"\[[^\]]*\]", " ", text)
    text = re.sub(r",?\s+\bL\d+[A-Z]?\b.*$", " ", text, flags=re.IGNORECASE)
    text = re.sub(r",?\s+\bLevel\s+\d+[A-Z]?\b.*$", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\b(MSCP|Surface|Basement)\b.*$", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\bDeck\b.*$", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\bLot\b.*$", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\bNear\b.*$", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" ,")


def load_jobs_from_excel(uploaded_file: Any) -> pd.DataFrame:
    try:
        jobs = pd.read_excel(uploaded_file)
    except Exception as exc:
        raise ValueError(f"Unable to read the Excel file: {exc}") from exc
    jobs.columns = [clean_text(column) for column in jobs.columns]
    return jobs


def validate_jobs(jobs: pd.DataFrame) -> tuple[pd.DataFrame, list[str], list[str]]:
    warnings: list[str] = []
    uploaded_count = len(jobs)
    missing_headers = [header for header in REQUIRED_JOB_HEADERS if header not in jobs.columns]
    if missing_headers:
        return pd.DataFrame(), missing_headers, warnings

    keep_columns = REQUIRED_JOB_HEADERS + [header for header in OPTIONAL_JOB_HEADERS if header in jobs.columns]
    jobs = jobs.loc[:, keep_columns].copy()
    jobs["_original_order"] = range(len(jobs))

    before_drop = len(jobs)
    jobs = jobs[
        jobs["Pickup Address"].apply(clean_text).ne("")
        & jobs["Drop-off Address"].apply(clean_text).ne("")
    ].copy()
    dropped = before_drop - len(jobs)
    if dropped:
        warnings.append(f"Dropped {dropped} row(s) with blank pickup or drop-off address.")

    duplicate_plates = jobs["Car Plate"].apply(clean_text).loc[lambda series: series.ne("")].duplicated(keep=False)
    duplicate_plate_values = sorted(
        jobs["Car Plate"]
        .apply(clean_text)
        .loc[lambda series: series.ne("") & series.duplicated(keep=False)]
        .unique()
    )
    if duplicate_plates.any():
        warnings.append("Duplicate car plate values found. They are allowed, but please verify them.")

    jobs["Pickup Zone"] = jobs["Pickup Address"].apply(infer_zone)
    jobs["Drop-off Zone"] = jobs["Drop-off Address"].apply(infer_zone)
    jobs = jobs.reset_index(drop=True)
    jobs.attrs["uploaded_count"] = uploaded_count
    jobs.attrs["blank_address_rows_dropped"] = dropped
    jobs.attrs["duplicate_plate_values"] = duplicate_plate_values
    return jobs, [], warnings


def load_and_validate_jobs(uploaded_file: Any) -> tuple[pd.DataFrame, list[str], list[str]]:
    return validate_jobs(load_jobs_from_excel(uploaded_file))


def _fetch_json(url: str, params: dict[str, Any], token: str | None = None, timeout: int = 15) -> dict[str, Any]:
    headers = {"User-Agent": "Lance-BlueSG-Route-Optimiser/1.0"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    request = Request(f"{url}?{urlencode(params)}", headers=headers)
    with urlopen(request, timeout=timeout) as response:
        return json.loads(response.read().decode("utf-8"))


def geocode_address_onemap(address: str, token: str | None = None) -> GeocodeResult:
    address = clean_text(address)
    if not address:
        return GeocodeResult(address, None, None, "fallback estimate", "Blank address")

    search_values = [address]
    cleaned_address = clean_address_for_geocoding(address)
    if cleaned_address and cleaned_address.casefold() != address.casefold():
        search_values.append(cleaned_address)

    last_error = "No OneMap geocoding result"
    for search_value in search_values:
        try:
            payload = _fetch_json(
                ONEMAP_SEARCH_URL,
                {"searchVal": search_value, "returnGeom": "Y", "getAddrDetails": "Y", "pageNum": 1},
                token=token,
            )
        except (HTTPError, URLError, TimeoutError, json.JSONDecodeError, OSError) as exc:
            last_error = f"OneMap geocoding failed: {exc}"
            continue

        results = payload.get("results") or []
        if not results:
            last_error = "No OneMap geocoding result"
            continue

        best = results[0]
        try:
            latitude = float(best["LATITUDE"])
            longitude = float(best["LONGITUDE"])
        except (KeyError, TypeError, ValueError) as exc:
            last_error = f"Invalid OneMap geocode result: {exc}"
            continue

        source = "OneMap" if search_value == address else "OneMap cleaned address"
        return GeocodeResult(address, latitude, longitude, source)

    return GeocodeResult(address, None, None, "fallback estimate", last_error)


def get_cached_geocode(address: str, token: str | None = None, use_onemap: bool = True) -> GeocodeResult:
    address = clean_text(address)
    if not use_onemap:
        return GeocodeResult(address, None, None, "fallback estimate", "OneMap disabled")

    memory_key = address.casefold()
    if memory_key in GEOCODE_MEMORY_CACHE:
        cached = GEOCODE_MEMORY_CACHE[memory_key]
        if cached.is_available:
            return GeocodeResult(address, cached.latitude, cached.longitude, "OneMap cache", cached.error)

    cache = _read_csv_cache(GEOCODE_CACHE_FILE, ["address", "latitude", "longitude"])
    if not cache.empty:
        matches = cache[cache["address"].astype(str).str.casefold() == address.casefold()]
        if not matches.empty:
            row = matches.iloc[0]
            result = GeocodeResult(address, float(row["latitude"]), float(row["longitude"]), "OneMap cache")
            GEOCODE_MEMORY_CACHE[memory_key] = result
            return result

    result = geocode_address_onemap(address, token=token)
    GEOCODE_MEMORY_CACHE[memory_key] = result
    if result.is_available:
        cache = pd.concat(
            [
                cache,
                pd.DataFrame(
                    [{"address": address, "latitude": result.latitude, "longitude": result.longitude}]
                ),
            ],
            ignore_index=True,
        ).drop_duplicates(subset=["address"], keep="last")
        _write_csv_cache(GEOCODE_CACHE_FILE, cache)
    return result


def _parse_onemap_route(payload: dict[str, Any]) -> tuple[float | None, float | None]:
    summary = payload.get("route_summary") or payload.get("routeSummary") or {}
    distance_m = (
        summary.get("total_distance")
        or summary.get("totalDistance")
        or payload.get("total_distance")
        or payload.get("distance")
    )
    duration_s = (
        summary.get("total_time")
        or summary.get("totalTime")
        or payload.get("total_time")
        or payload.get("duration")
    )

    distance_km = None
    duration_min = None
    try:
        if distance_m is not None:
            distance_km = round(float(distance_m) / 1000, 2)
    except (TypeError, ValueError):
        distance_km = None
    try:
        if duration_s is not None:
            duration_min = round(float(duration_s) / 60, 1)
    except (TypeError, ValueError):
        duration_min = None
    return distance_km, duration_min


def get_onemap_route_cost(
    from_geocode: GeocodeResult,
    to_geocode: GeocodeResult,
    token: str | None = None,
) -> TravelCost:
    if not from_geocode.is_available or not to_geocode.is_available:
        missing = from_geocode.error or to_geocode.error or "Missing coordinates"
        return TravelCost(None, None, "fallback estimate", missing)

    key = (
        f"{from_geocode.latitude:.6f},{from_geocode.longitude:.6f}|"
        f"{to_geocode.latitude:.6f},{to_geocode.longitude:.6f}"
    )
    if key in ROUTE_MEMORY_CACHE:
        cached = ROUTE_MEMORY_CACHE[key]
        return TravelCost(cached.distance_km, cached.duration_min, "OneMap cache", cached.error)

    cache = _read_csv_cache(ROUTE_CACHE_FILE, ["route_key", "distance_km", "duration_min"])
    if not cache.empty:
        matches = cache[cache["route_key"] == key]
        if not matches.empty:
            row = matches.iloc[0]
            result = TravelCost(float(row["distance_km"]), float(row["duration_min"]), "OneMap cache")
            ROUTE_MEMORY_CACHE[key] = result
            return result

    try:
        payload = _fetch_json(
            ONEMAP_ROUTE_URL,
            {
                "start": f"{from_geocode.latitude},{from_geocode.longitude}",
                "end": f"{to_geocode.latitude},{to_geocode.longitude}",
                "routeType": "drive",
            },
            token=token,
        )
    except (HTTPError, URLError, TimeoutError, json.JSONDecodeError, OSError) as exc:
        return TravelCost(None, None, "fallback estimate", f"OneMap routing failed: {exc}")

    distance_km, duration_min = _parse_onemap_route(payload)
    if distance_km is None and duration_min is None:
        return TravelCost(None, None, "fallback estimate", "OneMap route returned no distance or duration")

    result = TravelCost(distance_km, duration_min, "OneMap")
    ROUTE_MEMORY_CACHE[key] = result
    cache = pd.concat(
        [
            cache,
            pd.DataFrame(
                [{"route_key": key, "distance_km": distance_km or 0, "duration_min": duration_min or 0}]
            ),
        ],
        ignore_index=True,
    ).drop_duplicates(subset=["route_key"], keep="last")
    _write_csv_cache(ROUTE_CACHE_FILE, cache)
    return result


def get_fallback_cost(
    from_address: Any,
    to_address: Any,
    from_zone: str | None = None,
    to_zone: str | None = None,
) -> TravelCost:
    from_text = clean_text(from_address)
    to_text = clean_text(to_address)
    if not from_text or not to_text:
        return TravelCost(UNKNOWN_DISTANCE_KM, UNKNOWN_DURATION_MIN, "fallback estimate", "Blank address")
    if from_text.casefold() == to_text.casefold():
        return TravelCost(0, 0, "fallback estimate")

    source_zone = from_zone or infer_zone(from_text)
    destination_zone = to_zone or infer_zone(to_text)
    if source_zone and destination_zone:
        distance = FALLBACK_ZONE_KM.get(source_zone, {}).get(destination_zone, UNKNOWN_DISTANCE_KM)
        duration = FALLBACK_ZONE_MINUTES.get(source_zone, {}).get(destination_zone, UNKNOWN_DURATION_MIN)
        return TravelCost(float(distance), float(duration), "fallback estimate")
    if source_zone == destination_zone and source_zone is not None:
        return TravelCost(SAME_UNKNOWN_DISTANCE_KM, SAME_UNKNOWN_DURATION_MIN, "fallback estimate")
    return TravelCost(UNKNOWN_DISTANCE_KM, UNKNOWN_DURATION_MIN, "fallback estimate", "Unknown zone")


def get_travel_cost(
    from_address: Any,
    to_address: Any,
    from_zone: str | None = None,
    to_zone: str | None = None,
    use_onemap: bool = True,
    token: str | None = None,
) -> TravelCost:
    if clean_text(from_address).casefold() == clean_text(to_address).casefold():
        return TravelCost(0, 0, "OneMap cache" if use_onemap else "fallback estimate")

    if use_onemap:
        from_geocode = get_cached_geocode(clean_text(from_address), token=token, use_onemap=True)
        to_geocode = get_cached_geocode(clean_text(to_address), token=token, use_onemap=True)
        onemap_cost = get_onemap_route_cost(from_geocode, to_geocode, token=token)
        if onemap_cost.distance_km is not None or onemap_cost.duration_min is not None:
            return onemap_cost
        fallback_cost = get_fallback_cost(from_address, to_address, from_zone=from_zone, to_zone=to_zone)
        return TravelCost(
            fallback_cost.distance_km,
            fallback_cost.duration_min,
            "fallback estimate",
            onemap_cost.error or fallback_cost.error,
        )

    return get_fallback_cost(from_address, to_address, from_zone=from_zone, to_zone=to_zone)


def get_cost_explanation() -> pd.DataFrame:
    rows = []
    for from_zone, destinations in FALLBACK_ZONE_MINUTES.items():
        for to_zone, duration in destinations.items():
            rows.append(
                {
                    "From Zone": from_zone,
                    "To Zone": to_zone,
                    "Fallback Distance KM": FALLBACK_ZONE_KM[from_zone][to_zone],
                    "Fallback Duration Min": duration,
                }
            )
    return pd.DataFrame(rows)


def calculate_route_zone_priority(
    rider_current_zone: str | None,
    pickup_zone: str | None,
    dropoff_zone: str | None,
) -> tuple[int, bool, bool]:
    same_zone_pickup = bool(rider_current_zone and pickup_zone and rider_current_zone == pickup_zone)
    same_zone_route = bool(pickup_zone and dropoff_zone and pickup_zone == dropoff_zone)
    stays_in_rider_zone = bool(
        rider_current_zone
        and pickup_zone
        and dropoff_zone
        and rider_current_zone == pickup_zone == dropoff_zone
    )
    if stays_in_rider_zone:
        return 0, same_zone_pickup, True
    if same_zone_route:
        return 1, same_zone_pickup, False
    return 2, same_zone_pickup, False


def calculate_zone_adjustment(
    rider_current_zone: str | None,
    pickup_zone: str | None,
    dropoff_zone: str | None,
) -> float:
    """
    Zone adjustment is strong but not absolute.

    Negative value = preferred. Positive value = penalty.
    Same-zone work is encouraged, but a clearly shorter/easier cross-zone
    movement can still win through the overall assignment score.
    """
    stays_in_rider_zone = bool(
        rider_current_zone
        and pickup_zone
        and dropoff_zone
        and rider_current_zone == pickup_zone == dropoff_zone
    )
    same_zone_pickup = bool(rider_current_zone and pickup_zone and rider_current_zone == pickup_zone)
    same_zone_route = bool(pickup_zone and dropoff_zone and pickup_zone == dropoff_zone)

    if stays_in_rider_zone:
        return -20.0
    if same_zone_pickup and not same_zone_route:
        return 20.0
    if not same_zone_pickup and same_zone_route:
        return 40.0
    if not same_zone_route:
        return 60.0
    return 0.0


def calculate_assignment_score(
    empty_distance_km: float | None,
    empty_duration_min: float | None,
    loaded_distance_km: float | None,
    loaded_duration_min: float | None,
    rider_current_zone: str | None,
    pickup_zone: str | None,
    dropoff_zone: str | None,
    rider_total_duration_min: float,
    rider_assigned_jobs: int,
    rider_max_jobs: int | None,
    optimise_by: str = "duration",
    empty_weight: float = DEFAULT_EMPTY_WEIGHT,
    loaded_weight: float = DEFAULT_LOADED_WEIGHT,
    soft_workload_min: float = DEFAULT_SOFT_WORKLOAD_MIN,
    workload_penalty_per_min: float = DEFAULT_WORKLOAD_PENALTY_PER_MIN,
    soft_adjusted_duration_min: float = DEFAULT_SOFT_ADJUSTED_DURATION_MIN,
    duration_buffer_multiplier: float = DEFAULT_DURATION_BUFFER_MULTIPLIER,
    duration_penalty_per_min: float = DEFAULT_DURATION_PENALTY_PER_MIN,
    max_jobs_overage_penalty: float = DEFAULT_MAX_JOB_OVERAGE_PENALTY,
    max_adjusted_duration_min: float = DEFAULT_MAX_ADJUSTED_DURATION_MIN,
) -> dict[str, float | int] | None:
    empty_distance = empty_distance_km if empty_distance_km is not None else UNKNOWN_DISTANCE_KM
    empty_duration = empty_duration_min if empty_duration_min is not None else UNKNOWN_DURATION_MIN
    loaded_distance = loaded_distance_km if loaded_distance_km is not None else UNKNOWN_DISTANCE_KM
    loaded_duration = loaded_duration_min if loaded_duration_min is not None else UNKNOWN_DURATION_MIN

    if optimise_by.casefold() == "distance":
        movement_score = (empty_distance * empty_weight) + (loaded_distance * loaded_weight)
    else:
        movement_score = (empty_duration * empty_weight) + (loaded_duration * loaded_weight)

    zone_adjustment = calculate_zone_adjustment(rider_current_zone, pickup_zone, dropoff_zone)
    candidate_total_duration = empty_duration + loaded_duration
    projected_duration = rider_total_duration_min + candidate_total_duration
    projected_adjusted_duration = projected_duration * duration_buffer_multiplier
    if projected_adjusted_duration > max_adjusted_duration_min:
        return None

    workload_penalty = (
        (projected_duration - soft_workload_min) * workload_penalty_per_min
        if projected_duration > soft_workload_min
        else 0.0
    )
    duration_penalty = (
        (projected_adjusted_duration - soft_adjusted_duration_min) * duration_penalty_per_min
        if projected_adjusted_duration > soft_adjusted_duration_min
        else 0.0
    )
    projected_jobs = rider_assigned_jobs + 1
    max_jobs_overage = max(0, projected_jobs - rider_max_jobs) if rider_max_jobs is not None else 0
    max_jobs_penalty = (max_jobs_overage**2) * max_jobs_overage_penalty
    assignment_score = (
        movement_score
        + zone_adjustment
        + workload_penalty
        + duration_penalty
        + max_jobs_penalty
    )

    return {
        "assignment_score": round(float(assignment_score), 3),
        "zone_adjustment": float(zone_adjustment),
        "workload_penalty": round(float(workload_penalty), 3),
        "duration_penalty": round(float(duration_penalty), 3),
        "max_jobs_penalty": round(float(max_jobs_penalty), 3),
        "projected_duration_min": round(float(projected_duration), 3),
        "projected_adjusted_duration_min": round(float(projected_adjusted_duration), 3),
        "max_jobs_overage": int(max_jobs_overage),
    }


def default_rider_table(count: int = 4) -> pd.DataFrame:
    defaults = [
        {"Rider Name": "Rider A", "Start Location": "Sengkang", "Start Zone": "North-East", "Max Jobs": 5},
        {"Rider Name": "Rider B", "Start Location": "Punggol", "Start Zone": "North-East", "Max Jobs": 5},
        {"Rider Name": "Rider C", "Start Location": "Yishun", "Start Zone": "North", "Max Jobs": 5},
        {"Rider Name": "Rider D", "Start Location": "Tampines", "Start Zone": "East", "Max Jobs": 5},
    ]
    while len(defaults) < count:
        rider_number = len(defaults) + 1
        defaults.append({"Rider Name": f"Rider {rider_number}", "Start Location": "", "Start Zone": "", "Max Jobs": None})
    return pd.DataFrame(defaults[:count])


def _normalise_rider_roster(rider_df: pd.DataFrame) -> pd.DataFrame:
    rider_df = rider_df.copy() if rider_df is not None else pd.DataFrame()
    for column in RIDER_COLUMNS:
        if column not in rider_df.columns:
            rider_df[column] = None
    rider_df = rider_df.loc[:, RIDER_COLUMNS].copy()
    rider_df["Rider Name"] = rider_df["Rider Name"].apply(clean_text)
    rider_df["Start Location"] = rider_df["Start Location"].apply(clean_text)
    rider_df["Start Zone"] = rider_df["Start Zone"].apply(clean_text)
    rider_df["Max Jobs"] = rider_df["Max Jobs"].apply(parse_optional_int)
    return rider_df


def ensure_rider_roster_workbook(path: Path = ROSTER_FILE) -> Path:
    if path.exists():
        try:
            workbook = pd.ExcelFile(path)
            try:
                existing_sheets = workbook.sheet_names
            finally:
                workbook.close()
        except Exception:
            existing_sheets = []
    else:
        existing_sheets = []

    if path.exists() and all(day in existing_sheets for day in WEEKDAY_SHEETS):
        return path

    sheets: dict[str, pd.DataFrame] = {}
    if path.exists() and existing_sheets:
        try:
            sheets = {
                sheet_name: _normalise_rider_roster(pd.read_excel(path, sheet_name=sheet_name))
                for sheet_name in existing_sheets
            }
        except Exception:
            sheets = {}

    for day in WEEKDAY_SHEETS:
        sheets.setdefault(day, default_rider_table(4))

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for day in WEEKDAY_SHEETS:
            _normalise_rider_roster(sheets[day]).to_excel(writer, sheet_name=day, index=False)
    return path


def load_rider_roster(day: str, path: Path = ROSTER_FILE) -> pd.DataFrame:
    if day not in WEEKDAY_SHEETS:
        raise ValueError(f"Unknown roster day: {day}")
    ensure_rider_roster_workbook(path)
    try:
        return _normalise_rider_roster(pd.read_excel(path, sheet_name=day))
    except ValueError:
        save_rider_roster(day, default_rider_table(4), path)
        return default_rider_table(4)


def save_rider_roster(day: str, rider_df: pd.DataFrame, path: Path = ROSTER_FILE) -> Path:
    if day not in WEEKDAY_SHEETS:
        raise ValueError(f"Unknown roster day: {day}")
    ensure_rider_roster_workbook(path)
    workbook = pd.ExcelFile(path)
    try:
        sheet_names = workbook.sheet_names
    finally:
        workbook.close()
    sheets = {
        sheet_name: _normalise_rider_roster(pd.read_excel(path, sheet_name=sheet_name))
        for sheet_name in sheet_names
        if sheet_name in WEEKDAY_SHEETS
    }
    sheets[day] = _normalise_rider_roster(rider_df)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in WEEKDAY_SHEETS:
            sheets.get(sheet_name, default_rider_table(4)).to_excel(writer, sheet_name=sheet_name, index=False)
    return path


def read_rider_roster_file(path: Path = ROSTER_FILE) -> bytes:
    ensure_rider_roster_workbook(path)
    return path.read_bytes()


def validate_riders(rider_df: pd.DataFrame) -> tuple[list[RiderState], list[str]]:
    errors: list[str] = []
    required_columns = ["Rider Name", "Start Location", "Start Zone", "Max Jobs"]
    for column in required_columns:
        if column not in rider_df.columns:
            errors.append(f"Missing rider column: {column}")

    if errors:
        return [], errors

    riders: list[RiderState] = []
    for index, row in rider_df.iterrows():
        rider_name = clean_text(row.get("Rider Name"))
        start_location = clean_text(row.get("Start Location"))
        if not rider_name and not start_location:
            continue
        if not rider_name or not start_location:
            errors.append(f"Rider row {index + 1} needs both Rider Name and Start Location.")
            continue
        riders.append(RiderState.from_row(row))

    if not riders:
        errors.append("Add at least one rider with a name and start location.")
    return riders, errors


def _combined_source(empty_cost: TravelCost, loaded_cost: TravelCost) -> str:
    if "public transport adjusted" in empty_cost.source.lower():
        base_empty_source = empty_cost.source.replace(", public transport adjusted", "")
        base_source = _combined_source(
            TravelCost(empty_cost.distance_km, empty_cost.duration_min, base_empty_source, empty_cost.error),
            loaded_cost,
        )
        return f"{base_source}, public transport adjusted"

    sources = {empty_cost.source, loaded_cost.source}
    if sources == {"OneMap"}:
        return "OneMap"
    if sources <= {"OneMap", "OneMap cache"}:
        return "OneMap cache" if "OneMap cache" in sources else "OneMap"
    if "fallback estimate" in sources:
        return "fallback estimate"
    return ", ".join(sorted(sources))


def adjust_empty_travel_for_public_transport(
    empty_cost: TravelCost,
    duration_multiplier: float = DEFAULT_EMPTY_TRAVEL_DURATION_MULTIPLIER,
    wait_buffer_min: float = DEFAULT_EMPTY_TRAVEL_WAIT_BUFFER_MIN,
) -> TravelCost:
    if empty_cost.duration_min is None:
        return empty_cost
    if empty_cost.duration_min <= 0:
        return empty_cost

    adjusted_duration = round((empty_cost.duration_min * duration_multiplier) + wait_buffer_min, 1)
    source = empty_cost.source
    if "public transport adjusted" not in source.lower():
        source = f"{source}, public transport adjusted"
    return TravelCost(empty_cost.distance_km, adjusted_duration, source, empty_cost.error)


def _job_uploaded_row(job: dict[str, Any]) -> int:
    return int(job.get("_original_order", 0)) + 2


def _route_status_for_adjusted_duration(adjusted_duration: float, base_cap: float) -> str:
    if adjusted_duration <= base_cap:
        return "OK"
    if adjusted_duration <= max(FORCE_COMPLETE_CAP_STAGES):
        return "Accepted under force-complete mode"
    return "Fail"


def _route_zone_for_job(job: dict[str, Any], key: str) -> str | None:
    address = clean_text(job.get(key))
    zone_key = "Pickup Zone" if key == "Pickup Address" else "Drop-off Zone"
    return job.get(zone_key) or infer_zone(address)


def optimise_vehicle_routes(
    jobs: pd.DataFrame,
    riders: list[RiderState],
    use_onemap: bool = True,
    optimise_by: str = "duration",
    token: str | None = None,
    progress_callback: ProgressCallback | None = None,
    empty_weight: float = DEFAULT_EMPTY_WEIGHT,
    loaded_weight: float = DEFAULT_LOADED_WEIGHT,
    soft_workload_min: float = DEFAULT_SOFT_WORKLOAD_MIN,
    workload_penalty_per_min: float = DEFAULT_WORKLOAD_PENALTY_PER_MIN,
    soft_adjusted_duration_min: float = DEFAULT_SOFT_ADJUSTED_DURATION_MIN,
    duration_penalty_per_min: float = DEFAULT_DURATION_PENALTY_PER_MIN,
    max_job_overage_penalty: float = DEFAULT_MAX_JOB_OVERAGE_PENALTY,
    duration_buffer_multiplier: float = DEFAULT_DURATION_BUFFER_MULTIPLIER,
    max_adjusted_duration_min: float = DEFAULT_MAX_ADJUSTED_DURATION_MIN,
    empty_travel_duration_multiplier: float = DEFAULT_EMPTY_TRAVEL_DURATION_MULTIPLIER,
    empty_travel_wait_buffer_min: float = DEFAULT_EMPTY_TRAVEL_WAIT_BUFFER_MIN,
    force_complete_assignment: bool = True,
    cluster_pressure_bonus_per_job: float = DEFAULT_CLUSTER_PRESSURE_BONUS_PER_JOB,
) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    token = token or get_onemap_token()
    remaining = jobs.sort_values("_original_order").to_dict("records")
    route_rows: list[dict[str, Any]] = []
    lookup_warnings: list[str] = []
    total_jobs = len(remaining)
    active_rider_count = max(1, len(riders))
    estimated_comparisons = max(1, active_rider_count * total_jobs * (total_jobs + 1) // 2)
    comparison_count = 0
    progress_base = 0.2 if use_onemap else 0
    progress_span = 0.78 if use_onemap else 0.98
    rider_sequences: dict[str, list[dict[str, Any]]] = {rider.name: [] for rider in riders}
    base_rider_state = {
        rider.name: {
            "start_location": rider.start_location,
            "start_zone": rider.start_zone,
            "max_jobs": rider.max_jobs,
        }
        for rider in riders
    }
    rejected_candidate_audit: list[dict[str, Any]] = []

    def report(status: str, **extra: Any) -> None:
        if progress_callback is None:
            return
        progress = progress_base + (progress_span * comparison_count / estimated_comparisons)
        payload = {
            "status": status,
            "assigned_jobs": len(route_rows),
            "total_jobs": total_jobs,
            "remaining_jobs": len(remaining),
            "comparison_count": comparison_count,
            "estimated_comparisons": estimated_comparisons,
            "progress": min(0.98, progress),
        }
        payload.update(extra)
        progress_callback(payload)

    def evaluate_rider_sequence(
        rider: RiderState,
        sequence: list[dict[str, Any]],
        cap: float,
    ) -> dict[str, Any]:
        current_location = base_rider_state[rider.name]["start_location"]
        current_zone = base_rider_state[rider.name]["start_zone"]
        empty_distance = 0.0
        empty_duration = 0.0
        loaded_distance = 0.0
        loaded_duration = 0.0
        rows: list[dict[str, Any]] = []
        warnings: list[str] = []

        for sequence_index, job in enumerate(sequence, start=1):
            pickup_address = clean_text(job["Pickup Address"])
            dropoff_address = clean_text(job["Drop-off Address"])
            pickup_zone = _route_zone_for_job(job, "Pickup Address")
            dropoff_zone = _route_zone_for_job(job, "Drop-off Address")

            empty_cost = get_travel_cost(
                current_location,
                pickup_address,
                current_zone,
                pickup_zone,
                use_onemap=use_onemap,
                token=token,
            )
            empty_cost = adjust_empty_travel_for_public_transport(
                empty_cost,
                duration_multiplier=empty_travel_duration_multiplier,
                wait_buffer_min=empty_travel_wait_buffer_min,
            )
            loaded_cost = get_travel_cost(
                pickup_address,
                dropoff_address,
                pickup_zone,
                dropoff_zone,
                use_onemap=use_onemap,
                token=token,
            )
            route_zone_priority, same_zone_pickup, route_stays_current_zone = calculate_route_zone_priority(
                current_zone,
                pickup_zone,
                dropoff_zone,
            )
            score_data = calculate_assignment_score(
                empty_distance_km=empty_cost.distance_km,
                empty_duration_min=empty_cost.duration_min,
                loaded_distance_km=loaded_cost.distance_km,
                loaded_duration_min=loaded_cost.duration_min,
                rider_current_zone=current_zone,
                pickup_zone=pickup_zone,
                dropoff_zone=dropoff_zone,
                rider_total_duration_min=empty_duration + loaded_duration,
                rider_assigned_jobs=sequence_index - 1,
                rider_max_jobs=base_rider_state[rider.name]["max_jobs"],
                optimise_by=optimise_by,
                empty_weight=empty_weight,
                loaded_weight=loaded_weight,
                soft_workload_min=soft_workload_min,
                workload_penalty_per_min=workload_penalty_per_min,
                soft_adjusted_duration_min=soft_adjusted_duration_min,
                duration_buffer_multiplier=duration_buffer_multiplier,
                duration_penalty_per_min=duration_penalty_per_min,
                max_jobs_overage_penalty=max_job_overage_penalty,
                max_adjusted_duration_min=cap,
            )
            if score_data is None:
                projected_raw = empty_duration + loaded_duration + (empty_cost.duration_min or 0) + (loaded_cost.duration_min or 0)
                return {
                    "valid": False,
                    "rows": rows,
                    "raw_duration": projected_raw,
                    "adjusted_duration": projected_raw * duration_buffer_multiplier,
                    "warnings": warnings,
                }

            if empty_cost.error:
                warnings.append(f"{current_location} -> {pickup_address}: {empty_cost.error}")
            if loaded_cost.error:
                warnings.append(f"{pickup_address} -> {dropoff_address}: {loaded_cost.error}")

            cost_source = _combined_source(empty_cost, loaded_cost)
            rows.append(
                {
                    "Rider": rider.name,
                    "Sequence": sequence_index,
                    "Uploaded Row": _job_uploaded_row(job),
                    "Start From": current_location,
                    "Empty Travel To Pickup": f"{current_location} -> {pickup_address}",
                    "Car Plate": clean_text(job.get("Car Plate")),
                    "Pickup Address": pickup_address,
                    "Pickup Lot": clean_text(job.get("Pickup Lot")),
                    "Drop-off Address": dropoff_address,
                    "Loaded Travel / Car Movement": f"{pickup_address} -> {dropoff_address}",
                    "Empty Distance KM": empty_cost.distance_km,
                    "Empty Duration Min": empty_cost.duration_min,
                    "Loaded Distance KM": loaded_cost.distance_km,
                    "Loaded Duration Min": loaded_cost.duration_min,
                    "Total Distance KM": round((empty_cost.distance_km or 0) + (loaded_cost.distance_km or 0), 2),
                    "Total Duration Min": round((empty_cost.duration_min or 0) + (loaded_cost.duration_min or 0), 1),
                    "Assignment Score": score_data["assignment_score"],
                    "Zone Adjustment": score_data["zone_adjustment"],
                    "Same Zone Pickup": "Yes" if same_zone_pickup else "No",
                    "Same Zone Route": "Yes" if route_stays_current_zone else "No",
                    "Route Zone Priority": route_zone_priority,
                    "Empty Weight": empty_weight,
                    "Loaded Weight": loaded_weight,
                    "Workload Penalty": score_data["workload_penalty"],
                    "Duration Penalty": score_data["duration_penalty"],
                    "Max Jobs Penalty": score_data["max_jobs_penalty"],
                    "Projected Rider Duration Min": round(float(score_data["projected_duration_min"]), 1),
                    "Projected Adjusted Duration Min": round(float(score_data["projected_adjusted_duration_min"]), 1),
                    "Cost Source": cost_source,
                    "Route Validation Status": "",
                }
            )

            empty_distance += empty_cost.distance_km or 0
            empty_duration += empty_cost.duration_min or 0
            loaded_distance += loaded_cost.distance_km or 0
            loaded_duration += loaded_cost.duration_min or 0
            current_location = dropoff_address
            current_zone = dropoff_zone

        raw_duration = empty_duration + loaded_duration
        return {
            "valid": raw_duration * duration_buffer_multiplier <= cap,
            "rows": rows,
            "empty_distance": empty_distance,
            "empty_duration": empty_duration,
            "loaded_distance": loaded_distance,
            "loaded_duration": loaded_duration,
            "raw_duration": raw_duration,
            "adjusted_duration": raw_duration * duration_buffer_multiplier,
            "final_location": current_location,
            "final_zone": current_zone,
            "warnings": warnings,
        }

    if use_onemap:
        unique_addresses = sorted(
            {
                clean_text(address)
                for address in (
                    list(jobs["Pickup Address"])
                    + list(jobs["Drop-off Address"])
                    + [rider.start_location for rider in riders]
                )
                if clean_text(address)
            }
        )
        for index, address in enumerate(unique_addresses, start=1):
            report(
                f"Geocoding address {index} of {len(unique_addresses)}",
                phase="Geocoding",
                current_address=address,
                progress=min(0.2, 0.2 * index / max(1, len(unique_addresses))),
            )
            result = get_cached_geocode(address, token=token, use_onemap=True)
            if result.error:
                lookup_warnings.append(f"{address}: {result.error}")
    else:
        report("Using fallback zone estimates only", phase="Fallback")

    # Semi-optimised greedy assignment: every round compares all remaining fixed
    # pickup-to-drop-off jobs against every rider's real current location.
    while remaining and riders:
        report("Comparing rider-job combinations", phase="Comparing")
        remaining_pickup_zone_counts: dict[str, int] = {}
        for remaining_job in remaining:
            remaining_zone = _route_zone_for_job(remaining_job, "Pickup Address") or "Unknown"
            remaining_pickup_zone_counts[remaining_zone] = remaining_pickup_zone_counts.get(remaining_zone, 0) + 1

        best_choice: tuple[
            tuple[float, float, float, int, float, int, int, int],
            int,
            RiderState,
            dict[str, Any],
            TravelCost,
            TravelCost,
            dict[str, float | int],
            bool,
            bool,
            int,
        ] | None = None

        for rider in riders:
            for job_index, job in enumerate(remaining):
                comparison_count += 1
                pickup_address = clean_text(job["Pickup Address"])
                dropoff_address = clean_text(job["Drop-off Address"])
                pickup_zone = job.get("Pickup Zone") or infer_zone(pickup_address)
                dropoff_zone = job.get("Drop-off Zone") or infer_zone(dropoff_address)

                empty_cost = get_travel_cost(
                    rider.current_location,
                    pickup_address,
                    rider.current_zone,
                    pickup_zone,
                    use_onemap=use_onemap,
                    token=token,
                )
                empty_cost = adjust_empty_travel_for_public_transport(
                    empty_cost,
                    duration_multiplier=empty_travel_duration_multiplier,
                    wait_buffer_min=empty_travel_wait_buffer_min,
                )
                loaded_cost = get_travel_cost(
                    pickup_address,
                    dropoff_address,
                    pickup_zone,
                    dropoff_zone,
                    use_onemap=use_onemap,
                    token=token,
                )
                # BlueSG-style relocation prioritises the rider's empty/public transport leg.
                # Zone fit, workload, and buffered duration are score components, not absolute first-sort gates.
                # The hard 3-hour adjusted duration cap still removes infeasible rider-job pairs.
                route_zone_priority, same_zone_pickup, route_stays_current_zone = calculate_route_zone_priority(
                    rider.current_zone,
                    pickup_zone,
                    dropoff_zone,
                )
                score_data = calculate_assignment_score(
                    empty_distance_km=empty_cost.distance_km,
                    empty_duration_min=empty_cost.duration_min,
                    loaded_distance_km=loaded_cost.distance_km,
                    loaded_duration_min=loaded_cost.duration_min,
                    rider_current_zone=rider.current_zone,
                    pickup_zone=pickup_zone,
                    dropoff_zone=dropoff_zone,
                    rider_total_duration_min=rider.empty_duration_min + rider.loaded_duration_min,
                    rider_assigned_jobs=rider.assigned_count,
                    rider_max_jobs=rider.max_jobs,
                    optimise_by=optimise_by,
                    empty_weight=empty_weight,
                    loaded_weight=loaded_weight,
                    soft_workload_min=soft_workload_min,
                    workload_penalty_per_min=workload_penalty_per_min,
                    soft_adjusted_duration_min=soft_adjusted_duration_min,
                    duration_buffer_multiplier=duration_buffer_multiplier,
                    duration_penalty_per_min=duration_penalty_per_min,
                    max_jobs_overage_penalty=max_job_overage_penalty,
                    max_adjusted_duration_min=max_adjusted_duration_min,
                )
                if score_data is None:
                    continue

                cluster_pressure_bonus = -cluster_pressure_bonus_per_job * remaining_pickup_zone_counts.get(pickup_zone or "Unknown", 0)
                greedy_assignment_score = float(score_data["assignment_score"]) + cluster_pressure_bonus
                candidate_rank = (
                    greedy_assignment_score,
                    empty_cost.distance_km if empty_cost.distance_km is not None else math.inf,
                    empty_cost.duration_min if empty_cost.duration_min is not None else math.inf,
                    route_zone_priority,
                    float(score_data["projected_adjusted_duration_min"]),
                    int(score_data["max_jobs_overage"]),
                    rider.assigned_count,
                    int(job["_original_order"]),
                )
                if best_choice is None or candidate_rank < best_choice[0]:
                    best_choice = (
                        candidate_rank,
                        job_index,
                        rider,
                        job,
                        empty_cost,
                        loaded_cost,
                        {**score_data, "assignment_score": round(greedy_assignment_score, 3)},
                        same_zone_pickup,
                        route_stays_current_zone,
                        route_zone_priority,
                    )

                if comparison_count == 1 or comparison_count % 10 == 0:
                    report(
                        "Comparing rider-job combinations",
                        phase="Routing",
                        current_rider=rider.name,
                        current_pickup=pickup_address,
                    )

        if best_choice is None:
            break

        (
            _,
            job_index,
            rider,
            job,
            empty_cost,
            loaded_cost,
            score_data,
            same_zone_pickup,
            route_stays_current_zone,
            route_zone_priority,
        ) = best_choice
        pickup_address = clean_text(job["Pickup Address"])
        dropoff_address = clean_text(job["Drop-off Address"])
        cost_source = _combined_source(empty_cost, loaded_cost)
        if empty_cost.error:
            lookup_warnings.append(f"{rider.current_location} -> {pickup_address}: {empty_cost.error}")
        if loaded_cost.error:
            lookup_warnings.append(f"{pickup_address} -> {dropoff_address}: {loaded_cost.error}")

        report(
            f"Assigned job {len(route_rows) + 1} of {total_jobs}",
            phase="Assigning",
            current_rider=rider.name,
            current_pickup=pickup_address,
            current_dropoff=dropoff_address,
        )

        route_rows.append(
            {
                "Rider": rider.name,
                "Sequence": rider.assigned_count + 1,
                "Uploaded Row": int(job.get("_original_order", len(route_rows))) + 2,
                "Start From": rider.current_location,
                "Empty Travel To Pickup": f"{rider.current_location} -> {pickup_address}",
                "Car Plate": clean_text(job.get("Car Plate")),
                "Pickup Address": pickup_address,
                "Pickup Lot": clean_text(job.get("Pickup Lot")),
                "Drop-off Address": dropoff_address,
                "Loaded Travel / Car Movement": f"{pickup_address} -> {dropoff_address}",
                "Empty Distance KM": empty_cost.distance_km,
                "Empty Duration Min": empty_cost.duration_min,
                "Loaded Distance KM": loaded_cost.distance_km,
                "Loaded Duration Min": loaded_cost.duration_min,
                "Total Distance KM": round((empty_cost.distance_km or 0) + (loaded_cost.distance_km or 0), 2),
                "Total Duration Min": round((empty_cost.duration_min or 0) + (loaded_cost.duration_min or 0), 1),
                "Assignment Score": score_data["assignment_score"],
                "Zone Adjustment": score_data["zone_adjustment"],
                "Same Zone Pickup": "Yes" if same_zone_pickup else "No",
                "Same Zone Route": "Yes" if route_stays_current_zone else "No",
                "Route Zone Priority": route_zone_priority,
                "Empty Weight": empty_weight,
                "Loaded Weight": loaded_weight,
                "Workload Penalty": score_data["workload_penalty"],
                "Duration Penalty": score_data["duration_penalty"],
                "Max Jobs Penalty": score_data["max_jobs_penalty"],
                "Projected Rider Duration Min": round(float(score_data["projected_duration_min"]), 1),
                "Projected Adjusted Duration Min": round(float(score_data["projected_adjusted_duration_min"]), 1),
                "Cost Source": cost_source,
                "Route Validation Status": "",
            }
        )

        rider.assigned_count += 1
        rider.empty_distance_km += empty_cost.distance_km or 0
        rider.empty_duration_min += empty_cost.duration_min or 0
        rider.loaded_distance_km += loaded_cost.distance_km or 0
        rider.loaded_duration_min += loaded_cost.duration_min or 0
        rider.current_location = dropoff_address
        rider.current_zone = job.get("Drop-off Zone") or infer_zone(dropoff_address)
        rider_sequences[rider.name].append(job)
        remaining.pop(job_index)

    cap_stages = [max_adjusted_duration_min]
    if force_complete_assignment:
        cap_stages = []
        for cap in [max_adjusted_duration_min] + FORCE_COMPLETE_CAP_STAGES:
            if cap not in cap_stages:
                cap_stages.append(float(cap))

    cap_used = max_adjusted_duration_min
    if remaining:
        for cap in cap_stages:
            cap_used = cap
            made_assignment = True
            while remaining and made_assignment:
                made_assignment = False
                best_insertion: tuple[float, float, int, RiderState, int, int, dict[str, Any], dict[str, Any]] | None = None

                current_evaluations = {
                    rider.name: evaluate_rider_sequence(rider, rider_sequences[rider.name], cap)
                    for rider in riders
                }
                for job_index, job in enumerate(remaining):
                    for rider in riders:
                        current_sequence = rider_sequences[rider.name]
                        current_duration = float(current_evaluations[rider.name].get("raw_duration", 0) or 0)
                        for insert_at in range(len(current_sequence) + 1):
                            candidate_sequence = current_sequence[:insert_at] + [job] + current_sequence[insert_at:]
                            evaluation = evaluate_rider_sequence(rider, candidate_sequence, cap)
                            added_duration = float(evaluation.get("raw_duration", math.inf)) - current_duration
                            projected_adjusted = float(evaluation.get("adjusted_duration", math.inf))
                            if not evaluation.get("valid"):
                                continue
                            candidate_rank = (
                                added_duration,
                                projected_adjusted,
                                int(job.get("_original_order", job_index)),
                            )
                            if best_insertion is None or candidate_rank < best_insertion[:3]:
                                best_insertion = (
                                    added_duration,
                                    projected_adjusted,
                                    int(job.get("_original_order", job_index)),
                                    rider,
                                    insert_at,
                                    job_index,
                                    job,
                                    evaluation,
                                )

                if best_insertion is not None:
                    _, _, _, rider, insert_at, job_index, job, _ = best_insertion
                    rider_sequences[rider.name].insert(insert_at, job)
                    remaining.pop(job_index)
                    made_assignment = True

            if not remaining:
                break

    final_evaluations = {
        rider.name: evaluate_rider_sequence(rider, rider_sequences[rider.name], cap_used)
        for rider in riders
    }
    route_rows = []
    lookup_warnings.extend(
        warning
        for evaluation in final_evaluations.values()
        for warning in evaluation.get("warnings", [])
    )
    for rider in riders:
        evaluation = final_evaluations[rider.name]
        route_rows.extend(evaluation.get("rows", []))
        rider.assigned_count = len(rider_sequences[rider.name])
        rider.empty_distance_km = float(evaluation.get("empty_distance", 0) or 0)
        rider.empty_duration_min = float(evaluation.get("empty_duration", 0) or 0)
        rider.loaded_distance_km = float(evaluation.get("loaded_distance", 0) or 0)
        rider.loaded_duration_min = float(evaluation.get("loaded_duration", 0) or 0)
        rider.current_location = clean_text(evaluation.get("final_location", base_rider_state[rider.name]["start_location"]))
        rider.current_zone = evaluation.get("final_zone") or base_rider_state[rider.name]["start_zone"]

    unassigned_details: list[dict[str, Any]] = []
    for job in remaining:
        candidate_rows = []
        cap_for_audit = cap_used
        for rider in riders:
            current_sequence = rider_sequences[rider.name]
            current_evaluation = evaluate_rider_sequence(rider, current_sequence, cap_for_audit)
            current_raw_duration = float(current_evaluation.get("raw_duration", 0) or 0)
            current_location = (
                clean_text(current_sequence[-1]["Drop-off Address"])
                if current_sequence
                else base_rider_state[rider.name]["start_location"]
            )
            for insert_at in range(len(current_sequence) + 1):
                candidate_sequence = current_sequence[:insert_at] + [job] + current_sequence[insert_at:]
                evaluation = evaluate_rider_sequence(rider, candidate_sequence, cap_for_audit)
                inserted_rows = [
                    row for row in evaluation.get("rows", [])
                    if int(row.get("Uploaded Row", -1)) == _job_uploaded_row(job)
                ]
                inserted_row = inserted_rows[0] if inserted_rows else {}
                candidate_rows.append(
                    {
                        "Car Plate": clean_text(job.get("Car Plate")),
                        "Pickup Address": clean_text(job.get("Pickup Address")),
                        "Drop-off Address": clean_text(job.get("Drop-off Address")),
                        "Candidate Rider": rider.name,
                        "Candidate Start/Current Location": current_location,
                        "Candidate Empty Duration Min": inserted_row.get("Empty Duration Min", ""),
                        "Candidate Loaded Duration Min": inserted_row.get("Loaded Duration Min", ""),
                        "Candidate Added Duration Min": round(
                            float(evaluation.get("raw_duration", 0) or 0) - current_raw_duration,
                            1,
                        ),
                        "Projected Raw Duration Min": round(float(evaluation.get("raw_duration", 0) or 0), 1),
                        "Projected Adjusted Duration Min": round(float(evaluation.get("adjusted_duration", 0) or 0), 1),
                        "Cap Used": cap_for_audit,
                        "Reason Rejected": (
                            "Projected adjusted duration exceeds cap"
                            if float(evaluation.get("adjusted_duration", math.inf)) > cap_for_audit
                            else "Not selected by rescue insertion ranking"
                        ),
                    }
                )

        candidate_rows = sorted(
            candidate_rows,
            key=lambda row: float(row["Projected Adjusted Duration Min"] or math.inf),
        )
        rejected_candidate_audit.extend(candidate_rows[:3])
        best = candidate_rows[0] if candidate_rows else {}
        best_adjusted = best.get("Projected Adjusted Duration Min", "")
        best_rider = best.get("Candidate Rider", "")
        best_location = best.get("Candidate Start/Current Location", "")
        unassigned_details.append(
            {
                "Uploaded Row": _job_uploaded_row(job),
                "Best Candidate Rider": best_rider,
                "Best Candidate Final Location": best_location,
                "Best Candidate Projected Raw Duration": best.get("Projected Raw Duration Min", ""),
                "Best Candidate Projected Adjusted Duration": best_adjusted,
                "Current Cap": cap_for_audit,
                "Reason": (
                    f"Could not assign under current cap. Best candidate was {best_rider}, "
                    f"projected adjusted duration {best_adjusted} min, cap {cap_for_audit} min."
                    if best
                    else "Could not assign under current cap. No rider candidate was available."
                ),
            }
        )

    route_df = format_route_output(pd.DataFrame(route_rows, columns=ROUTE_COLUMNS), riders)
    route_df.attrs["unassigned_details"] = unassigned_details
    route_df.attrs["rejected_candidate_audit"] = rejected_candidate_audit
    route_df.attrs["force_complete_cap_used"] = cap_used
    summary_df = pd.DataFrame(
        [
            {
                "Rider": rider.name,
                "Total Jobs": rider.assigned_count,
                "Total Empty Distance KM": round(rider.empty_distance_km, 2),
                "Total Empty Duration Min": round(rider.empty_duration_min, 1),
                "Total Loaded Distance KM": round(rider.loaded_distance_km, 2),
                "Total Loaded Duration Min": round(rider.loaded_duration_min, 1),
                "Total Route Distance KM": round(rider.empty_distance_km + rider.loaded_distance_km, 2),
                "Total Route Duration Min": round(rider.empty_duration_min + rider.loaded_duration_min, 1),
                "Adjusted Route Duration Min": round(
                    (rider.empty_duration_min + rider.loaded_duration_min) * duration_buffer_multiplier,
                    1,
                ),
                "Within 3 Hours": _route_status_for_adjusted_duration(
                    (rider.empty_duration_min + rider.loaded_duration_min) * duration_buffer_multiplier,
                    DEFAULT_MAX_ADJUSTED_DURATION_MIN,
                ),
                "Final Location": rider.current_location,
            }
            for rider in riders
        ],
        columns=SUMMARY_COLUMNS,
    )
    summary_df = format_summary_output(summary_df, route_df)
    report("Finished route optimisation", phase="Finished", progress=1.0)
    return route_df, summary_df, sorted(set(lookup_warnings))


def validate_route_chain(route_df: pd.DataFrame, riders: list[RiderState]) -> dict[tuple[str, int], str]:
    validation: dict[tuple[str, int], str] = {}
    start_locations = {rider.name: rider.start_location for rider in riders}

    for rider_name, rider_routes in route_df.groupby("Rider", sort=False):
        expected_start = start_locations.get(rider_name, "")
        for _, row in rider_routes.sort_values("Sequence").iterrows():
            sequence = int(row["Sequence"])
            actual_start = clean_text(row["Start From"])
            if actual_start.casefold() == clean_text(expected_start).casefold():
                validation[(rider_name, sequence)] = "OK"
            else:
                validation[(rider_name, sequence)] = f"Check: expected start from {expected_start}"
            expected_start = clean_text(row["Drop-off Address"])

    return validation


def format_route_output(route_df: pd.DataFrame, riders: list[RiderState]) -> pd.DataFrame:
    if route_df.empty:
        return pd.DataFrame(columns=ROUTE_COLUMNS)
    validation = validate_route_chain(route_df, riders)
    route_df = route_df.copy()
    route_df["Route Validation Status"] = route_df.apply(
        lambda row: validation.get((row["Rider"], int(row["Sequence"])), "Check route chain"),
        axis=1,
    )
    return route_df.loc[:, ROUTE_COLUMNS]


def format_summary_output(summary_df: pd.DataFrame, route_df: pd.DataFrame) -> pd.DataFrame:
    if summary_df.empty:
        return pd.DataFrame(columns=SUMMARY_COLUMNS)

    summary_df = summary_df.copy()
    if "Adjusted Route Duration Min" not in summary_df.columns or summary_df["Adjusted Route Duration Min"].isna().all():
        summary_df["Adjusted Route Duration Min"] = (
            summary_df["Total Route Duration Min"].fillna(0).astype(float) * DEFAULT_DURATION_BUFFER_MULTIPLIER
        ).round(1)
    summary_df["Total Route Duration Hours"] = (
        summary_df["Total Route Duration Min"].fillna(0).astype(float) / 60
    ).round(2)
    if "Within 3 Hours" not in summary_df.columns or summary_df["Within 3 Hours"].isna().all():
        summary_df["Within 3 Hours"] = summary_df["Adjusted Route Duration Min"].apply(
            lambda value: "OK" if float(value or 0) <= DEFAULT_MAX_ADJUSTED_DURATION_MIN else "Fail"
        )

    total_duration = summary_df["Total Route Duration Min"].fillna(0).astype(float)
    summary_df["Empty Travel %"] = summary_df.apply(
        lambda row: round(
            (float(row["Total Empty Duration Min"] or 0) / float(row["Total Route Duration Min"])) * 100,
            1,
        )
        if float(row["Total Route Duration Min"] or 0) > 0
        else 0,
        axis=1,
    )
    summary_df["Loaded Travel %"] = summary_df.apply(
        lambda row: round(
            (float(row["Total Loaded Duration Min"] or 0) / float(row["Total Route Duration Min"])) * 100,
            1,
        )
        if float(row["Total Route Duration Min"] or 0) > 0
        else 0,
        axis=1,
    )

    active_durations = total_duration[summary_df["Total Jobs"].fillna(0).astype(int) > 0]
    average_duration = float(active_durations.mean()) if not active_durations.empty else 0

    fallback_share_by_rider: dict[str, float] = {}
    if not route_df.empty:
        for rider, rider_routes in route_df.groupby("Rider"):
            fallback_rows = rider_routes["Cost Source"].astype(str).str.contains("fallback", case=False, na=False)
            fallback_share_by_rider[str(rider)] = float(fallback_rows.mean()) if len(fallback_rows) else 0

    comments = []
    for _, row in summary_df.iterrows():
        rider = clean_text(row["Rider"])
        duration = float(row["Total Route Duration Min"] or 0)
        adjusted_duration = float(row["Adjusted Route Duration Min"] or 0)
        jobs = int(row["Total Jobs"] or 0)
        fallback_share = fallback_share_by_rider.get(rider, 0)

        if jobs == 0:
            comment = "No jobs assigned"
        elif adjusted_duration > DEFAULT_MAX_ADJUSTED_DURATION_MIN:
            comment = "Fails 3-hour adjusted cap"
        elif fallback_share >= 0.5:
            comment = "Estimate only - verify if needed"
        elif average_duration > 0 and duration > average_duration * 1.2:
            comment = "Heavier route"
        elif average_duration > 0 and duration < average_duration * 0.8:
            comment = "Lighter route"
        else:
            comment = "Balanced"
        comments.append(comment)

    summary_df["Workload Comment"] = comments
    return summary_df.loc[:, SUMMARY_COLUMNS]


def _write_title(ws: Any, title: str, row: int = 1) -> None:
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1F2937")


def _write_rows(ws: Any, rows: list[list[Any]], start_row: int = 1) -> int:
    for row_offset, values in enumerate(rows):
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=start_row + row_offset, column=column, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    return start_row + len(rows)


def _style_header_row(ws: Any, row: int, max_column: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for column in range(1, max_column + 1):
        cell = ws.cell(row=row, column=column)
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def _autosize_columns(ws: Any, max_width: int = 45) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 10
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, min(len(str(cell.value)), max_width))
        ws.column_dimensions[column_letter].width = max_length + 2


def _style_route_sheet(ws: Any, header_row: int, table_start_row: int, table_end_row: int) -> None:
    _style_header_row(ws, header_row, len(ROUTE_COLUMNS))
    ws.freeze_panes = f"A{header_row + 1}"
    address_columns = {
        "Start From",
        "Empty Travel To Pickup",
        "Pickup Address",
        "Drop-off Address",
        "Loaded Travel / Car Movement",
    }
    header_map = {ws.cell(header_row, col).value: col for col in range(1, len(ROUTE_COLUMNS) + 1)}
    previous_rider = None
    rider_fill_a = PatternFill("solid", fgColor="F8FAFC")
    rider_fill_b = PatternFill("solid", fgColor="EEF6FF")
    fallback_fill = PatternFill("solid", fgColor="FFE7A3")
    onemap_fill = PatternFill("solid", fgColor="C6EFCE")
    ok_fill = PatternFill("solid", fgColor="C6EFCE")
    warning_fill = PatternFill("solid", fgColor="F4B183")
    rider_band = 0

    for row in range(table_start_row, table_end_row + 1):
        rider = ws.cell(row, header_map["Rider"]).value
        if rider != previous_rider:
            rider_band += 1
            previous_rider = rider
        row_fill = rider_fill_a if rider_band % 2 else rider_fill_b
        for col in range(1, len(ROUTE_COLUMNS) + 1):
            cell = ws.cell(row, col)
            cell.fill = row_fill
            cell.alignment = Alignment(
                wrap_text=ws.cell(header_row, col).value in address_columns,
                vertical="top",
            )

        cost_cell = ws.cell(row, header_map["Cost Source"])
        if "fallback" in str(cost_cell.value).lower():
            cost_cell.fill = fallback_fill
        elif "onemap" in str(cost_cell.value).lower():
            cost_cell.fill = onemap_fill

        validation_cell = ws.cell(row, header_map["Route Validation Status"])
        validation_cell.fill = ok_fill if validation_cell.value == "OK" else warning_fill


def _style_summary_sheet(ws: Any, rider_summary_header_row: int) -> None:
    _style_header_row(ws, 3, 2)
    _style_header_row(ws, rider_summary_header_row, len(SUMMARY_COLUMNS))
    ws.freeze_panes = f"A{rider_summary_header_row + 1}"
    for row in range(rider_summary_header_row + 1, ws.max_row + 1):
        comment = str(ws.cell(row, SUMMARY_COLUMNS.index("Workload Comment") + 1).value or "")
        if "Heavier" in comment:
            fill = PatternFill("solid", fgColor="F4B183")
        elif "Lighter" in comment:
            fill = PatternFill("solid", fgColor="D9EAD3")
        elif "Estimate only" in comment:
            fill = PatternFill("solid", fgColor="FFE7A3")
        else:
            fill = PatternFill("solid", fgColor="E2F0D9")
        ws.cell(row, SUMMARY_COLUMNS.index("Workload Comment") + 1).fill = fill


def _overall_summary(route_df: pd.DataFrame, summary_df: pd.DataFrame, jobs_df: pd.DataFrame | None) -> list[list[Any]]:
    active_summary = summary_df[summary_df["Total Jobs"].fillna(0).astype(int) > 0]
    total_duration = float(summary_df["Total Route Duration Min"].fillna(0).sum()) if not summary_df.empty else 0
    average_duration = float(active_summary["Total Route Duration Min"].mean()) if not active_summary.empty else 0
    longest = ""
    shortest = ""
    if not active_summary.empty:
        longest_row = active_summary.loc[active_summary["Total Route Duration Min"].idxmax()]
        shortest_row = active_summary.loc[active_summary["Total Route Duration Min"].idxmin()]
        longest = f"{longest_row['Rider']} ({longest_row['Total Route Duration Min']} min)"
        shortest = f"{shortest_row['Rider']} ({shortest_row['Total Route Duration Min']} min)"

    cost_source = route_df["Cost Source"].astype(str) if not route_df.empty else pd.Series(dtype=str)
    onemap_count = int(cost_source.str.contains("OneMap", case=False, na=False).sum())
    fallback_count = int(cost_source.str.contains("fallback", case=False, na=False).sum())
    uploaded_jobs = (
        int(jobs_df.attrs.get("uploaded_count", len(jobs_df)))
        if jobs_df is not None
        else "Not provided"
    )
    fallback_warning = (
        "Warning: most route costs are fallback estimates."
        if fallback_count > onemap_count and fallback_count > 0
        else "Cost source mix looks acceptable."
    )

    return [
        ["Metric", "Value"],
        ["Total riders used", int((summary_df["Total Jobs"].fillna(0).astype(int) > 0).sum())],
        ["Total jobs uploaded", uploaded_jobs],
        ["Total jobs assigned", len(route_df)],
        ["Total estimated route duration", f"{round(total_duration, 1)} min / {round(total_duration / 60, 2)} hours"],
        ["Average duration per rider", f"{round(average_duration, 1)} min"],
        ["Longest rider route", longest],
        ["Shortest rider route", shortest],
        ["Number of OneMap route calculations", onemap_count],
        ["Number of fallback estimates used", fallback_count],
        ["Fallback warning", fallback_warning],
    ]


def _data_quality_rows(
    route_df: pd.DataFrame,
    jobs_df: pd.DataFrame | None,
    validation_warnings: list[str] | None,
    lookup_warnings: list[str] | None,
) -> list[list[Any]]:
    uploaded_jobs = len(jobs_df) if jobs_df is not None else "Not provided"
    blank_address_count = 0
    duplicate_plates = []
    if jobs_df is not None and not jobs_df.empty:
        uploaded_jobs = int(jobs_df.attrs.get("uploaded_count", len(jobs_df)))
        blank_address_count = int(jobs_df.attrs.get("blank_address_rows_dropped", 0))
        duplicate_plates = list(jobs_df.attrs.get("duplicate_plate_values", []))
        if not duplicate_plates:
            plates = jobs_df["Car Plate"].apply(clean_text)
            duplicate_plates = sorted(plates[plates.ne("") & plates.duplicated(keep=False)].unique())

    fallback_routes = route_df[route_df["Cost Source"].astype(str).str.contains("fallback", case=False, na=False)]
    validation_failures = route_df[route_df["Route Validation Status"].ne("OK")]
    geocode_failures = [
        warning
        for warning in (lookup_warnings or [])
        if "geocod" in warning.lower() or "no onemap" in warning.lower()
    ]

    action = "No major issues found. Routes can be reviewed and sent to riders."
    if not validation_failures.empty:
        action = "Do not send this route yet. Fix route chaining issue first."
    elif not fallback_routes.empty:
        action = "Some routes used fallback estimates. Review these manually if accuracy is important."

    rows = [
        ["Check", "Result"],
        ["Number of jobs uploaded", uploaded_jobs],
        ["Number of jobs assigned", len(route_df)],
        ["Number of unassigned jobs", max(0, int(uploaded_jobs) - len(route_df)) if isinstance(uploaded_jobs, int) else "Not provided"],
        ["Duplicate car plates detected", ", ".join(duplicate_plates) if duplicate_plates else "None detected"],
        ["Blank pickup/drop-off addresses detected", blank_address_count],
        ["Addresses that failed OneMap geocoding", "\n".join(geocode_failures) if geocode_failures else "None detected"],
        ["Routes that used fallback estimates", len(fallback_routes)],
        ["Route validation failures", len(validation_failures)],
        ["Validation/input warnings", "\n".join(validation_warnings or []) if validation_warnings else "None"],
        ["User Action Required", action],
    ]
    return rows


def build_unassigned_jobs_df(jobs_df: pd.DataFrame | None, route_df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Uploaded Row",
        "Car Plate",
        "Pickup Address",
        "Pickup Lot",
        "Drop-off Address",
        "Pickup Zone",
        "Drop-off Zone",
        "Best Candidate Rider",
        "Best Candidate Final Location",
        "Best Candidate Projected Raw Duration",
        "Best Candidate Projected Adjusted Duration",
        "Current Cap",
        "Reason",
    ]
    if jobs_df is None or jobs_df.empty:
        return pd.DataFrame(columns=columns)

    jobs = jobs_df.copy()
    if "_original_order" not in jobs.columns:
        jobs["_original_order"] = range(len(jobs))
    jobs["Uploaded Row"] = jobs["_original_order"].astype(int) + 2

    assigned_rows: set[int] = set()
    if route_df is not None and not route_df.empty and "Uploaded Row" in route_df.columns:
        assigned_rows = set(route_df["Uploaded Row"].dropna().astype(int).tolist())

    unassigned = jobs[~jobs["Uploaded Row"].isin(assigned_rows)].copy()
    if unassigned.empty:
        return pd.DataFrame(columns=columns)

    details = {}
    if route_df is not None:
        details = {
            int(detail.get("Uploaded Row")): detail
            for detail in route_df.attrs.get("unassigned_details", [])
            if detail.get("Uploaded Row") is not None
        }

    for detail_column in [
        "Best Candidate Rider",
        "Best Candidate Final Location",
        "Best Candidate Projected Raw Duration",
        "Best Candidate Projected Adjusted Duration",
        "Current Cap",
        "Reason",
    ]:
        unassigned[detail_column] = unassigned["Uploaded Row"].apply(
            lambda row_number: details.get(int(row_number), {}).get(detail_column, "")
        )
    unassigned["Reason"] = unassigned["Reason"].replace(
        "",
        "Could not assign under current cap. No candidate audit was available.",
    )
    for column in columns:
        if column not in unassigned.columns:
            unassigned[column] = ""
    return unassigned.loc[:, columns].reset_index(drop=True)


def _write_how_to_read_sheet(writer: pd.ExcelWriter) -> None:
    wb = writer.book
    ws = wb.create_sheet("How To Read This", 0)
    _write_title(ws, "How To Read This Export", 1)
    rows = [
        ["What this file is for", "This file assigns vehicle relocation jobs to riders and sequences them."],
        ["How each route works", "Each row has two legs: Empty Travel To Pickup and Loaded Travel / Car Movement."],
        ["Empty Travel To Pickup", "The rider travels without the car from the current location to the pickup address."],
        ["Loaded Travel / Car Movement", "The rider drives the assigned car from pickup address to drop-off address."],
        ["Important route logic", "After each drop-off, the rider's new starting point becomes the previous drop-off address. The rider does not reset to the original start location."],
        ["Example", "If Rider D drops a car at 150 Tampines Street 12, the next row for Rider D should start from 150 Tampines Street 12."],
        ["Rider", "Assigned rider or driver."],
        ["Sequence", "Order that rider should follow."],
        ["Start From", "Where the rider starts this job from."],
        ["Pickup Address / Pickup Lot", "Where the car is collected, including parking lot or deck information."],
        ["Drop-off Address", "Where the car must be moved to."],
        ["Empty / Loaded Distance and Duration", "Empty is before collecting the car. Loaded is while moving the assigned car. Total is empty plus loaded."],
        ["Cost Source", "OneMap means distance/time came from OneMap. fallback estimate means rough zone-based estimates were used because OneMap was unavailable."],
        ["Route Validation Status", "OK means sequence 1 starts from the rider start location, and later rows start from the previous drop-off address."],
        ["Fallback estimates", "Treat fallback estimates as planning estimates, not exact travel times."],
    ]
    _write_rows(ws, rows, 3)
    _style_header_row(ws, 3, 2)
    _autosize_columns(ws, 70)


def _rider_cost_accuracy_note(rider_routes: pd.DataFrame) -> str:
    if rider_routes.empty:
        return "No assigned route."
    cost_sources = rider_routes["Cost Source"].astype(str)
    fallback_count = int(cost_sources.str.contains("fallback", case=False, na=False).sum())
    if fallback_count == 0:
        return "Based on OneMap distance/time estimates."
    if fallback_count == len(rider_routes):
        return "Timing is rough estimate only."
    return "Some legs use fallback estimates. Timing should be manually reviewed."


def _route_validation_summary(rider_routes: pd.DataFrame) -> str:
    if rider_routes.empty:
        return "No route"
    return "OK" if rider_routes["Route Validation Status"].eq("OK").all() else "Check route chain"


def _cost_source_summary(rider_routes: pd.DataFrame) -> str:
    if rider_routes.empty:
        return "No route"
    cost_sources = rider_routes["Cost Source"].astype(str)
    fallback_count = int(cost_sources.str.contains("fallback", case=False, na=False).sum())
    onemap_count = int(cost_sources.str.contains("OneMap", case=False, na=False).sum())
    if fallback_count == 0:
        return "All OneMap"
    if onemap_count == 0:
        return "All fallback estimates"
    return f"Mixed: {onemap_count} OneMap, {fallback_count} fallback"


def _manager_note(rider_routes: pd.DataFrame) -> str:
    if _route_validation_summary(rider_routes) != "OK":
        return "Do not dispatch until fixed"
    if rider_routes["Cost Source"].astype(str).str.contains("fallback", case=False, na=False).any():
        return "Review timing manually"
    return "Ready to dispatch"


def _write_rider_instructions_sheet(writer: pd.ExcelWriter, route_df: pd.DataFrame, summary_df: pd.DataFrame) -> None:
    ws = writer.book.create_sheet("Rider Instructions")
    _write_title(ws, "Rider Instructions", 1)
    ws.cell(
        row=2,
        column=1,
        value="Simple dispatch instructions generated from the Optimised Routes sheet. No re-optimisation is done here.",
    )
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    headers = [
        "Step",
        "Start From",
        "Go To Pickup",
        "Collect Car Plate",
        "Pickup Lot",
        "Drop-off At",
        "Simple Instruction",
        "Estimated Step Duration Min",
        "Estimated Step Distance KM",
    ]
    section_fill = PatternFill("solid", fgColor="1F4E78")
    current_row = 4
    summary_by_rider = summary_df.set_index("Rider") if not summary_df.empty else pd.DataFrame()

    if route_df.empty:
        ws.cell(row=current_row, column=1, value="No rider routes were assigned.")
        _autosize_columns(ws, 60)
        return

    for rider, rider_routes in route_df.sort_values(["Rider", "Sequence"]).groupby("Rider", sort=False):
        rider_summary = summary_by_rider.loc[rider] if rider in summary_by_rider.index else None
        start_location = clean_text(rider_routes.iloc[0]["Start From"])
        final_location = clean_text(rider_routes.iloc[-1]["Drop-off Address"])
        total_distance = float(rider_routes["Total Distance KM"].fillna(0).sum())
        total_duration = float(rider_routes["Total Duration Min"].fillna(0).sum())
        validation_status = _route_validation_summary(rider_routes)
        cost_note = _rider_cost_accuracy_note(rider_routes)
        if rider_summary is not None:
            start_location = clean_text(rider_routes.iloc[0]["Start From"])
            final_location = clean_text(rider_summary.get("Final Location", final_location))
            total_distance = float(rider_summary.get("Total Route Distance KM", total_distance) or 0)
            total_duration = float(rider_summary.get("Total Route Duration Min", total_duration) or 0)

        ws.cell(row=current_row, column=1, value=f"Rider: {rider}")
        ws.cell(row=current_row, column=1).fill = section_fill
        ws.cell(row=current_row, column=1).font = Font(bold=True, color="FFFFFF", size=12)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        current_row += 1

        header_rows = [
            ["Rider Name", rider],
            ["Total Cars Assigned", len(rider_routes)],
            ["Start Location", start_location],
            ["Final Location", final_location],
            ["Estimated Total Distance KM", round(total_distance, 2)],
            ["Estimated Total Duration Min", round(total_duration, 1)],
            ["Route Validation Status", validation_status],
            ["Cost Accuracy Note", cost_note],
        ]
        current_row = _write_rows(ws, header_rows, current_row)
        current_row += 1

        for col, header in enumerate(headers, start=1):
            ws.cell(row=current_row, column=col, value=header)
        _style_header_row(ws, current_row, len(headers))
        current_row += 1

        for _, route in rider_routes.sort_values("Sequence").iterrows():
            pickup = clean_text(route["Pickup Address"])
            dropoff = clean_text(route["Drop-off Address"])
            start_from = clean_text(route["Start From"])
            car_plate = clean_text(route["Car Plate"])
            pickup_lot = clean_text(route["Pickup Lot"]) or "the stated pickup lot"
            if int(route["Sequence"]) == 1:
                instruction = (
                    f"Start from {start_from}. Go to {pickup}. Collect car {car_plate} "
                    f"at {pickup_lot}. Drive it to {dropoff}."
                )
            else:
                instruction = (
                    f"From {start_from}, go to {pickup}. Collect car {car_plate} "
                    f"at {pickup_lot}. Drive it to {dropoff}."
                )
            row_values = [
                int(route["Sequence"]),
                start_from,
                pickup,
                car_plate,
                clean_text(route["Pickup Lot"]),
                dropoff,
                instruction,
                route["Total Duration Min"],
                route["Total Distance KM"],
            ]
            for col, value in enumerate(row_values, start=1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.alignment = Alignment(wrap_text=col in {2, 3, 6, 7}, vertical="top")
            current_row += 1

        current_row += 1

    ws.column_dimensions["G"].width = 90
    _autosize_columns(ws, 45)
    ws.column_dimensions["G"].width = 90


def _write_manager_dispatch_summary_sheet(
    writer: pd.ExcelWriter,
    route_df: pd.DataFrame,
    summary_df: pd.DataFrame,
) -> None:
    ws = writer.book.create_sheet("Manager Dispatch Summary")
    _write_title(ws, "Manager Dispatch Summary", 1)
    ws.cell(
        row=2,
        column=1,
        value="One-row dispatch view for managers. Use Manager Notes before sending routes to riders.",
    )
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    headers = [
        "Rider",
        "Cars Assigned",
        "Start Location",
        "First Pickup",
        "Final Drop-off",
        "Total Estimated Duration Min",
        "Total Estimated Distance KM",
        "Route Validation Status",
        "Cost Source Summary",
        "Manager Notes",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=header)
    _style_header_row(ws, 4, len(headers))
    ws.freeze_panes = "A5"

    summary_by_rider = summary_df.set_index("Rider") if not summary_df.empty else pd.DataFrame()
    row_idx = 5
    if route_df.empty:
        ws.cell(row=row_idx, column=1, value="No rider routes were assigned.")
    else:
        for rider, rider_routes in route_df.sort_values(["Rider", "Sequence"]).groupby("Rider", sort=False):
            rider_summary = summary_by_rider.loc[rider] if rider in summary_by_rider.index else None
            first_route = rider_routes.sort_values("Sequence").iloc[0]
            final_route = rider_routes.sort_values("Sequence").iloc[-1]
            total_duration = float(rider_routes["Total Duration Min"].fillna(0).sum())
            total_distance = float(rider_routes["Total Distance KM"].fillna(0).sum())
            final_dropoff = clean_text(final_route["Drop-off Address"])
            if rider_summary is not None:
                total_duration = float(rider_summary.get("Total Route Duration Min", total_duration) or 0)
                total_distance = float(rider_summary.get("Total Route Distance KM", total_distance) or 0)
                final_dropoff = clean_text(rider_summary.get("Final Location", final_dropoff))

            values = [
                rider,
                len(rider_routes),
                clean_text(first_route["Start From"]),
                clean_text(first_route["Pickup Address"]),
                final_dropoff,
                round(total_duration, 1),
                round(total_distance, 2),
                _route_validation_summary(rider_routes),
                _cost_source_summary(rider_routes),
                _manager_note(rider_routes),
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_idx += 1

    note_col = headers.index("Manager Notes") + 1
    for row in range(5, ws.max_row + 1):
        note = str(ws.cell(row, note_col).value or "")
        if "Ready" in note:
            fill = PatternFill("solid", fgColor="C6EFCE")
        elif "Review" in note:
            fill = PatternFill("solid", fgColor="FFE7A3")
        else:
            fill = PatternFill("solid", fgColor="F4B183")
        ws.cell(row, note_col).fill = fill
    _autosize_columns(ws, 45)


def _write_unassigned_jobs_sheet(
    writer: pd.ExcelWriter,
    jobs_df: pd.DataFrame | None,
    route_df: pd.DataFrame,
) -> None:
    ws = writer.book.create_sheet("Unassigned Jobs")
    _write_title(ws, "Unassigned Jobs", 1)
    ws.cell(
        row=2,
        column=1,
        value="Uploaded jobs that were not assigned to any rider. Use the Uploaded Row number to find the job in the original file.",
    )
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    unassigned_df = build_unassigned_jobs_df(jobs_df, route_df)
    if unassigned_df.empty:
        ws.cell(row=4, column=1, value="All uploaded valid jobs were assigned.")
        _autosize_columns(ws, 45)
        return

    header_row = 4
    for col, column_name in enumerate(unassigned_df.columns, start=1):
        ws.cell(row=header_row, column=col, value=column_name)
    _style_header_row(ws, header_row, len(unassigned_df.columns))

    for row_idx, row in enumerate(unassigned_df.itertuples(index=False), start=header_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A5"
    _autosize_columns(ws, 55)


def _write_rejected_candidate_audit_sheet(writer: pd.ExcelWriter, route_df: pd.DataFrame) -> None:
    ws = writer.book.create_sheet("Rejected Candidate Audit")
    _write_title(ws, "Rejected Candidate Audit", 1)
    ws.cell(
        row=2,
        column=1,
        value="Top rejected rider options for jobs that remained unassigned after normal and rescue assignment passes.",
    )
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    audit_df = pd.DataFrame(route_df.attrs.get("rejected_candidate_audit", []))
    if audit_df.empty:
        ws.cell(row=4, column=1, value="No rejected candidate audit rows for this run.")
        _autosize_columns(ws, 45)
        return

    header_row = 4
    for col, column_name in enumerate(audit_df.columns, start=1):
        ws.cell(row=header_row, column=col, value=column_name)
    _style_header_row(ws, header_row, len(audit_df.columns))
    for row_idx, row in enumerate(audit_df.itertuples(index=False), start=header_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A5"
    _autosize_columns(ws, 55)


def export_routes_to_excel(
    route_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    jobs_df: pd.DataFrame | None = None,
    validation_warnings: list[str] | None = None,
    lookup_warnings: list[str] | None = None,
) -> bytes:
    route_df = route_df.sort_values(["Rider", "Sequence"]).reset_index(drop=True) if not route_df.empty else route_df
    summary_df = format_summary_output(summary_df, route_df)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _write_how_to_read_sheet(writer)

        route_note = (
            "Follow each rider's rows in sequence order. For each row, the rider first travels from "
            "Start From to the Pickup Address without the car. After collecting the car, the rider "
            "moves it from Pickup Address to Drop-off Address. The next row starts from the previous "
            "drop-off location."
        )
        route_df.to_excel(writer, sheet_name="Optimised Routes", index=False, startrow=4)
        route_ws = writer.sheets["Optimised Routes"]
        _write_title(route_ws, "Optimised Routes", 1)
        route_ws.cell(row=2, column=1, value=route_note)
        route_ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        route_ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=min(len(ROUTE_COLUMNS), 8))
        if not route_df.empty:
            _style_route_sheet(route_ws, header_row=5, table_start_row=6, table_end_row=5 + len(route_df))
        _autosize_columns(route_ws, 45)

        summary_ws = writer.book.create_sheet("Summary")
        _write_title(summary_ws, "Overall Route Summary", 1)
        overall_rows = _overall_summary(route_df, summary_df, jobs_df)
        next_row = _write_rows(summary_ws, overall_rows, 3) + 2
        _write_title(summary_ws, "Rider Workload Summary", next_row)
        rider_header_row = next_row + 2
        for col, column_name in enumerate(SUMMARY_COLUMNS, start=1):
            summary_ws.cell(row=rider_header_row, column=col, value=column_name)
        for row_idx, row in enumerate(summary_df.itertuples(index=False), start=rider_header_row + 1):
            for col_idx, value in enumerate(row, start=1):
                summary_ws.cell(row=row_idx, column=col_idx, value=value)
        _style_summary_sheet(summary_ws, rider_header_row)
        _autosize_columns(summary_ws, 45)

        _write_unassigned_jobs_sheet(writer, jobs_df, route_df)
        _write_rejected_candidate_audit_sheet(writer, route_df)
        _write_rider_instructions_sheet(writer, route_df, summary_df)
        _write_manager_dispatch_summary_sheet(writer, route_df, summary_df)

        fallback_ws = writer.book.create_sheet("Fallback Cost Guide")
        _write_title(fallback_ws, "Fallback Cost Guide", 1)
        fallback_rows = [
            ["Important", "These are rough estimated costs used only when OneMap is unavailable."],
            ["Same-zone routes", "Same zone usually has lower estimated cost."],
            ["Different-zone routes", "Different or far zones have higher estimated cost."],
            ["Accuracy", "These are not exact driving times. For accurate results, OneMap should be enabled and working."],
        ]
        table_start = _write_rows(fallback_ws, fallback_rows, 3) + 2
        fallback_df = get_cost_explanation()
        for col, column_name in enumerate(fallback_df.columns, start=1):
            fallback_ws.cell(row=table_start, column=col, value=column_name)
        for row_idx, row in enumerate(fallback_df.itertuples(index=False), start=table_start + 1):
            for col_idx, value in enumerate(row, start=1):
                fallback_ws.cell(row=row_idx, column=col_idx, value=value)
        _style_header_row(fallback_ws, table_start, len(fallback_df.columns))
        _autosize_columns(fallback_ws, 45)

        quality_ws = writer.book.create_sheet("Data Quality Checks")
        _write_title(quality_ws, "Data Quality Checks", 1)
        quality_rows = _data_quality_rows(route_df, jobs_df, validation_warnings, lookup_warnings)
        _write_rows(quality_ws, quality_rows, 3)
        _style_header_row(quality_ws, 3, 2)
        _autosize_columns(quality_ws, 70)
    return output.getvalue()


def optimise_routes(
    jobs: pd.DataFrame,
    riders: list[RiderState],
    use_onemap: bool = True,
    optimise_by: str = "duration",
) -> tuple[pd.DataFrame, pd.DataFrame]:
    route_df, summary_df, _ = optimise_vehicle_routes(
        jobs,
        riders,
        use_onemap=use_onemap,
        optimise_by=optimise_by,
        token=get_onemap_token(),
    )
    return route_df, summary_df


def build_excel_download(route_df: pd.DataFrame, summary_df: pd.DataFrame) -> bytes:
    return export_routes_to_excel(route_df, summary_df)
